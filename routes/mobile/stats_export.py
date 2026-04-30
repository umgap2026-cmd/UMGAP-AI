from datetime import date, timedelta, datetime
from calendar import monthrange
from io import BytesIO
from collections import defaultdict
from flask import Blueprint, request, send_file
from decimal import Decimal

from flask import Blueprint, request, send_file
from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from db import get_conn
from core import mobile_api_response, mobile_api_login_required

mobile_stats_export_bp = Blueprint("mobile_stats_export", __name__)

MONTH_ID = ["","Januari","Februari","Maret","April","Mei","Juni",
            "Juli","Agustus","September","Oktober","November","Desember"]

C = {
    "navy":"0B1733","blue":"1565C0","blue_mid":"1E88E5",
    "blue_lite":"E3F2FD","cyan_lite":"F0F7FF","white":"FFFFFF",
    "grey_bg":"F5F5F5","grey_txt":"616161",
    "green":"2E7D32","green_bg":"E8F5E9",
    "yellow":"E65100","yellow_bg":"FFF8E1",
    "red":"C62828","red_bg":"FFEBEE",
    "purple":"6A1B9A","purple_bg":"F3E5F5",
}

STATUS_MAP = {
    "PRESENT": ("Tepat Waktu", "green",  "green_bg"),
    "LATE":    ("Terlambat",   "yellow", "yellow_bg"),
    "SICK":    ("Sakit",       "blue",   "blue_lite"),
    "LEAVE":   ("Izin",        "purple", "purple_bg"),
    "ABSENT":  ("Absen",       "red",    "red_bg"),
}


def _ft(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def _fill(k):
    return PatternFill("solid", fgColor=C.get(k, k))

def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _bdr(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _rp(v):
    try:
        n = int(v or 0)
        return f"Rp {n:,}".replace(",", ".")
    except Exception:
        return "Rp 0"

def _wib(dt):
    """UTC/naive datetime → WIB string HH:MM."""
    if not dt: return "-"
    try:
        # checkin_at disimpan sebagai WIB naive (lihat _now_wib_naive di core.py)
        return dt.strftime("%H:%M")
    except Exception:
        return "-"

def _fmtd(d):
    if not d: return "-"
    days = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]
    return f"{days[d.weekday()]}, {d.day} {MONTH_ID[d.month]} {d.year}"

def _resolve(params):
    df = params.get("date_from", "").strip()
    dt = params.get("date_to",   "").strip()
    mo = params.get("month",     "").strip()
    if df and dt:
        s = datetime.strptime(df, "%Y-%m-%d").date()
        e = datetime.strptime(dt, "%Y-%m-%d").date()
        lbl = (f"{s.day} {MONTH_ID[s.month]} {s.year} - "
               f"{e.day} {MONTH_ID[e.month]} {e.year}")
        return s, e + timedelta(days=1), lbl, f"UMGAP_Kehadiran_{df}_sd_{dt}.xlsx"
    if not mo:
        t = date.today(); mo = f"{t.year:04d}-{t.month:02d}"
    y, m = int(mo[:4]), int(mo[5:7])
    s = date(y, m, 1); e = date(y, m, monthrange(y, m)[1])
    return s, e + timedelta(days=1), f"{MONTH_ID[m]} {y}", \
           f"UMGAP_Kehadiran_{y}_{m:02d}.xlsx"


# ── Deteksi kolom gaji yang tersedia ────────────────────────────────
def _col_exists(cur, table, column):
    cur.execute("""
        SELECT 1 FROM information_schema.columns
        WHERE table_name=%s AND column_name=%s LIMIT 1;
    """, (table, column))
    return cur.fetchone() is not None

def _salary_exprs(cur):
    has_ps      = _col_exists(cur, "payroll_settings", "daily_salary")
    has_monthly = has_ps and _col_exists(cur, "payroll_settings", "monthly_salary")
    has_saltype = has_ps and _col_exists(cur, "payroll_settings", "salary_type")
    if has_ps:
        de  = "COALESCE(ps.daily_salary, 0)"
        me  = "COALESCE(ps.monthly_salary, 0)" if has_monthly else "0"
        se  = "COALESCE(ps.salary_type, 'daily')" if has_saltype else "'daily'"
        je  = "LEFT JOIN payroll_settings ps ON ps.user_id = u.id"
        ge  = (", ps.daily_salary"
               + (", ps.monthly_salary" if has_monthly else "")
               + (", ps.salary_type"    if has_saltype else ""))
    else:
        de = me = "0"; se = "'daily'"; je = ""; ge = ""
    return de, me, se, je, ge


# ── Sheet 1: satu baris per record absensi ───────────────────────────
def _fetch_detail(s, e):
    conn = get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        de, me, se, je, _ = _salary_exprs(cur)
        # attendance table: work_date, status, arrival_type, checkin_at, note
        # TIDAK ada checkin_time / checkout_time di schema ini
        cur.execute(f"""
            SELECT
                u.name          AS employee_name,
                {de}            AS daily_salary,
                {me}            AS monthly_salary,
                {se}            AS salary_type,
                a.work_date,
                a.status,
                a.arrival_type,
                a.checkin_at,   -- WIB naive (dari _now_wib_naive)
                a.note
            FROM attendance a
            JOIN users u ON u.id = a.user_id
            {je}
            WHERE u.role = 'employee'
              AND a.work_date >= %s AND a.work_date < %s
            ORDER BY u.name ASC, a.work_date ASC;
        """, (s, e))
        return [dict(r) for r in cur.fetchall()]
    except Exception as ex:
        import traceback
        print(f"[export._fetch_detail] {ex}\n{traceback.format_exc()}")
        raise
    finally:
        cur.close(); conn.close()


# ── Sheet 2: ringkasan per karyawan ─────────────────────────────────
def _fetch_summary(s, e):
    conn = get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        de, me, se, je, ge = _salary_exprs(cur)
        cur.execute(f"""
            SELECT
                u.id,
                u.name          AS employee_name,
                {de}            AS daily_salary,
                {me}            AS monthly_salary,
                {se}            AS salary_type,
                COALESCE(SUM(CASE WHEN a.status='PRESENT' THEN 1 ELSE 0 END),0) AS present_days,
                COALESCE(SUM(CASE WHEN a.arrival_type='LATE' THEN 1 ELSE 0 END),0) AS late_days,
                COALESCE(SUM(CASE WHEN a.status='SICK'    THEN 1 ELSE 0 END),0) AS sick_days,
                COALESCE(SUM(CASE WHEN a.status='LEAVE'   THEN 1 ELSE 0 END),0) AS leave_days,
                COALESCE(SUM(CASE WHEN a.status='ABSENT'  THEN 1 ELSE 0 END),0) AS absent_days
            FROM users u
            {je}
            LEFT JOIN attendance a ON a.user_id = u.id
                AND a.work_date >= %s AND a.work_date < %s
            WHERE u.role = 'employee'
            GROUP BY u.id, u.name{ge}
            ORDER BY u.name ASC;
        """, (s, e))
        return [dict(r) for r in cur.fetchall()]
    except Exception as ex:
        import traceback
        print(f"[export._fetch_summary] {ex}\n{traceback.format_exc()}")
        raise
    finally:
        cur.close(); conn.close()


# ── Build Excel ──────────────────────────────────────────────────────
def _title(ws, r1, r2, title, sub):
    ws.merge_cells(r1); c = ws[r1.split(":")[0]]
    c.value = title; c.font = _ft(True, 13, "FFFFFF")
    c.fill = _fill("navy"); c.alignment = _aln("center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(r2); c2 = ws[r2.split(":")[0]]
    c2.value = sub; c2.font = _ft(size=9, color="FFFFFF")
    c2.fill = _fill("blue_mid"); c2.alignment = _aln("center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

def _hdr(ws, row, cols):
    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _ft(True, 10, "FFFFFF"); c.fill = _fill("blue")
        c.alignment = _aln("center"); c.border = _bdr("1565C0")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22

def _build(detail, summary, label):
    wb  = openpyxl.Workbook()
    gen = (datetime.utcnow() + timedelta(hours=7)).strftime("%d %b %Y %H:%M WIB")
    sub = f"Periode: {label}   |   Dicetak: {gen}"

    # ════ SHEET 1: DETAIL KEHADIRAN ════
    ws1 = wb.active; ws1.title = "Detail Kehadiran"
    ws1.sheet_view.showGridLines = False; ws1.freeze_panes = "A5"
    _title(ws1, "A1:J1", "A2:J2",
           "LAPORAN DETAIL KEHADIRAN — UMGAP", sub)

    # 10 kolom (tanpa checkout karena tidak ada di schema)
    C1 = [("No",4), ("Tanggal",22), ("Nama Karyawan",24), ("Status",14),
          ("Jam Masuk",12), ("Catatan",22),
          ("Gaji Harian",16), ("Jenis",10), ("Dihitung",10), ("Total Gaji",18)]
    _hdr(ws1, 4, C1)

    DR = 5
    for i, r in enumerate(detail, 1):
        row  = DR + i - 1
        even = "cyan_lite" if i % 2 == 0 else "white"
        rs   = (r.get("status")       or "").upper()
        at   = (r.get("arrival_type") or "").upper()
        dk   = "LATE" if at == "LATE" and rs == "PRESENT" else rs
        ls, fc, bg = STATUS_MAP.get(dk, (dk, "grey_txt", "white"))

        daily   = int(r["daily_salary"]   or 0)
        monthly = int(r["monthly_salary"] or 0)
        st      = r["salary_type"] or "daily"
        counted = rs in ("PRESENT", "SICK", "LEAVE")
        dh      = 1 if counted else 0
        tg      = monthly if st == "monthly" else (daily if counted else 0)

        rd = [
            (i,                        "center", "grey_txt", even),
            (_fmtd(r["work_date"]),    "left",   "000000",   even),
            (r["employee_name"],       "left",   "000000",   even),
            (ls,                       "center", fc,         bg),
            (_wib(r.get("checkin_at")),"center", "000000",   even),
            (r.get("note") or "",      "left",   "grey_txt", even),
            (_rp(daily) if st != "monthly" else "-",
                                       "right",  "000000",   even),
            ("Bulanan" if st == "monthly" else "Harian",
                                       "center", "grey_txt", even),
            (dh,                       "center", "blue",     even),
            (_rp(tg),                  "right",  "blue",     even),
        ]
        for col, (val, ha, fc2, bg2) in enumerate(rd, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = _ft(size=9, color=C.get(fc2, fc2), bold=(col == 10))
            c.fill = _fill(bg2); c.alignment = _aln(ha); c.border = _bdr()
        ws1.row_dimensions[row].height = 17

    last1 = (DR + len(detail) - 1) if detail else DR
    tr    = last1 + 2
    ws1.merge_cells(f"A{tr}:F{tr}")
    c = ws1.cell(row=tr, column=1, value="TOTAL HARI TERHITUNG")
    c.font = _ft(True, 10, "FFFFFF"); c.fill = _fill("navy")
    c.alignment = _aln("right")
    for col in [7, 9, 10]:
        cl  = get_column_letter(col)
        val = f"=SUM({cl}{DR}:{cl}{last1})" if detail else 0
        ce  = ws1.cell(row=tr, column=col, value=val)
        ce.font = _ft(True, 10, "FFFFFF"); ce.fill = _fill("navy")
        ce.alignment = _aln("center")
    ws1.row_dimensions[tr].height = 22

    lr = tr + 2; ws1.merge_cells(f"A{lr}:J{lr}")
    ws1.cell(row=lr, column=1,
             value="* Hijau=Tepat Waktu | Kuning=Terlambat | Merah=Absen | Biru=Sakit | Ungu=Izin"
             ).font = _ft(size=8, color=C["grey_txt"])

    # ════ SHEET 2: RINGKASAN PER KARYAWAN ════
    ws2 = wb.create_sheet("Ringkasan")
    ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A5"
    _title(ws2, "A1:L1", "A2:L2",
           "RINGKASAN PENGGAJIAN — UMGAP", sub)

    C2 = [("No",4), ("Nama Karyawan",26), ("Jenis Gaji",12), ("Gaji Harian",16),
          ("Tepat Waktu",12), ("Terlambat",11), ("Sakit",9), ("Izin",9),
          ("Absen",9), ("Total Hadir",11), ("Kehadiran %",12), ("Total Gaji",18)]
    _hdr(ws2, 4, C2)

    grand = defaultdict(int); SR = 5
    for i, r in enumerate(summary, 1):
        row  = SR + i - 1
        even = "cyan_lite" if i % 2 == 0 else "white"
        daily   = int(r["daily_salary"]   or 0)
        monthly = int(r["monthly_salary"] or 0)
        st      = r["salary_type"] or "daily"
        p  = int(r["present_days"] or 0); la = int(r["late_days"]  or 0)
        si = int(r["sick_days"]    or 0); lv = int(r["leave_days"] or 0)
        ab = int(r["absent_days"]  or 0)
        wk = p + la + si + lv; td = wk + ab
        pct  = round(wk / td * 100, 1) if td > 0 else 0.0
        gaji = monthly if st == "monthly" else wk * daily
        grand["p"]  += p;  grand["la"] += la; grand["si"] += si
        grand["lv"] += lv; grand["ab"] += ab; grand["wk"] += wk
        grand["gaji"] += gaji
        pfc = "green" if pct >= 80 else ("yellow" if pct >= 60 else "red")
        rd2 = [
            (i,          "center", "grey_txt"),
            (r["employee_name"], "left", "000000"),
            ("Bulanan" if st=="monthly" else "Harian", "center", "grey_txt"),
            ("-" if st=="monthly" else _rp(daily), "right", "000000"),
            (p,  "center", "green"),  (la, "center", "yellow"),
            (si, "center", "blue"),   (lv, "center", "purple"),
            (ab, "center", "red"),    (wk, "center", "blue_mid"),
            (f"{pct}%", "center", pfc), (_rp(gaji), "right", "blue"),
        ]
        for col, (val, ha, fc2) in enumerate(rd2, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = _ft(size=9, color=C.get(fc2, fc2), bold=(col == 12))
            c.fill = _fill(even); c.alignment = _aln(ha); c.border = _bdr()
        ws2.row_dimensions[row].height = 17

    last2 = (SR + len(summary) - 1) if summary else SR
    gr    = last2 + 2
    ws2.merge_cells(f"A{gr}:C{gr}")
    c = ws2.cell(row=gr, column=1, value="GRAND TOTAL")
    c.font = _ft(True, 11, "FFFFFF"); c.fill = _fill("navy")
    c.alignment = _aln("center")
    for col, val in [(4,"-"), (5,grand["p"]), (6,grand["la"]), (7,grand["si"]),
                     (8,grand["lv"]), (9,grand["ab"]), (10,grand["wk"]),
                     (11,""), (12,_rp(grand["gaji"]))]:
        c = ws2.cell(row=gr, column=col, value=val)
        c.font = _ft(True, 10, "FFFFFF"); c.fill = _fill("navy")
        c.alignment = _aln("right" if col == 12 else "center")
    ws2.row_dimensions[gr].height = 24

    # Statistik ringkasan
    te  = len(summary)
    td2 = grand["wk"] + grand["ab"]
    avg = round(grand["wk"] / td2 * 100, 1) if td2 > 0 else 0
    sr2 = gr + 2
    ws2.merge_cells(f"A{sr2}:D{sr2}")
    h = ws2.cell(row=sr2, column=1, value="STATISTIK PERIODE")
    h.font = _ft(True, 10, "FFFFFF"); h.fill = _fill("blue")
    h.alignment = _aln("center"); ws2.row_dimensions[sr2].height = 20

    for j, (lbl, val, fc2) in enumerate([
        ("Total Karyawan",        te,               "blue"),
        ("Total Hari Hadir",      grand["p"],        "green"),
        ("Total Hari Terlambat",  grand["la"],       "yellow"),
        ("Total Hari Sakit",      grand["si"],       "blue_mid"),
        ("Total Hari Izin",       grand["lv"],       "purple"),
        ("Total Hari Absen",      grand["ab"],       "red"),
        ("Rata-rata Kehadiran %", f"{avg}%",         "blue"),
        ("Total Estimasi Gaji",   _rp(grand["gaji"]),"blue"),
    ]):
        rr = sr2 + 1 + j; ev = "grey_bg" if j % 2 == 0 else "white"
        ws2.merge_cells(f"A{rr}:C{rr}")
        lc = ws2.cell(row=rr, column=1, value=lbl)
        lc.font = _ft(size=9); lc.fill = _fill(ev)
        lc.alignment = _aln("left"); lc.border = _bdr()
        vc = ws2.cell(row=rr, column=4, value=val)
        vc.font = _ft(True, 10, C.get(fc2, fc2))
        vc.fill = _fill(ev); vc.alignment = _aln("center"); vc.border = _bdr()
        ws2.row_dimensions[rr].height = 17

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── Endpoint ─────────────────────────────────────────────────────────
@mobile_stats_export_bp.route("/stats/export", methods=["GET", "OPTIONS"])
@mobile_api_login_required
def mobile_stats_export():
    if request.method == "OPTIONS":
        return mobile_api_response(ok=True, message="OK", data={}, status_code=200)
    role = str(request.mobile_user.get("role") or "").strip().lower()
    if role not in ("admin", "owner"):
        return mobile_api_response(ok=False, message="Akses ditolak.", status_code=403)
    if not HAS_OPENPYXL:
        return mobile_api_response(
            ok=False,
            message="openpyxl belum terinstall. Tambahkan openpyxl>=3.1.2 ke requirements.txt",
            status_code=500)
    try:
        start, end_exc, label, filename = _resolve(request.args)
    except Exception as e:
        return mobile_api_response(
            ok=False, message=f"Format tanggal tidak valid: {e}", status_code=400)
    try:
        detail = _fetch_detail(start, end_exc)
        summ   = _fetch_summary(start, end_exc)
        buf    = _build(detail, summ, label)
    except Exception as e:
        import traceback
        print(f"[stats_export] ERROR: {e}\n{traceback.format_exc()}")
        return mobile_api_response(
            ok=False, message=f"Gagal membuat Excel: {e}", status_code=500)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@mobile_stats_export_bp.route("/owner/export/excel", methods=["GET", "OPTIONS"])
@mobile_api_login_required
def owner_export_excel():
    """Export laporan keuangan + gaji owner ke Excel"""
    if request.method == "OPTIONS":
        from core import mobile_api_response
        return mobile_api_response(ok=True, message="OK", data={})

    role = str(request.mobile_user.get("role") or "").strip().lower()
    if role not in ("owner", "admin"):
        from core import mobile_api_response
        return mobile_api_response(ok=False, message="Akses ditolak.", status_code=403)

    month = request.args.get("month", "")
    if not month:
        today = date.today()
        month = f"{today.year:04d}-{today.month:02d}"

    year = int(month.split("-")[0])
    mon  = int(month.split("-")[1])
    start = date(year, mon, 1)
    end   = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)

    def _c(v):
        if v is None: return 0.0
        if isinstance(v, Decimal): return float(v)
        try: return float(v)
        except: return 0.0

    def _rp(v):
        n = int(_c(v))
        if n == 0: return "Rp -"
        return f"Rp {n:,}".replace(",", ".")

    conn = get_conn()
    cur  = conn.cursor(cursor_factory=RealDictCursor)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────────────
        navy_fill   = PatternFill("solid", fgColor="0D1B4E")
        teal_fill   = PatternFill("solid", fgColor="0E7490")
        green_fill  = PatternFill("solid", fgColor="166534")
        amber_fill  = PatternFill("solid", fgColor="92400E")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font  = Font(bold=True, color="0D1B4E", size=13)
        sub_font    = Font(bold=True, color="374151", size=10)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        def _hdr(ws, row, col, text, fill=navy_fill):
            c = ws.cell(row=row, column=col, value=text)
            c.font  = header_font
            c.fill  = fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
            return c

        def _cell(ws, row, col, val, bold=False, align="left", color=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font   = Font(bold=bold, color=color or "111827", size=10)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = thin_border
            return c

        # ════════════════════════════════════════
        # SHEET 1: RINGKASAN KEUANGAN
        # ════════════════════════════════════════
        names = ['','Januari','Februari','Maret','April','Mei','Juni',
                 'Juli','Agustus','September','Oktober','November','Desember']
        period_label = f"{names[mon]} {year}"

        ws1 = wb.active
        ws1.title = "Keuangan"
        ws1.column_dimensions['A'].width = 32
        ws1.column_dimensions['B'].width = 22

        ws1.merge_cells('A1:B1')
        t = ws1['A1']
        t.value = f"LAPORAN KEUANGAN UMGAP — {period_label.upper()}"
        t.font  = Font(bold=True, color="0D1B4E", size=14)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 32

        ws1.merge_cells('A2:B2')
        ws1['A2'].value = f"Dicetak: {date.today().strftime('%d/%m/%Y')}"
        ws1['A2'].font  = Font(color="6B7280", size=9)
        ws1['A2'].alignment = Alignment(horizontal="center")

        row = 4
        _hdr(ws1, row, 1, "Keterangan", navy_fill)
        _hdr(ws1, row, 2, "Jumlah", navy_fill)
        row += 1

        # Ambil data keuangan
        revenue = buying = expense = 0.0
        if True:
            try:
                cur.execute("""
                    SELECT
                        COALESCE(SUM(CASE WHEN type='JUAL_GUDANG' THEN total_amount ELSE 0 END),0) AS rev,
                        COALESCE(SUM(CASE WHEN type='BELI_GUDANG' THEN total_amount ELSE 0 END),0) AS buy,
                        COALESCE(SUM(CASE WHEN type='PENGELUARAN' THEN total_amount ELSE 0 END),0) AS exp
                    FROM fin_transactions WHERE created_at::date >= %s AND created_at::date < %s;
                """, (start, end))
                r = cur.fetchone() or {}
                revenue += _c(r.get("rev")); buying += _c(r.get("buy")); expense += _c(r.get("exp"))
            except: pass
            try:
                cur.execute("""
                    SELECT
                        COALESCE(SUM(CASE WHEN type='JUAL' THEN subtotal ELSE 0 END),0) AS rev,
                        COALESCE(SUM(CASE WHEN type='BELI' THEN subtotal ELSE 0 END),0) AS buy,
                        COALESCE(SUM(CASE WHEN type='EXPENSE' THEN subtotal ELSE 0 END),0) AS exp
                    FROM fin_trip_items WHERE created_at::date >= %s AND created_at::date < %s;
                """, (start, end))
                r = cur.fetchone() or {}
                revenue += _c(r.get("rev")); buying += _c(r.get("buy")); expense += _c(r.get("exp"))
            except: pass

        gross = revenue - buying
        net_pre = gross - expense

        items_keu = [
            ("PENDAPATAN", None),
            ("  Omzet Penjualan (Gudang + Jakarta)", revenue),
            ("PENGELUARAN", None),
            ("  Modal Beli Barang", buying),
            ("  Biaya Operasional", expense),
            ("LABA KOTOR", gross),
            ("LABA SEBELUM GAJI", net_pre),
        ]

        for label, val in items_keu:
            is_total = label.startswith("LABA")
            is_cat   = val is None
            _cell(ws1, row, 1, label, bold=is_total or is_cat,
                  color="0D1B4E" if is_total else ("374151" if is_cat else "111827"))
            if val is not None:
                c = _cell(ws1, row, 2, _rp(val), bold=is_total, align="right",
                           color="166534" if (is_total and val >= 0) else
                                 ("DC2626" if (is_total and val < 0) else "111827"))
            else:
                ws1.cell(row=row, column=2).fill = PatternFill("solid", fgColor="F3F4F6")
            if is_total:
                for col in [1, 2]:
                    ws1.cell(row=row, column=col).fill = PatternFill("solid", fgColor="ECFDF5")
            row += 1

        # Gaji
        salary_total = 0
        row += 1
        _hdr(ws1, row, 1, "GAJI KARYAWAN", teal_fill)
        _hdr(ws1, row, 2, "Estimasi", teal_fill)
        row += 1
        try:
            cur.execute("""
                SELECT u.name,
                    COALESCE(ps.daily_salary, 0) AS daily,
                    COALESCE(ps.monthly_salary, 0) AS monthly,
                    COALESCE(SUM(CASE WHEN a.status='PRESENT' THEN 1 ELSE 0 END),0) AS prs,
                    COALESCE(SUM(CASE WHEN a.arrival_type='LATE' THEN 1 ELSE 0 END),0) AS late,
                    COALESCE(SUM(CASE WHEN a.status='SICK' THEN 1 ELSE 0 END),0) AS sick
                FROM users u
                LEFT JOIN payroll_settings ps ON ps.user_id = u.id
                LEFT JOIN attendance a ON a.user_id = u.id
                    AND a.work_date >= %s AND a.work_date < %s
                WHERE u.role = 'employee'
                GROUP BY u.id, u.name, ps.daily_salary, ps.monthly_salary
                ORDER BY u.name;
            """, (start, end))
            for r in cur.fetchall():
                daily   = int(_c(r["daily"])); monthly = int(_c(r["monthly"]))
                worked  = int(r["prs"] or 0) + int(r["late"] or 0) + int(r["sick"] or 0)
                gaji    = monthly if monthly > 0 else worked * daily
                salary_total += gaji
                _cell(ws1, row, 1, r["name"])
                _cell(ws1, row, 2, _rp(gaji), align="right")
                row += 1
        except: pass

        # Total gaji + laba bersih
        net_final = net_pre - salary_total
        _cell(ws1, row, 1, "TOTAL GAJI", bold=True, color="92400E")
        _cell(ws1, row, 2, _rp(salary_total), bold=True, align="right", color="92400E")
        for col in [1,2]:
            ws1.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFBEB")
        row += 1
        _cell(ws1, row, 1, "LABA BERSIH (setelah gaji)", bold=True,
              color="166534" if net_final >= 0 else "DC2626")
        _cell(ws1, row, 2, _rp(net_final), bold=True, align="right",
              color="166534" if net_final >= 0 else "DC2626")
        for col in [1,2]:
            ws1.cell(row=row, column=col).fill = PatternFill(
                "solid", fgColor="ECFDF5" if net_final >= 0 else "FEF2F2")

        ws1.freeze_panes = "A5"

        # ════════════════════════════════════════
        # SHEET 2: TOP MATERIAL
        # ════════════════════════════════════════
        ws2 = wb.create_sheet("Top Material")
        for col, (w, h) in enumerate(zip([5,22,16,16,10], []), 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

        ws2['A1'].value = f"TOP MATERIAL TERJUAL — {period_label}"
        ws2['A1'].font  = title_font
        ws2.merge_cells('A1:E1')
        ws2.row_dimensions[1].height = 24

        hdrs = ["#", "Material", "Total Kg", "Nilai (Rp)", "Transaksi"]
        fills = [navy_fill]*5
        for i, (h, f) in enumerate(zip(hdrs, fills), 1):
            _hdr(ws2, 2, i, h, f)

        try:
            cur.execute("""
                SELECT m.name,
                    COALESCE(SUM(i.qty_kg),0) AS kg,
                    COALESCE(SUM(i.subtotal),0) AS nilai,
                    COUNT(*) AS txn
                FROM fin_transaction_items i
                JOIN fin_materials m ON m.id = i.material_id
                JOIN fin_transactions t ON t.id = i.transaction_id
                WHERE t.type='JUAL_GUDANG'
                  AND t.created_at::date >= %s AND t.created_at::date < %s
                  AND i.material_id IS NOT NULL
                GROUP BY m.name ORDER BY nilai DESC LIMIT 10;
            """, (start, end))
            for rank, r in enumerate(cur.fetchall(), 1):
                row2 = rank + 2
                _cell(ws2, row2, 1, rank, align="center", bold=(rank<=3))
                _cell(ws2, row2, 2, r["name"], bold=(rank<=3))
                _cell(ws2, row2, 3, f"{_c(r['kg']):.1f}", align="right")
                _cell(ws2, row2, 4, _rp(r["nilai"]), align="right",
                      color="166534" if rank == 1 else "111827", bold=(rank==1))
                _cell(ws2, row2, 5, int(r["txn"] or 0), align="center")
                if rank <= 3:
                    for col in range(1, 6):
                        ws2.cell(row=row2, column=col).fill = \
                            PatternFill("solid", fgColor="F0FDF4")
        except: pass

        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 22
        ws2.column_dimensions['C'].width = 14
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 12

        # ════════════════════════════════════════
        # SHEET 3: ABSENSI + GAJI
        # ════════════════════════════════════════
        ws3 = wb.create_sheet("SDM & Gaji")
        ws3['A1'].value = f"REKAP SDM & GAJI — {period_label}"
        ws3['A1'].font  = title_font
        ws3.merge_cells('A1:H1')
        ws3.row_dimensions[1].height = 24

        hdrs3 = ["Nama", "Hadir", "Terlambat", "Sakit", "Izin", "Absen", "Hari Kerja", "Est. Gaji"]
        for i, h in enumerate(hdrs3, 1):
            _hdr(ws3, 2, i, h, teal_fill)

        try:
            cur.execute("""
                SELECT u.name,
                    COALESCE(ps.daily_salary, 0) AS daily,
                    COALESCE(ps.monthly_salary, 0) AS monthly,
                    COALESCE(SUM(CASE WHEN a.status='PRESENT' THEN 1 ELSE 0 END),0) AS prs,
                    COALESCE(SUM(CASE WHEN a.arrival_type='LATE' THEN 1 ELSE 0 END),0) AS late,
                    COALESCE(SUM(CASE WHEN a.status='SICK' THEN 1 ELSE 0 END),0) AS sick,
                    COALESCE(SUM(CASE WHEN a.status='LEAVE' THEN 1 ELSE 0 END),0) AS leave,
                    COALESCE(SUM(CASE WHEN a.status='ABSENT' THEN 1 ELSE 0 END),0) AS absent
                FROM users u
                LEFT JOIN payroll_settings ps ON ps.user_id = u.id
                LEFT JOIN attendance a ON a.user_id = u.id
                    AND a.work_date >= %s AND a.work_date < %s
                WHERE u.role = 'employee'
                GROUP BY u.id, u.name, ps.daily_salary, ps.monthly_salary
                ORDER BY u.name;
            """, (start, end))
            for ri, r in enumerate(cur.fetchall(), 3):
                daily = int(_c(r["daily"])); monthly = int(_c(r["monthly"]))
                prs   = int(r["prs"] or 0);  late = int(r["late"] or 0)
                sick  = int(r["sick"] or 0); leave = int(r["leave"] or 0)
                absent = int(r["absent"] or 0)
                worked = prs + late + sick + leave
                gaji   = monthly if monthly > 0 else worked * daily
                row_data = [r["name"], prs, late, sick, leave, absent, worked, _rp(gaji)]
                for ci, val in enumerate(row_data, 1):
                    _cell(ws3, ri, ci, val,
                          align="right" if ci > 1 else "left",
                          bold=(ci == 8))
                    if ri % 2 == 0:
                        ws3.cell(row=ri, column=ci).fill = \
                            PatternFill("solid", fgColor="F9FAFB")
        except: pass

        for ci, w in enumerate([22,8,10,8,8,8,10,16], 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

        # ════════════════════════════════════════
        # SHEET 4: TRIP JAKARTA
        # ════════════════════════════════════════
        ws4 = wb.create_sheet("Perjalanan Jakarta")
        ws4['A1'].value = f"PERJALANAN JAKARTA — {period_label}"
        ws4['A1'].font  = title_font
        ws4.merge_cells('A1:F1')
        ws4.row_dimensions[1].height = 24

        hdrs4 = ["Tanggal", "Keterangan", "Status", "Pemasukan", "Pengeluaran", "Hasil Bersih"]
        for i, h in enumerate(hdrs4, 1):
            _hdr(ws4, 2, i, h, navy_fill)

        try:
            cur.execute("""
                SELECT trip_date, note, status,
                       total_income, total_expense, net_result
                FROM fin_trips
                WHERE created_at::date >= %s AND created_at::date < %s
                ORDER BY trip_date;
            """, (start, end))
            for ri, r in enumerate(cur.fetchall(), 3):
                net = _c(r["net_result"])
                row4 = [str(r["trip_date"]), r["note"] or "-", r["status"],
                        _rp(r["total_income"]), _rp(r["total_expense"]), _rp(net)]
                for ci, val in enumerate(row4, 1):
                    _cell(ws4, ri, ci, val,
                          color="166534" if (ci == 6 and net >= 0) else
                                "DC2626" if (ci == 6 and net < 0) else "111827",
                          bold=(ci == 6))
        except: pass

        for ci, w in enumerate([12,24,12,16,16,16], 1):
            ws4.column_dimensions[get_column_letter(ci)].width = w

        # ── Save ──────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f"UMGAP_Laporan_{month}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        import traceback
        print(f"[owner_export] ERROR: {e}\n{traceback.format_exc()}")
        from core import mobile_api_response
        return mobile_api_response(ok=False, message=str(e), status_code=500)
    finally:
        cur.close(); conn.close()
