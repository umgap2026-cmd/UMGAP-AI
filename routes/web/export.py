import io
import calendar
from datetime import date, timedelta

from flask import Blueprint, request, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from psycopg2.extras import RealDictCursor

from db import get_conn
from core import admin_required, ensure_hr_v2_schema, _parse_date, get_payroll_report

export_bp = Blueprint("export", __name__)


# =========================
# HELPERS
# =========================
THIN = Side(style="thin", color="D9DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_SUBHEADER = PatternFill("solid", fgColor="EEF4FA")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")

FONT_TITLE = Font(color="FFFFFF", bold=True, size=13)
FONT_HEADER = Font(bold=True)
FONT_BOLD = Font(bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _parse_user_ids(args):
    """Terima "user_ids" (koma-pisah, dari checkbox multi-pilih) ATAU
    "user_id" tunggal (kompatibel dgn link lama /range_user.xlsx).
    Kosong/None = semua karyawan (tidak difilter) -- termasuk yg tidak
    digaji tapi tetap tercatat absensinya terpisah."""
    ids_str = (args.get("user_ids") or "").strip()
    if ids_str:
        ids = []
        for part in ids_str.split(","):
            part = part.strip()
            if part.isdigit():
                ids.append(int(part))
        return ids or None
    single = (args.get("user_id") or "").strip()
    if single.isdigit():
        return [int(single)]
    return None


def rupiah_excel(value):
    try:
        return int(round(float(value or 0)))
    except Exception:
        return 0


def auto_fit(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        try:
            col_letter = get_column_letter(col_cells[0].column)
        except Exception:
            continue

        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)

        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def style_table_range(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER


# =========================
# EXPORT EXCEL
# =========================
@export_bp.route("/admin/data/range.xlsx")
@export_bp.route("/admin/data/range_user.xlsx")
def export_range():
    deny = admin_required()
    if deny:
        return deny

    ensure_hr_v2_schema()

    # Terima start_date/end_date (nama resmi) ATAU start/end (dipakai UI
    # admin_dashboard.html) -- sebelumnya mismatch nama ini bikin filter
    # tanggal dari UI tidak pernah benar-benar kepakai.
    start_str = (request.args.get("start_date") or request.args.get("start") or "").strip()
    end_str = (request.args.get("end_date") or request.args.get("end") or "").strip()
    user_ids = _parse_user_ids(request.args)

    start_date = _parse_date(start_str)
    end_date = _parse_date(end_str)

    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1)
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        # Ambil user employee (atau cuma yg dipilih kalau user_ids diisi --
        # sengaja bisa lebih dari 1, mis. buat pisahin yg tidak digaji tapi
        # tetap perlu direkap absensinya sendiri)
        cur.execute("""
            SELECT
                u.id,
                u.name,
                COALESCE(p.daily_salary, 0) AS daily_salary,
                COALESCE(p.monthly_salary, 0) AS monthly_salary
            FROM users u
            LEFT JOIN payroll_settings p ON p.user_id = u.id
            WHERE u.role = 'employee'
              AND (%s::int[] IS NULL OR u.id = ANY(%s::int[]))
            ORDER BY u.name ASC;
        """, (user_ids, user_ids))
        employees = cur.fetchall()

        # Ambil absensi rentang tanggal (termasuk jam pulang/checkout_at)
        cur.execute("""
            SELECT
                a.user_id,
                a.work_date,
                a.status,
                a.arrival_type,
                a.note,
                a.checkin_at,
                a.checkout_at
            FROM attendance a
            WHERE a.work_date >= %s
              AND a.work_date <= %s
            ORDER BY a.user_id ASC, a.work_date ASC;
        """, (start_date, end_date))
        attendance_rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    # Mapping absensi
    attendance_map = {}
    for row in attendance_rows:
        key = (row["user_id"], row["work_date"])
        attendance_map[key] = row

    # Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Absensi"

    # =========================
    # SHEET 1 - ABSENSI
    # =========================
    current_row = 1

    title = f"LAPORAN ABSENSI KARYAWAN ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})"
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    c = ws1.cell(current_row, 1, title)
    c.fill = FILL_TITLE
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER
    current_row += 2

    all_dates = []
    d = start_date
    while d <= end_date:
        all_dates.append(d)
        d += timedelta(days=1)

    for idx, emp in enumerate(employees, start=1):
        emp_name = emp["name"]

        # Header user
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        c = ws1.cell(current_row, 1, f"{idx}. {emp_name}")
        c.fill = FILL_SUBHEADER
        c.font = Font(bold=True, size=12)
        c.alignment = ALIGN_LEFT
        c.border = BORDER
        current_row += 1

        headers = ["No", "Tanggal", "Hari", "Status", "Tipe Datang", "Jam Masuk", "Jam Pulang", "Catatan"]
        for col, h in enumerate(headers, start=1):
            cell = ws1.cell(current_row, col, h)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER
        current_row += 1

        start_data_row = current_row

        present_count = 0
        sick_count = 0
        leave_count = 0
        absent_count = 0
        late_count = 0

        for no, dt in enumerate(all_dates, start=1):
            row = attendance_map.get((emp["id"], dt))
            status = row["status"] if row else "-"
            arrival_type = row["arrival_type"] if row and row.get("arrival_type") else "-"
            note = row["note"] if row and row.get("note") else ""
            # checkin_at/checkout_at disimpan sbg WIB naive (_now_wib_naive di
            # core.py) -- TANPA offset tambahan, beda dgn kolom UTC lainnya.
            checkin = ""
            if row and row.get("checkin_at"):
                try:
                    checkin = row["checkin_at"].strftime("%H:%M")
                except Exception:
                    checkin = str(row["checkin_at"])
            checkout = ""
            if row and row.get("checkout_at"):
                try:
                    checkout = row["checkout_at"].strftime("%H:%M")
                except Exception:
                    checkout = str(row["checkout_at"])

            if status == "PRESENT":
                present_count += 1
            elif status == "SICK":
                sick_count += 1
            elif status == "LEAVE":
                leave_count += 1
            elif status == "ABSENT":
                absent_count += 1

            if arrival_type == "LATE":
                late_count += 1

            values = [
                no,
                dt.strftime("%d/%m/%Y"),
                calendar.day_name[dt.weekday()],
                status,
                arrival_type,
                checkin,
                checkout,
                note,
            ]

            for col, val in enumerate(values, start=1):
                cell = ws1.cell(current_row, col, val)
                cell.border = BORDER
                cell.alignment = ALIGN_CENTER if col != 8 else ALIGN_LEFT

            current_row += 1

        # Summary per user
        ws1.cell(current_row, 1, "Ringkasan").font = FONT_BOLD
        ws1.cell(current_row, 1).fill = FILL_TOTAL
        ws1.cell(current_row, 1).border = BORDER

        summaries = [
            f"Hadir: {present_count}",
            f"Terlambat: {late_count}",
            f"Sakit: {sick_count}",
            f"Izin: {leave_count}",
            f"Absen: {absent_count}",
        ]

        for i, text in enumerate(summaries, start=2):
            cell = ws1.cell(current_row, i, text)
            cell.fill = FILL_TOTAL
            cell.font = FONT_BOLD
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER

        for i in (7, 8):
            ws1.cell(current_row, i, "").fill = FILL_TOTAL
            ws1.cell(current_row, i).border = BORDER

        current_row += 2

    ws1.freeze_panes = "A3"
    auto_fit(ws1, min_width=10, max_width=24)
    ws1.column_dimensions["H"].width = 28

    # =========================
    # SHEET 2 - REKAP GAJI
    # =========================
    ws2 = wb.create_sheet("Rekap Gaji")

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    c = ws2.cell(1, 1, f"REKAP GAJI KARYAWAN ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})")
    c.fill = FILL_TITLE
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # SATU SUMBER KEBENARAN dgn halaman Gaji Karyawan & Export Print --
    # base gaji + penyesuaian manual (potongan/bonus) + status dibayar
    # SELALU sama persis di sini.
    payroll_employees, workdays = get_payroll_report(start_date, end_date, user_ids)

    headers = [
        "No", "Nama", "Hari Kerja", "Hadir", "Sakit", "Izin", "Absen",
        "Terlambat", "Gaji Harian", "Gaji Pokok", "Penyesuaian",
        "Total Dibayar", "Status Bayar", "Keterangan Penyesuaian",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(3, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER

    row_idx = 4
    grand_total_salary = 0

    for no, emp in enumerate(payroll_employees, start=1):
        grand_total_salary += emp["final_total"]

        adj_note = "; ".join(
            f"{a['date_str']}: {a['note'] or ('Bonus' if a['amount'] > 0 else 'Potongan')}"
            for a in emp["adjustments"]
        )
        pay = emp.get("payment")
        pay_status = f"Sudah dibayar ({pay['paid_at_wib']})" if pay else "Belum dibayar"

        values = [
            no,
            emp["name"],
            workdays,
            emp["present_count"],
            emp["sick_count"],
            emp["leave_count"],
            emp["absent_count"],
            emp["late_count"],
            rupiah_excel(emp["daily_salary"]),
            rupiah_excel(emp["base_salary"]),
            rupiah_excel(emp["adjustments_total"]),
            rupiah_excel(emp["final_total"]),
            pay_status,
            adj_note,
        ]

        for col, val in enumerate(values, start=1):
            cell = ws2.cell(row_idx, col, val)
            cell.border = BORDER
            if col in (1, 3, 4, 5, 6, 7, 8):
                cell.alignment = ALIGN_CENTER
            elif col in (9, 10, 11, 12):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '#,##0'
            else:
                cell.alignment = ALIGN_LEFT

        row_idx += 1

    # Total akhir
    for col in range(1, 14 + 1):
        ws2.cell(row_idx, col).border = BORDER
        ws2.cell(row_idx, col).fill = FILL_TOTAL

    ws2.cell(row_idx, 1, "TOTAL").font = FONT_BOLD
    ws2.cell(row_idx, 1).alignment = ALIGN_CENTER
    ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=11)
    ws2.cell(row_idx, 12, grand_total_salary).font = FONT_BOLD
    ws2.cell(row_idx, 12).alignment = ALIGN_RIGHT
    ws2.cell(row_idx, 12).number_format = '#,##0'

    ws2.freeze_panes = "A4"
    auto_fit(ws2, min_width=10, max_width=24)
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["J"].width = 16

    # =========================
    # OUTPUT
    # =========================
    file_obj = io.BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)

    filename = f"laporan_absensi_gaji_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

    return send_file(
        file_obj,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================
# EXPORT PRINT (continuous form, sama kertas dgn nota)
# =========================
@export_bp.route("/admin/data/range_print")
def export_range_print():
    """Rekap absensi & gaji, versi print continuous form -- data & rumus
    gaji SAMA PERSIS dgn export_range() (Sheet Absensi + Rekap Gaji di
    atas), cuma disusun jadi 1 halaman cetak per kertas dot-matrix,
    bukan file .xlsx."""
    deny = admin_required()
    if deny:
        return deny

    ensure_hr_v2_schema()

    start_str = (request.args.get("start_date") or request.args.get("start") or "").strip()
    end_str = (request.args.get("end_date") or request.args.get("end") or "").strip()
    user_ids = _parse_user_ids(request.args)

    start_date = _parse_date(start_str)
    end_date = _parse_date(end_str)

    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1)
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    # SATU SUMBER KEBENARAN dgn halaman Gaji Karyawan & Excel Sheet "Rekap
    # Gaji" -- base gaji + penyesuaian manual (potongan/bonus) + status
    # dibayar SELALU sama persis di sini.
    result_employees, workdays = get_payroll_report(start_date, end_date, user_ids)
    grand_total_salary = sum(e["final_total"] for e in result_employees)

    return render_template(
        "admin_attendance_print.html",
        employees=result_employees,
        start_date=start_date,
        end_date=end_date,
        workdays=workdays,
        grand_total_salary=grand_total_salary,
    )