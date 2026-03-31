# ==================== IMPORTS ====================
import os
import io
import re
import ssl
import time
import random
import smtplib
import calendar
import hmac
import hashlib
from datetime import datetime, date, timedelta
from email.message import EmailMessage
from functools import wraps
from decimal import Decimal
import uuid
import psycopg2
import cv2
import atexit
import json


# Flask & Extensions
from flask import (
    Flask, render_template, request, redirect, session, abort,
    jsonify, url_for, flash, Response, send_file,
)
import pytz
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

# Database
from psycopg2.extras import RealDictCursor
from db import get_conn

# AI
from openai import OpenAI
from openai import RateLimitError, APIError, AuthenticationError

# Excel Export
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# OAuth
from authlib.integrations.flask_client import OAuth

# Timezone
from zoneinfo import ZoneInfo

# Serial port untuk print thermal via Bluetooth COM port (Windows)
try:
    import serial
    import serial.tools.list_ports
    SERIAL_OK = True
except ImportError:
    SERIAL_OK = False

# Load .env
load_dotenv()

from werkzeug.middleware.proxy_fix import ProxyFix

from werkzeug.utils import secure_filename


from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

# ==================== APP CONFIG ====================
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-me")
IS_PROD = os.getenv("RENDER") == "true" or os.getenv("FLASK_ENV") == "production"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = True if IS_PROD else False
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["PREFERRED_URL_SCHEME"] = "https" if IS_PROD else "http"
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

wib = pytz.timezone("Asia/Jakarta")

@app.after_request
def add_mobile_api_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response

# OpenAI Client
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
OPENAI_API_KEY = (os.getenv("OPENAI_API_KEY") or "").strip()
oa_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

# OAuth Setup
oauth = OAuth(app)
GOOGLE_CLIENT_ID = (os.getenv("GOOGLE_CLIENT_ID") or "").strip()
GOOGLE_CLIENT_SECRET = (os.getenv("GOOGLE_CLIENT_SECRET") or "").strip()

if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
    oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )



# ==================== HELPER FUNCTIONS ====================
def is_logged_in():
    return "user_id" in session

def is_admin():
    return session.get("role") == "admin"

def admin_guard():
    if not is_logged_in():
        return redirect("/login")
    if not is_admin():
        abort(403)
    return None

def admin_required():
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if session.get("role") != "admin":
        flash("Akses ditolak. Hanya admin.", "danger")
        return redirect(url_for("dashboard"))
    return None

UPLOAD_ATT_USER_DIR = os.path.join("static", "uploads", "attendance_user")

def _ensure_att_user_upload_dir():
    os.makedirs(UPLOAD_ATT_USER_DIR, exist_ok=True)

UPLOAD_QA_DIR = os.path.join("static", "uploads", "quick_attendance")

def _ensure_upload_dir():
    os.makedirs(UPLOAD_QA_DIR, exist_ok=True)

def cleanup_old_quick_attendance_photos():
    """
    Hapus foto quick attendance yang bukan hari ini.
    Nama file kita buat prefix: qa_YYYY_MM_DD_...
    """
    _ensure_upload_dir()
    today_prefix = "qa_" + date.today().strftime("%Y_%m_%d") + "_"
    for fn in os.listdir(UPLOAD_QA_DIR):
        if fn.startswith("qa_") and not fn.startswith(today_prefix):
            try:
                os.remove(os.path.join(UPLOAD_QA_DIR, fn))
            except Exception as e:
                print("cleanup error:", fn, e)

def is_token_valid(token: str) -> bool:
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(
            "SELECT 1 FROM attendance_links WHERE token=%s AND is_active=TRUE LIMIT 1;",
            (token,)
        )
        return cur.fetchone() is not None
    finally:
        cur.close()
        conn.close()

def _public_ip():
    # Render biasanya set X-Forwarded-For
    xf = request.headers.get("X-Forwarded-For", "")
    if xf:
        return xf.split(",")[0].strip()
    return request.remote_addr

def login_required():
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def _parse_manual_wib_naive(manual_dt):
    if not manual_dt:
        return None
    try:
        return datetime.strptime(manual_dt.strip(), "%Y-%m-%dT%H:%M")
    except Exception:
        return None

def _now_wib_naive_from_form():
    client_ts = request.form.get("client_ts")
    if client_ts and client_ts.isdigit():
        now_wib_aware = datetime.fromtimestamp(int(client_ts) / 1000, tz=ZoneInfo("Asia/Jakarta"))
    else:
        now_wib_aware = datetime.now(ZoneInfo("Asia/Jakarta"))
    return now_wib_aware.replace(tzinfo=None)

def _utc_naive_to_wib_naive(dt):
    """
    DB kamu sekarang banyak menyimpan timestamp naive yang tampil seperti UTC.
    Fungsi ini geser +7 jam agar sinkron ke WIB.
    """
    if not dt:
        return None
    return dt + timedelta(hours=7)

def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def pick(options):
    return random.choice(options)

def rupiah(s):
    try:
        n = int(s)
        return f"Rp {n:,}".replace(",", ".")
    except:
        return f"Rp {s}"

def _rupiah(value):
    try:
        n = int(value)
        return f"Rp {n:,}".replace(",", ".")
    except:
        return f"Rp {value}"

def _pick(rng, items):
    """Pick a random item from a list using the provided RNG."""
    return items[rng.randrange(len(items))]

def _otp_hash(email, otp):
    salt = (os.getenv("RESET_OTP_SALT") or "umgap-reset-salt").encode("utf-8")
    msg = (email.lower().strip() + ":" + otp.strip()).encode("utf-8")
    return hashlib.sha256(salt + msg).hexdigest()


# =========================
# Helper invoice
# =========================

def _make_invoice_no():
    now = datetime.now(ZoneInfo("Asia/Jakarta"))
    return "INV-" + now.strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:5].upper()


def _safe_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default

def _safe_decimal(v, default="0"):
    try:
        raw = str(v or "").strip().replace(",", ".")
        if not raw:
            return Decimal(default)
        return Decimal(raw)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _decimal_to_display(v):
    d = _safe_decimal(v)
    s = format(d.normalize(), "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s.replace(".", ",")

def _invoice_rows_from_form(form):
    product_ids = form.getlist("product_id[]")
    qtys = form.getlist("qty[]")

    rows = []
    for i in range(min(len(product_ids), len(qtys))):
        pid = _safe_int(product_ids[i], 0)
        qty = _safe_decimal(qtys[i], "0")

        if pid > 0 and qty > 0:
            rows.append({
                "product_id": pid,
                "qty": qty
            })
    return rows

CCTV_SNAPSHOT_DIR = os.path.join("static", "uploads", "cctv_snapshots")

def _ensure_cctv_snapshot_dir():
    os.makedirs(CCTV_SNAPSHOT_DIR, exist_ok=True)

def save_cctv_event_snapshot(camera_code, frame, event_type="MOTION", motion_score=0, note=""):
    _ensure_cctv_event_snapshot_dir()

    filename = f"{camera_code}_{event_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.jpg"
    save_path = os.path.join(CCTV_EVENT_SNAPSHOT_DIR, filename)

    ok = cv2.imwrite(save_path, frame)
    if not ok:
        return None

    rel_path = f"uploads/cctv_events/{filename}"

    ensure_cctv_report_schema()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            INSERT INTO cctv_event_snapshots
                (camera_code, event_type, snapshot_path, motion_score, note)
            VALUES
                (%s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            camera_code,
            event_type,
            rel_path,
            int(motion_score or 0),
            note or "",
        ))
        row = cur.fetchone()
        conn.commit()
        return row["id"] if row else None
    finally:
        cur.close()
        conn.close()    

UPLOAD_INVOICE_LOGO_DIR = os.path.join("static", "uploads", "invoice_logo")

def _ensure_invoice_logo_dir():
    os.makedirs(UPLOAD_INVOICE_LOGO_DIR, exist_ok=True)

def _save_company_logo(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    _ensure_invoice_logo_dir()

    raw_name = secure_filename(file_storage.filename or "")
    ext = os.path.splitext(raw_name)[1].lower()
    if ext not in [".png", ".jpg", ".jpeg", ".webp"]:
        ext = ".png"

    filename = f"inv_logo_{datetime.now(ZoneInfo('Asia/Jakarta')).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    save_path = os.path.join(UPLOAD_INVOICE_LOGO_DIR, filename)
    file_storage.save(save_path)

    return f"uploads/invoice_logo/{filename}"

def _utc_naive_to_wib_string(dt, fmt="%d/%m/%Y %H:%M"):
    if not dt:
        return "-"
    try:
        return (dt + timedelta(hours=7)).strftime(fmt)
    except Exception:
        return "-"
# =========================
# APP MOBOLE API
# =========================
def ensure_mobile_api_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mobile_api_tokens (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                token TEXT NOT NULL UNIQUE,
                device_name VARCHAR(120),
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                last_used_at TIMESTAMP
            );
        """)
        conn.commit()
    finally:
        cur.close()
        conn.close()

#-------- HELPER ANDROID ---------
def mobile_api_response(ok=True, message="", data=None, status_code=200):
    payload = {
        "ok": ok,
        "message": message,
        "data": data if data is not None else {}
    }
    return jsonify(payload), status_code


def get_bearer_token():
    auth = request.headers.get("Authorization", "").strip()
    if not auth.lower().startswith("bearer "):
        return None
    return auth[7:].strip()


def get_mobile_api_user():
    token = get_bearer_token()
    if not token:
        return None

    ensure_mobile_api_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                t.id AS token_id,
                t.user_id,
                t.token,
                t.is_active,
                u.id,
                u.name,
                u.email,
                u.role,
                COALESCE(u.points, 0) AS points,
                COALESCE(u.points_admin, 0) AS points_admin
            FROM mobile_api_tokens t
            JOIN users u ON u.id = t.user_id
            WHERE t.token=%s
              AND t.is_active=TRUE
            LIMIT 1;
        """, (token,))
        row = cur.fetchone()

        if row:
            cur.execute("""
                UPDATE mobile_api_tokens
                SET last_used_at=CURRENT_TIMESTAMP
                WHERE id=%s;
            """, (row["token_id"],))
            conn.commit()

        return row
    finally:
        cur.close()
        conn.close()


def mobile_api_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_mobile_api_user()
        if not user:
            return mobile_api_response(
                ok=False,
                message="Unauthorized. Token tidak valid atau belum login.",
                status_code=401
            )
        request.mobile_user = user
        return fn(*args, **kwargs)
    return wrapper



# =========================
# SIMPAN INVOICE
# =========================

def _save_invoice(is_admin_mode=False):
    ensure_invoice_schema()

    created_by = session.get("user_id")
    customer_name = (request.form.get("customer_name") or "").strip()
    customer_phone = (request.form.get("customer_phone") or "").strip()
    company_name = (request.form.get("company_name") or "").strip()
    payment_method = (request.form.get("payment_method") or "CASH").strip().upper()
    print_size = (request.form.get("print_size") or "80mm").strip()
    notes = (request.form.get("notes") or "").strip()
    discount = max(0, _safe_int(request.form.get("discount"), 0))
    is_paid = str(request.form.get("is_paid") or "1").strip() in ("1", "true", "True", "on", "yes")
    paid_at = datetime.utcnow() if is_paid else None

    logo_file = request.files.get("company_logo")
    company_logo_path = _save_company_logo(logo_file)

    target_user_id = created_by
    if is_admin_mode:
        target_user_id = _safe_int(request.form.get("employee_id"), created_by)

    item_rows = _invoice_rows_from_form(request.form)
    if not item_rows:
        return redirect("/admin/invoice/new" if is_admin_mode else "/invoice/new")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        invoice_no = _make_invoice_no()

        final_items = []
        subtotal = Decimal("0")

        for row in item_rows:
            cur.execute("""
                SELECT id, name, price
                FROM products
                WHERE id=%s AND is_global=TRUE
                LIMIT 1;
            """, (row["product_id"],))
            p = cur.fetchone()

            if not p:
                continue

            qty = _safe_decimal(row["qty"], "0")
            price = Decimal(str(int(p.get("price") or 0)))
            line_subtotal = (qty * price).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            subtotal += line_subtotal

            final_items.append({
                "product_id": p["id"],
                "product_name": p["name"],
                "qty": qty,
                "price": int(price),
                "subtotal": int(line_subtotal)
            })

        if not final_items:
            return redirect("/admin/invoice/new" if is_admin_mode else "/invoice/new")

        grand_total = max(0, subtotal - discount)

        cur.execute("""
            INSERT INTO invoices
                (invoice_no, created_by, customer_name, customer_phone, company_name, company_logo_path,
                 print_size, payment_method, subtotal, discount, grand_total, notes, is_paid, paid_at)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (
            invoice_no,
            created_by,
            customer_name,
            customer_phone,
            company_name,
            company_logo_path,
            print_size,
            payment_method,
            subtotal,
            discount,
            grand_total,
            notes,
            is_paid,
            paid_at
        ))

        invoice_id = (cur.fetchone() or {}).get("id")

        for item in final_items:
            cur.execute("""
                INSERT INTO invoice_items
                (invoice_id, product_id, product_name, qty, price, subtotal)
                VALUES (%s,%s,%s,%s,%s,%s);
            """, (
                invoice_id,
                item["product_id"],
                item["product_name"],
                str(item["qty"]),
                item["price"],
                item["subtotal"]
            ))

            cur.execute("""
                INSERT INTO sales_submissions
                (user_id, product_id, qty, note, status, created_at)
                VALUES (%s,%s,%s,%s,'APPROVED',CURRENT_TIMESTAMP);
            """, (
                target_user_id,
                item["product_id"],
                int(Decimal(str(item["qty"])).quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
                f"INVOICE {invoice_no}"
            ))

        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect(f"/invoice/{invoice_id}")

# === CCTV MULTI CAMERA CONFIG ===
CCTV_CAMERAS = [
    {
        "code": "CAM01",
        "name": "Kamera 1",
        "location": "Area 1",
        "rtsp_url": os.getenv("CCTV_CAM01_RTSP", "").strip(),
    },
    {
        "code": "CAM02",
        "name": "Kamera 2",
        "location": "Gudang",
        "rtsp_url": os.getenv("CCTV_CAM02_RTSP", "").strip(),
    },
    {
        "code": "CAM03",
        "name": "Kamera 3",
        "location": "Bak Muat",
        "rtsp_url": os.getenv("CCTV_CAM03_RTSP", "").strip(),
    },
    {
        "code": "CAM04",
        "name": "Kamera 4",
        "location": "Area Dalam",
        "rtsp_url": os.getenv("CCTV_CAM04_RTSP", "").strip(),
    },
]

# === MODE TEST CCTV (PAKAI VIDEO LOKAL SAAT CCTV ASLI TIDAK BISA DIAKSES) ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USE_TEST_VIDEO = (os.getenv("USE_TEST_VIDEO") or "true").strip().lower() == "true"

TEST_VIDEO_DEFAULT = os.path.join(BASE_DIR, "static", "sample.mp4")

TEST_VIDEO_MAP = {
    "CAM01": os.path.join(BASE_DIR, "static", "sample1.mp4"),
    "CAM02": os.path.join(BASE_DIR, "static", "sample2.mp4"),
    "CAM03": os.path.join(BASE_DIR, "static", "sample3.mp4"),
    "CAM04": os.path.join(BASE_DIR, "static", "sample4.mp4"),
}


def get_camera_url(camera_code):
    camera_code = (camera_code or "").strip().upper()

    # kalau mode test aktif, pakai video lokal
    if USE_TEST_VIDEO:
        path = TEST_VIDEO_MAP.get(camera_code, TEST_VIDEO_DEFAULT)
        if os.path.exists(path):
            return path
        return TEST_VIDEO_DEFAULT

    # kalau mode test mati, pakai RTSP asli dari config
    cam = get_camera_config(camera_code)
    if cam and (cam.get("rtsp_url") or "").strip():
        return cam["rtsp_url"].strip()

    return None

CCTV_ROIS = {
    "CAM01": (120, 120, 1100, 650),
    "CAM02": (80, 80, 1150, 680),
    "CAM03": (100, 100, 1100, 650),
    "CAM04": (100, 100, 1100, 650),
}

CCTV_EVENT_SNAPSHOT_DIR = os.path.join("static", "uploads", "cctv_events")

def _ensure_cctv_event_snapshot_dir():
    os.makedirs(CCTV_EVENT_SNAPSHOT_DIR, exist_ok=True)

CCTV_ACTIVITY_STATE = {}
CCTV_ZONE_ACTIVITY_STATE = {}
CCTV_BG_SUBTRACTORS = {}
CCTV_LAST_SNAPSHOT_TS = {}

def detect_motion_simple(frame, camera_code):
    if camera_code not in CCTV_BG_SUBTRACTORS:
        CCTV_BG_SUBTRACTORS[camera_code] = cv2.createBackgroundSubtractorMOG2(
            history=300,
            varThreshold=25,
            detectShadows=True
        )

    roi = CCTV_ROIS.get(camera_code)
    if roi:
        x1, y1, x2, y2 = roi
        crop = frame[y1:y2, x1:x2]
    else:
        x1, y1, x2, y2 = 0, 0, frame.shape[1], frame.shape[0]
        crop = frame

    if crop is None or crop.size == 0:
        return False, 0, (x1, y1, x2, y2)

    fgbg = CCTV_BG_SUBTRACTORS[camera_code]

    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (21, 21), 0)

    fgmask = fgbg.apply(blur)
    _, th = cv2.threshold(fgmask, 200, 255, cv2.THRESH_BINARY)

    motion_pixels = cv2.countNonZero(th)
    moving = motion_pixels > 2500

    return moving, motion_pixels, (x1, y1, x2, y2)

def get_camera_config(camera_code):
    camera_code = (camera_code or "").strip().upper()
    for cam in CCTV_CAMERAS:
        if (cam.get("code") or "").strip().upper() == camera_code:
            return cam
    return None


def _make_offline_frame_multi(camera_name, message):
    import numpy as np

    frame = np.zeros((720, 1280, 3), dtype=np.uint8)
    frame[:] = (36, 44, 62)

    cv2.putText(
        frame,
        camera_name or "CCTV",
        (60, 120),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.5,
        (255, 255, 255),
        3,
        cv2.LINE_AA,
    )
    cv2.putText(
        frame,
        message,
        (60, 200),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.9,
        (180, 210, 255),
        2,
        cv2.LINE_AA,
    )
    cv2.putText(
        frame,
        "Periksa RTSP URL / user / password / channel DVR.",
        (60, 250),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        (210, 220, 230),
        2,
        cv2.LINE_AA,
    )
    return frame


def _jpeg_chunk_multi(frame):
    ok, buffer = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
    if not ok:
        return (
            b"--frame\r\n"
            b"Content-Type: image/jpeg\r\n\r\n" + b"" + b"\r\n"
        )
    return (
        b"--frame\r\n"
        b"Content-Type: image/jpeg\r\n\r\n" + buffer.tobytes() + b"\r\n"
    )


def generate_frames_multi(rtsp_url, camera_name="CCTV", camera_code="CAM"):
    cap = None

    while True:
        try:
            if cap is None or not cap.isOpened():
                cap = cv2.VideoCapture(rtsp_url)
                try:
                    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                except Exception:
                    pass

                if cap is None or not cap.isOpened():
                    frame = _make_offline_frame_multi(camera_name, "Sumber video tidak bisa dibuka")
                    yield _jpeg_chunk_multi(frame)
                    time.sleep(1)
                    continue

            ok, frame = cap.read()

            # kalau video lokal habis, ulang dari awal
            if not ok or frame is None:
                if str(rtsp_url).lower().endswith((".mp4", ".avi", ".mov", ".mkv")):
                    try:
                        cap.release()
                    except Exception:
                        pass
                    cap = cv2.VideoCapture(rtsp_url)
                    ok, frame = cap.read()

                if not ok or frame is None:
                    try:
                        cap.release()
                    except Exception:
                        pass
                    cap = None

                    frame = _make_offline_frame_multi(camera_name, "Stream gagal dibaca, mencoba ulang...")
                    yield _jpeg_chunk_multi(frame)
                    time.sleep(1)
                    continue

            moving, motion_score, roi = detect_motion_simple(frame, camera_code)

            status = "MOVING" if moving else "IDLE"
            color = (0, 200, 0) if moving else (0, 0, 255)

            now_ts = time.time()
            last_snap_ts = CCTV_LAST_SNAPSHOT_TS.get(camera_code, 0)

            # snapshot otomatis tiap 5 menit saat ada gerakan
            if moving and (now_ts - last_snap_ts) >= 300:
                snap_id = save_cctv_event_snapshot(
                    camera_code=camera_code,
                    frame=frame,
                    event_type="MOTION",
                    motion_score=motion_score,
                    note="Snapshot otomatis karena ada gerakan",
                )
                if snap_id:
                    CCTV_LAST_SNAPSHOT_TS[camera_code] = now_ts

            h, w = frame.shape[:2]
            max_w = 1280
            if w > max_w:
                ratio = max_w / float(w)
                frame = cv2.resize(frame, (int(w * ratio), int(h * ratio)))

            x1, y1, x2, y2 = roi
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

            ts = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            cv2.rectangle(frame, (12, 12), (560, 100), (20, 20, 20), -1)
            cv2.putText(frame, str(camera_name), (22, 35),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2, cv2.LINE_AA)
            cv2.putText(frame, ts, (22, 58),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (220, 220, 220), 1, cv2.LINE_AA)
            cv2.putText(frame, f"STATUS: {status}", (22, 82),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2, cv2.LINE_AA)
            cv2.putText(frame, f"SCORE: {motion_score}", (250, 82),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2, cv2.LINE_AA)

            yield _jpeg_chunk_multi(frame)

        except GeneratorExit:
            break
        except Exception as e:
            if cap is not None:
                try:
                    cap.release()
                except Exception:
                    pass
                cap = None

            frame = _make_offline_frame_multi(camera_name, f"Error stream: {str(e)[:80]}")
            yield _jpeg_chunk_multi(frame)
            time.sleep(1)

    if cap is not None:
        try:
            cap.release()
        except Exception:
            pass

def _seconds_to_hms(total_seconds):
    total_seconds = int(total_seconds or 0)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def _seconds_to_hms(total_seconds):
    total_seconds = int(total_seconds or 0)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _get_active_employee_zones(camera_code):
    ensure_cctv_activity_schema()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                z.id,
                z.camera_code,
                z.user_id,
                z.zone_name,
                z.x1, z.y1, z.x2, z.y2,
                z.idle_threshold_seconds,
                z.is_active,
                u.name AS user_name
            FROM cctv_employee_zones z
            LEFT JOIN users u ON u.id = z.user_id
            WHERE z.camera_code=%s
              AND z.is_active=TRUE
            ORDER BY z.id ASC;
        """, (camera_code,))
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()

def ensure_cctv_report_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_event_snapshots (
                id SERIAL PRIMARY KEY,
                camera_code VARCHAR(20) NOT NULL,
                event_type VARCHAR(30) NOT NULL DEFAULT 'MOTION',
                snapshot_path TEXT NOT NULL,
                motion_score INTEGER NOT NULL DEFAULT 0,
                note TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_snapshot_labels (
                id SERIAL PRIMARY KEY,
                snapshot_id INTEGER NOT NULL REFERENCES cctv_event_snapshots(id) ON DELETE CASCADE,
                subject_label VARCHAR(100) NOT NULL,
                note TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        conn.commit()
    finally:
        cur.close()
        conn.close()

def _flush_cctv_activity(camera_code, seconds_to_add, status, motion_score=0):
    if seconds_to_add <= 0:
        return

    ensure_cctv_activity_schema()
    today = date.today()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO cctv_activity_daily
                (camera_code, work_date, moving_seconds, idle_seconds, last_status, last_motion_score, updated_at)
            VALUES
                (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (camera_code, work_date)
            DO UPDATE SET
                moving_seconds = cctv_activity_daily.moving_seconds + EXCLUDED.moving_seconds,
                idle_seconds   = cctv_activity_daily.idle_seconds + EXCLUDED.idle_seconds,
                last_status    = EXCLUDED.last_status,
                last_motion_score = EXCLUDED.last_motion_score,
                updated_at     = CURRENT_TIMESTAMP;
        """, (
            camera_code,
            today,
            seconds_to_add if status == "MOVING" else 0,
            seconds_to_add if status == "IDLE" else 0,
            status,
            int(motion_score or 0),
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()


def _flush_zone_activity(camera_code, zone_id, user_id, seconds_to_add, status, motion_score=0):
    if seconds_to_add <= 0:
        return

    ensure_cctv_activity_schema()
    today = date.today()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO cctv_employee_activity_daily
                (camera_code, zone_id, user_id, work_date,
                 moving_seconds, idle_seconds, absent_seconds,
                 last_status, last_motion_score, updated_at)
            VALUES
                (%s, %s, %s, %s,
                 %s, %s, %s,
                 %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (zone_id, work_date)
            DO UPDATE SET
                moving_seconds = cctv_employee_activity_daily.moving_seconds + EXCLUDED.moving_seconds,
                idle_seconds   = cctv_employee_activity_daily.idle_seconds + EXCLUDED.idle_seconds,
                absent_seconds = cctv_employee_activity_daily.absent_seconds + EXCLUDED.absent_seconds,
                last_status    = EXCLUDED.last_status,
                last_motion_score = EXCLUDED.last_motion_score,
                updated_at     = CURRENT_TIMESTAMP;
        """, (
            camera_code,
            zone_id,
            user_id,
            today,
            seconds_to_add if status == "MOVING" else 0,
            seconds_to_add if status == "IDLE" else 0,
            seconds_to_add if status == "ABSENT" else 0,
            status,
            int(motion_score or 0),
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()


def _save_idle_snapshot(camera_code, zone_id, user_id, frame, roi, idle_seconds, motion_score):
    _ensure_cctv_snapshot_dir()

    x1, y1, x2, y2 = roi
    crop = frame[y1:y2, x1:x2]
    if crop is None or crop.size == 0:
        return None

    fname = f"idle_{camera_code}_{zone_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.jpg"
    fpath = os.path.join(CCTV_SNAPSHOT_DIR, fname)

    ok = cv2.imwrite(fpath, crop)
    if not ok:
        return None

    rel_path = f"uploads/cctv_snapshots/{fname}"

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO cctv_idle_snapshots
                (camera_code, zone_id, user_id, snapshot_path, idle_seconds, motion_score, status)
            VALUES
                (%s, %s, %s, %s, %s, %s, 'IDLE');
        """, (
            camera_code,
            zone_id,
            user_id,
            rel_path,
            int(idle_seconds or 0),
            int(motion_score or 0),
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return rel_path


def detect_motion_simple(frame, camera_code, roi=None):
    key = f"{camera_code}:{roi if roi else 'FULL'}"

    if key not in CCTV_BG_SUBTRACTORS:
        CCTV_BG_SUBTRACTORS[key] = cv2.createBackgroundSubtractorMOG2(
            history=300,
            varThreshold=25,
            detectShadows=True
        )

    if roi:
        x1, y1, x2, y2 = roi
        crop = frame[y1:y2, x1:x2]
    else:
        x1, y1, x2, y2 = 0, 0, frame.shape[1], frame.shape[0]
        crop = frame

    if crop is None or crop.size == 0:
        return False, 0, (x1, y1, x2, y2)

    fgbg = CCTV_BG_SUBTRACTORS[key]

    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (21, 21), 0)
    fgmask = fgbg.apply(blur)

    _, th = cv2.threshold(fgmask, 200, 255, cv2.THRESH_BINARY)
    motion_pixels = cv2.countNonZero(th)

    moving = motion_pixels > 2500
    return moving, motion_pixels, (x1, y1, x2, y2)

# ==================== SCHEMA ENSURERS ====================
def ensure_points_schema():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS points INTEGER DEFAULT 0;")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS points_admin INTEGER DEFAULT 0;")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS points_logs (
            id SERIAL PRIMARY KEY,
            user_id INT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            admin_id INT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            delta INT NOT NULL,
            note TEXT,
            created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW()
        );
    """)
    conn.commit()
    cur.close()
    conn.close()

def init_points_v1():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS points INTEGER NOT NULL DEFAULT 0;")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS points_admin INTEGER NOT NULL DEFAULT 0;")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS points_logs (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            admin_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            delta INTEGER NOT NULL,
            note TEXT,
            created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()
    conn.close()

def ensure_hr_v2_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS payroll_settings (
                user_id INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
                daily_salary INTEGER NOT NULL DEFAULT 0,
                monthly_salary INTEGER NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        cur.execute("ALTER TABLE payroll_settings ADD COLUMN IF NOT EXISTS daily_salary INTEGER NOT NULL DEFAULT 0;")
        cur.execute("ALTER TABLE payroll_settings ADD COLUMN IF NOT EXISTS monthly_salary INTEGER NOT NULL DEFAULT 0;")
        cur.execute("ALTER TABLE payroll_settings ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;")
        cur.execute("ALTER TABLE attendance ADD COLUMN IF NOT EXISTS arrival_type VARCHAR(20) NOT NULL DEFAULT 'ONTIME';")
        cur.execute("ALTER TABLE attendance ADD COLUMN IF NOT EXISTS checkin_at TIMESTAMP NULL;")
        conn.commit()
    finally:
        cur.close()
        conn.close()

def ensure_password_reset_schema():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_otps (
            id SERIAL PRIMARY KEY,
            email TEXT NOT NULL,
            otp_hash TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            used BOOLEAN NOT NULL DEFAULT FALSE,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()
    conn.close()

def ensure_announcements_schema():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS announcements (
            id SERIAL PRIMARY KEY,
            title VARCHAR(255) NOT NULL,
            message TEXT NOT NULL,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS announcement_reads (
            id SERIAL PRIMARY KEY,
            announcement_id INTEGER NOT NULL REFERENCES announcements(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            read_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(announcement_id, user_id)
        );
    """)
    conn.commit()
    cur.close()
    conn.close()

def get_unread_notifications(user_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT COUNT(*) AS total
            FROM announcements a
            WHERE a.is_active = TRUE
              AND a.id NOT IN (
                SELECT announcement_id
                FROM announcement_reads
                WHERE user_id = %s
              )
        """, (user_id,))
        row = cur.fetchone() or {"total": 0}
        return int(row.get("total", 0) or 0)
    finally:
        cur.close()
        conn.close()

def get_notif_count():
    conn = get_conn()
    # PAKAI RealDictCursor kalau kamu konsisten pakai dict di app
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("SELECT COUNT(*) AS total FROM announcements WHERE is_active = TRUE;")
        row = cur.fetchone() or {"total": 0}
        return int(row.get("total", 0) or 0)
    finally:
        cur.close()
        conn.close()

def ensure_attendance_links_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 1) pastikan tabel ada (minimal)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance_links (
                id SERIAL PRIMARY KEY,
                token TEXT UNIQUE NOT NULL
            );
        """)

        # 2) migrasi kolom-kolom yang mungkin belum ada (untuk DB lama)
        cur.execute("ALTER TABLE attendance_links ADD COLUMN IF NOT EXISTS title TEXT;")
        cur.execute("ALTER TABLE attendance_links ADD COLUMN IF NOT EXISTS created_by INTEGER;")
        cur.execute("ALTER TABLE attendance_links ADD COLUMN IF NOT EXISTS created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP;")
        cur.execute("ALTER TABLE attendance_links ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE;")

        conn.commit()
    finally:
        cur.close()
        conn.close()

def ensure_cctv_activity_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_activity_daily (
                id SERIAL PRIMARY KEY,
                camera_code VARCHAR(20) NOT NULL,
                work_date DATE NOT NULL,
                moving_seconds INTEGER NOT NULL DEFAULT 0,
                idle_seconds INTEGER NOT NULL DEFAULT 0,
                last_status VARCHAR(20),
                last_motion_score INTEGER NOT NULL DEFAULT 0,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(camera_code, work_date)
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_employee_zones (
                id SERIAL PRIMARY KEY,
                camera_code VARCHAR(20) NOT NULL,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                zone_name VARCHAR(100) NOT NULL,
                x1 INTEGER NOT NULL,
                y1 INTEGER NOT NULL,
                x2 INTEGER NOT NULL,
                y2 INTEGER NOT NULL,
                idle_threshold_seconds INTEGER NOT NULL DEFAULT 600,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_employee_activity_daily (
                id SERIAL PRIMARY KEY,
                camera_code VARCHAR(20) NOT NULL,
                zone_id INTEGER REFERENCES cctv_employee_zones(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                work_date DATE NOT NULL,
                moving_seconds INTEGER NOT NULL DEFAULT 0,
                idle_seconds INTEGER NOT NULL DEFAULT 0,
                absent_seconds INTEGER NOT NULL DEFAULT 0,
                last_status VARCHAR(20),
                last_motion_score INTEGER NOT NULL DEFAULT 0,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(zone_id, work_date)
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cctv_idle_snapshots (
                id SERIAL PRIMARY KEY,
                camera_code VARCHAR(20) NOT NULL,
                zone_id INTEGER REFERENCES cctv_employee_zones(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                snapshot_path TEXT NOT NULL,
                idle_seconds INTEGER NOT NULL DEFAULT 0,
                motion_score INTEGER NOT NULL DEFAULT 0,
                status VARCHAR(20) NOT NULL DEFAULT 'IDLE',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        conn.commit()
    finally:
        cur.close()
        conn.close()

def _flush_cctv_activity(camera_code, seconds_to_add, status, motion_score=0):
    if seconds_to_add <= 0:
        return

    ensure_cctv_activity_schema()

    today = date.today()
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO cctv_activity_daily
                (camera_code, work_date, moving_seconds, idle_seconds, last_status, last_motion_score, updated_at)
            VALUES
                (%s, %s,
                 %s,
                 %s,
                 %s,
                 %s,
                 CURRENT_TIMESTAMP)
            ON CONFLICT (camera_code, work_date)
            DO UPDATE SET
                moving_seconds = cctv_activity_daily.moving_seconds + EXCLUDED.moving_seconds,
                idle_seconds   = cctv_activity_daily.idle_seconds + EXCLUDED.idle_seconds,
                last_status    = EXCLUDED.last_status,
                last_motion_score = EXCLUDED.last_motion_score,
                updated_at     = CURRENT_TIMESTAMP;
        """, (
            camera_code,
            today,
            seconds_to_add if status == "MOVING" else 0,
            seconds_to_add if status == "IDLE" else 0,
            status,
            int(motion_score or 0),
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()

# ============== INVOICE GENERATOR ================
def ensure_invoice_schema():
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                invoice_no VARCHAR(50) UNIQUE NOT NULL,
                created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                customer_name VARCHAR(150),
                company_name VARCHAR(150),
                company_logo_path TEXT,
                print_size VARCHAR(10) NOT NULL DEFAULT '80mm',
                payment_method VARCHAR(30) DEFAULT 'CASH',
                subtotal INTEGER NOT NULL DEFAULT 0,
                grand_total INTEGER NOT NULL DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS company_name VARCHAR(150);")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS company_logo_path TEXT;")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS customer_phone VARCHAR(30);")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS discount INTEGER NOT NULL DEFAULT 0;")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS is_paid BOOLEAN NOT NULL DEFAULT TRUE;")
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS paid_at TIMESTAMP NULL;")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoice_items (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER NOT NULL REFERENCES invoices(id) ON DELETE CASCADE,
                product_id INTEGER REFERENCES products(id) ON DELETE SET NULL,
                product_name VARCHAR(150) NOT NULL,
                qty NUMERIC(12,3) NOT NULL DEFAULT 1,
                price INTEGER NOT NULL DEFAULT 0,
                subtotal INTEGER NOT NULL DEFAULT 0
            );
        """)

        cur.execute("""
            ALTER TABLE invoice_items
            ALTER COLUMN qty TYPE NUMERIC(12,3)
            USING qty::numeric;
        """)

        conn.commit()
    finally:
        cur.close()
        conn.close()

# ==================== AI CAPTION GENERATOR ====================
def generate_caption_ai(product, price, style, brand="", platform="Instagram", notes=""):
    product = (product or "").strip()
    price = (price or "").strip()
    style = (style or "Santai").strip()
    brand = (brand or "").strip()
    platform = (platform or "Instagram").strip()
    notes = (notes or "").strip()

    brand_rule = f'- WAJIB sebut brand "{brand}" minimal 1x di tiap versi.\n' if brand else ""
    price_hint = f'- Cantumkan harga "{price}" dengan format yang enak dibaca.\n' if price else ""

    platform_rules = {
        "WhatsApp": "Gaya WhatsApp: Ringkas, 4–7 baris, CTA chat/wa, hashtag 0–2.",
        "TikTok": "Gaya TikTok: Wajib HOOK 1 baris, 8–14 baris, ajakan komentar, hashtag 5–9.",
        "Instagram": "Gaya Instagram: Semi storytelling, 9–15 baris, CTA DM/komentar, hashtag 5–10."
    }
    plat_rule = platform_rules.get(platform, platform_rules["Instagram"])
    notes_block = f'Catatan pendukung: "{notes}"\n' if notes else ""

    prompt = f"""
Kamu copywriter UMKM Indonesia. Buat caption ORGANIK, manusiawi.

DATA: Produk: "{product}", Platform: "{platform}", Tone: "{style}"
{notes_block}

ATURAN: {brand_rule}{price_hint}Tidak klaim medis. Bahasa Indonesia sehari-hari.

{plat_rule}

Buat 3 versi berbeda: V1 (manfaat), V2 (cerita), V3 (promo halus).
Format: V1: ... V2: ... V3: ...
"""

    r = client.responses.create(
        model="gpt-4.1-mini",
        input=prompt,
        temperature=0.9,
        max_output_tokens=700
    )
    return (r.output_text or "").strip()

def build_caption(data):
    seed = time.time_ns()
    rng = random.Random(seed)

    template = (data.get("template") or "promo").strip()
    biz_type = (data.get("biz_type") or "produk").strip()
    tone = (data.get("tone") or "santai").strip()
    product = (data.get("product") or "").strip()
    price = _rupiah((data.get("price") or "").strip())
    wa = (data.get("wa") or "").strip()
    location = (data.get("location") or "").strip()
    extra = (data.get("extra") or "").strip()

    loc_line = f"📍 Lokasi: {location}\n" if location else ""
    extra_line = f"📝 Catatan: {extra}\n" if extra else ""

    hooks = {
        "santai": ["Lagi cari yang pas? Cek ini dulu 👇", "Info cepat, siapa tau cocok 👇", "Gas cek detailnya ya 👇"],
        "formal": ["Berikut informasi penawaran kami:", "Rincian penawaran saat ini:", "Detail layanan/produk:"],
        "sales": ["Jangan sampai kelewatan!", "Terbatas! Amankan sekarang!", "Kesempatan bagus—gas sekarang!"]
    }

    benefits_produk = ["Kualitas terjaga", "Cocok untuk kebutuhan harian", "Praktis & siap pakai", "Packing aman"]
    benefits_jasa = ["Pengerjaan rapi & profesional", "Tepat waktu", "Harga transparan", "Bisa konsultasi dulu"]

    ctas = {
        "santai": ["Chat aja ya 👉", "Langsung WA ya 👉", "Siap bantu order 👉"],
        "formal": ["Silakan hubungi:", "Hubungi admin:", "Reservasi melalui:"],
        "sales": ["Order sekarang!", "Amankan slot sekarang!", "Langsung WA!"]
    }

    hashtags = ["#UMKM", "#Promo", "#LocalBrand", "#Indonesia", "#BisnisLokal"]

    benefit = _pick(rng, benefits_produk if biz_type == "produk" else benefits_jasa)
    hook = _pick(rng, hooks.get(tone, hooks["santai"]))
    cta = _pick(rng, ctas.get(tone, ctas["santai"]))
    tagline = _pick(rng, ["✨", "🔥", "📌", "✅", "💡", "⚡"])

    s = _pick(rng, ["A", "B", "C"])

    if template == "promo":
        promo_line = _pick(rng, ["Harga spesial periode terbatas.", "Bisa tanya detail dulu ya.", "Order sekarang, proses mudah."])
        if s == "A":
            caption = f"{hook}\n\n{tagline} {product}\n💰 Mulai {price}\n✅ {benefit}\n{loc_line}{extra_line}📌 {promo_line}\n\n{cta} {wa}\n{_pick(rng, hashtags)} {_pick(rng, hashtags)}"
        elif s == "B":
            caption = f"{tagline} PROMO!\n{product} (mulai {price})\n- {benefit}\n- {promo_line}\n{loc_line}{extra_line}{cta} {wa}"
        else:
            caption = f"{hook}\n🎯 {product} • {price}\n✅ {benefit}\n{loc_line}{extra_line}⏳ {promo_line}\n📲 {cta} {wa}"

    elif template == "new":
        intro = _pick(rng, ["Rilis!", "Baru tersedia!", "New arrival!"])
        if s == "A":
            caption = f"{hook}\n\n✨ {intro} {product}\n💰 Harga: {price}\n✅ {benefit}\n{loc_line}{extra_line}{cta} {wa}\n{_pick(rng, hashtags)}"
        elif s == "B":
            caption = f"✨ {intro}\n{product}\nHarga {price}\nKeunggulan: {benefit}\n{loc_line}{extra_line}{cta} {wa}"
        else:
            caption = f"{hook}\n🆕 {product} — {price}\n✅ {benefit}\n{loc_line}{extra_line}📩 {cta} {wa}"

    elif template == "testi":
        testis = ["\"Respon cepat, prosesnya gampang.\"", "\"Hasilnya rapi, sesuai harapan.\"", "\"Worth it! Bakal order lagi.\""]
        testi = _pick(rng, testis)
        if s == "A":
            caption = f"{hook}\n\n⭐ Testimoni tentang {product}:\n{testi}\n\n💰 Mulai {price}\n✅ {benefit}\n{loc_line}{extra_line}{cta} {wa}"
        elif s == "B":
            caption = f"⭐ TESTIMONI\n{testi}\nProduk: {product}\nMulai: {price}\n{loc_line}{extra_line}{cta} {wa}"
        else:
            caption = f"{hook}\n⭐ {testi}\n📌 {product} • {price}\n✅ {benefit}\n{loc_line}{extra_line}📲 {cta} {wa}"

    else:
        rem = _pick(rng, ["Slot terbatas, amankan dulu ya.", "Bisa booking sekarang biar kebagian.", "Yang butuh cepat, ini waktunya!"])
        if s == "A":
            caption = f"{hook}\n\n⏰ Reminder: {product}\n💰 Mulai {price}\n✅ {benefit}\n{loc_line}{extra_line}📌 {rem}\n\n{cta} {wa}"
        elif s == "B":
            caption = f"⏰ REMINDER\n{product} (mulai {price})\n- {benefit}\n- {rem}\n{loc_line}{extra_line}{cta} {wa}"
        else:
            caption = f"{hook}\n⏰ {product} • {price}\n✅ {benefit}\n{loc_line}{extra_line}📌 {rem}\n📲 {cta} {wa}"

    return caption.strip(), hex(seed)[-6:]

# ==================== EMAIL ====================
def send_email(to_email, subject, body):
    host = (os.getenv("SMTP_HOST") or "").strip()
    port = int(os.getenv("SMTP_PORT") or "587")
    user = (os.getenv("SMTP_USER") or "").strip()
    passwd = (os.getenv("SMTP_PASS") or "").strip()
    mail_from = (os.getenv("SMTP_FROM") or user).strip()

    if not host or not user or not passwd:
        raise RuntimeError("SMTP belum dikonfigurasi.")

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    context = ssl.create_default_context()
    with smtplib.SMTP(host, port) as s:
        s.starttls(context=context)
        s.login(user, passwd)
        s.send_message(msg)

# ==================== ROUTES ====================

@app.route("/")
def landing():
    if is_logged_in():
        return redirect("/admin/dashboard" if session.get("role") == "admin" else "/dashboard")
    return render_template("landing.html")

@app.route("/db-check")
def db_check():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT 1 AS ok;")
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row

@app.route("/init-db")
def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(120) UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role VARCHAR(20) DEFAULT 'employee',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()
    conn.close()
    return "OK: tabel users siap."



@app.route("/init-products")
def init_products():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            name VARCHAR(120) NOT NULL,
            price INTEGER DEFAULT 0,
            is_global BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()
    conn.close()
    return "OK: tabel products siap."

@app.route("/init-hr-v2")
def init_hr_v2():
    ensure_hr_v2_schema()
    return "OK: HR v2 tables/columns ensured."

# ---------- AUTH ----------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "GET":
        return render_template("register.html", error=None)

    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    if not name or not email or not password:
        return render_template("register.html", error="Semua field wajib diisi.")

    pw_hash = generate_password_hash(password)

    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            "INSERT INTO users (name, email, password_hash, role) VALUES (%s, %s, %s, 'employee') RETURNING id;",
            (name, email, pw_hash),
        )
        user_id = (cur.fetchone() or {}).get("id")
        conn.commit()
        cur.close()
        conn.close()

        if not user_id:
            return render_template("register.html", error="Gagal membuat user (DB).")

        session["user_id"] = user_id
        session["user_name"] = name
        session["role"] = "employee"
        return redirect("/")
    except Exception:
        return render_template("register.html", error="Email sudah terdaftar / DB error.")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", error=None)

    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, name, email, password_hash, role FROM users WHERE email=%s;", (email,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if not user or not check_password_hash(user["password_hash"], password):
        return render_template("login.html", error="Email atau password salah.")

    session["user_id"] = user["id"]
    session["user_name"] = user["name"]
    session["role"] = user.get("role", "user")
    return redirect("/")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/login/google")
def login_google():
    if not (GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET):
        return "Google OAuth belum dikonfigurasi", 500
    redirect_uri = url_for("google_callback", _external=True)
    return oauth.google.authorize_redirect(redirect_uri)

@app.route("/auth/google/callback")
def google_callback():
    token = oauth.google.authorize_access_token()
    userinfo = token.get("userinfo") or oauth.google.get("https://openidconnect.googleapis.com/v1/userinfo").json()
    email = userinfo.get("email", "").lower()
    name = userinfo.get("name", "User")

    if not email:
        return "Email Google tidak ditemukan", 400

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, name, role FROM users WHERE lower(email)=%s LIMIT 1;", (email,))
    u = cur.fetchone()

    if not u:
        rand_pw = hashlib.sha256(f"{email}:{time.time()}".encode()).hexdigest()
        pw_hash = generate_password_hash(rand_pw)
        cur.execute(
            "INSERT INTO users (name, email, password_hash, role) VALUES (%s, %s, %s, 'employee') RETURNING id, name, role;",
            (name, email, pw_hash),
        )
        u = cur.fetchone()

    conn.commit()
    cur.close()
    conn.close()

    session.clear()
    session["user_id"] = u["id"]
    session["user_name"] = u["name"]
    session["role"] = u["role"]
    return redirect("/admin/dashboard" if u["role"] == "admin" else "/dashboard")

@app.route("/forgot", methods=["GET", "POST"])
def forgot_password():
    if request.method == "GET":
        return render_template("forgot_password.html")

    ensure_password_reset_schema()
    email = (request.form.get("email") or "").strip().lower()

    if not email:
        return "Email wajib diisi.", 400

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id FROM users WHERE lower(email)=%s LIMIT 1;", (email,))
    u = cur.fetchone()
    cur.close()
    conn.close()

    if not u:
        return render_template("forgot_password.html", sent=True)

    otp = f"{random.randint(0, 999999):06d}"
    otp_h = _otp_hash(email, otp)
    expires_at = datetime.utcnow() + timedelta(minutes=10)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE password_reset_otps SET used=TRUE WHERE email=%s AND used=FALSE;", (email,))
    cur.execute(
        "INSERT INTO password_reset_otps (email, otp_hash, expires_at, used) VALUES (%s, %s, %s, FALSE);",
        (email, otp_h, expires_at)
    )
    conn.commit()
    cur.close()
    conn.close()

    try:
        send_email(
            to_email=email,
            subject="UMGAP • Kode OTP Reset Password",
            body=f"Halo,\n\nKode OTP reset password kamu: {otp}\nBerlaku 10 menit.\n\nJika kamu tidak meminta reset, abaikan email ini.",
        )
    except Exception as e:
        return f"Gagal kirim email OTP. Error: {str(e)}", 500

    return render_template("forgot_password.html", sent=True, email=email)

@app.route("/reset", methods=["GET", "POST"])
def reset_password():
    if request.method == "GET":
        email = (request.args.get("email") or "").strip().lower()
        return render_template("reset_password.html", email=email)

    ensure_password_reset_schema()
    email = (request.form.get("email") or "").strip().lower()
    otp = (request.form.get("otp") or "").strip()
    new_password = (request.form.get("new_password") or "").strip()
    confirm = (request.form.get("confirm_password") or "").strip()

    if not email or not otp or not new_password:
        return "Email, OTP, dan password baru wajib diisi.", 400
    if new_password != confirm:
        return "Konfirmasi password tidak sama.", 400
    if len(new_password) < 6:
        return "Password minimal 6 karakter.", 400

    otp_h = _otp_hash(email, otp)
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT id, otp_hash, expires_at, used FROM password_reset_otps WHERE email=%s AND used=FALSE ORDER BY created_at DESC LIMIT 1;",
        (email,)
    )
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return "OTP tidak ditemukan / sudah dipakai.", 400

    if datetime.utcnow() > row["expires_at"]:
        cur.execute("UPDATE password_reset_otps SET used=TRUE WHERE id=%s;", (row["id"],))
        conn.commit()
        cur.close()
        conn.close()
        return "OTP sudah kedaluwarsa.", 400

    if not hmac.compare_digest(row["otp_hash"], otp_h):
        cur.close()
        conn.close()
        return "OTP salah.", 400

    pw_hash = generate_password_hash(new_password)
    cur.execute("UPDATE users SET password_hash=%s WHERE lower(email)=%s;", (pw_hash, email))
    cur.execute("UPDATE password_reset_otps SET used=TRUE WHERE id=%s;", (row["id"],))
    conn.commit()
    cur.close()
    conn.close()

    return redirect("/login")

# ---------- DASHBOARD ----------
@app.route("/dashboard")
def dashboard():
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        user_id = session.get("user_id")

        # ambil poin user
        cur.execute("SELECT points_admin, name FROM users WHERE id=%s LIMIT 1;", (user_id,))
        u = cur.fetchone() or {}

        # notifikasi / pengumuman yang belum dibaca user
        cur.execute("""
            SELECT a.id, a.title, a.message, a.created_at
            FROM announcements a
            LEFT JOIN announcement_reads ar
              ON ar.announcement_id = a.id
             AND ar.user_id = %s
            WHERE a.is_active = TRUE
              AND ar.id IS NULL
            ORDER BY a.created_at DESC
            LIMIT 20;
        """, (user_id,))
        announcements = cur.fetchall()

        notif_count = len(announcements)

    finally:
        cur.close()
        conn.close()

    return render_template(
        "dashboard.html",
        user_name=session.get("user_name"),
        points_admin=(u.get("points_admin") or 0),
        notif_count=notif_count,
        announcements=announcements
    )
# ---------- ADMIN ----------
@app.route("/admin")
def admin_home():
    admin_guard()
    return redirect("/admin/dashboard")

@app.route("/admin/dashboard")
def admin_dashboard():
    deny = admin_required()
    if deny:
        return deny

    if os.getenv('RUN_SCHEMA_ON_REQUEST','').lower()=='true':
        ensure_points_schema()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT COUNT(*) AS total FROM users WHERE role='employee';")
    total_employees = cur.fetchone()["total"]

    today = date.today()
    cur.execute("SELECT COUNT(*) AS total FROM attendance WHERE work_date=%s AND status='PRESENT';", (today,))
    total_attendance_today = cur.fetchone()["total"]

    cur.execute("SELECT COUNT(*) AS total FROM products;")
    total_products = cur.fetchone()["total"]

    cur.execute("SELECT id, name, email FROM users WHERE role='employee' ORDER BY name ASC;")
    employees = cur.fetchall()

    cur.close()
    conn.close()

    notif_count = get_notif_count()

    return render_template(
        "admin_dashboard.html",
        user_name=session.get("user_name", "Admin"),
        notif_count=int(notif_count or 0),
        total_employees=total_employees,
        total_attendance_today=total_attendance_today,
        total_products=total_products,
        employees=employees
    )

@app.route("/admin/users")
def admin_users():
    admin_guard()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT u.id, u.name, u.email, u.role, COALESCE(p.daily_salary, 0) AS daily_salary
            FROM users u
            LEFT JOIN payroll_settings p ON p.user_id=u.id
            ORDER BY u.id DESC;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    return render_template("admin_users.html", rows=rows, error=None)


@app.route("/admin/users/create", methods=["POST"])
def admin_users_create():
    admin_guard()
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    role = (request.form.get("role") or "employee").strip()
    daily_salary = int(request.form.get("daily_salary") or "0")

    if not name or not email or not password:
        return redirect("/admin/users")

    pw_hash = generate_password_hash(password)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            INSERT INTO users (name, email, password_hash, role)
            VALUES (%s, %s, %s, %s)
            RETURNING id;
        """, (name, email, pw_hash, role))
        row = cur.fetchone() or {}
        uid = row.get("id")

        if uid:
            cur.execute("""
                INSERT INTO payroll_settings (user_id, daily_salary)
                VALUES (%s, %s)
                ON CONFLICT (user_id)
                DO UPDATE SET daily_salary=EXCLUDED.daily_salary, updated_at=CURRENT_TIMESTAMP;
            """, (uid, daily_salary))

        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/users")


@app.route("/admin/users/update/<int:uid>", methods=["POST"])
def admin_users_update(uid):
    admin_guard()
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    role = (request.form.get("role") or "employee").strip()
    daily_salary = int(request.form.get("daily_salary") or "0")
    new_password = (request.form.get("new_password") or "").strip()

    conn = get_conn()
    cur = conn.cursor()
    try:
        if new_password:
            pw_hash = generate_password_hash(new_password)
            cur.execute("""
                UPDATE users
                SET name=%s, email=%s, role=%s, password_hash=%s
                WHERE id=%s;
            """, (name, email, role, pw_hash, uid))
        else:
            cur.execute("""
                UPDATE users
                SET name=%s, email=%s, role=%s
                WHERE id=%s;
            """, (name, email, role, uid))

        cur.execute("""
            INSERT INTO payroll_settings (user_id, daily_salary)
            VALUES (%s, %s)
            ON CONFLICT (user_id)
            DO UPDATE SET daily_salary=EXCLUDED.daily_salary, updated_at=CURRENT_TIMESTAMP;
        """, (uid, daily_salary))

        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/users")


@app.route("/admin/users/delete/<int:uid>", methods=["POST"])
def admin_users_delete(uid):
    if uid == session.get("user_id"):
        return redirect("/admin/users")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM users WHERE id=%s;", (uid,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/users")

@app.route("/admin/quick-attendance-links", methods=["GET","POST"])
def admin_quick_attendance_links():
    deny = admin_required()
    if deny:
        return deny

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()

            if action == "create":
                label = (request.form.get("label") or "").strip() or "Link Absensi"
                token = uuid.uuid4().hex

                cur.execute("""
                    INSERT INTO attendance_links (token, label, created_by, is_active)
                    VALUES (%s,%s,%s,TRUE)
                """, (token,label,session.get("user_id")))

                conn.commit()

            elif action == "toggle":
                link_id = int(request.form.get("id"))

                cur.execute("""
                    UPDATE attendance_links
                    SET is_active = NOT is_active
                    WHERE id=%s
                """,(link_id,))

                conn.commit()

            elif action == "delete":
                link_id = int(request.form.get("id"))
                cur.execute("DELETE FROM attendance_links WHERE id=%s", (link_id,))
                conn.commit()

        cur.execute("""
            SELECT id, token, label, created_at, is_active
            FROM attendance_links
            ORDER BY created_at DESC
            LIMIT 50
        """)

        links = cur.fetchall()

        base_url = request.host_url.rstrip("/")

        return render_template(
            "admin_quick_attendance_links.html",
            links=links,
            base_url=base_url
        )

    finally:
        cur.close()
        conn.close()

# ---------- ADMIN: APPROVAL ATTENDANCE (QUICK + USER LOGIN) ----------
@app.route("/admin/attendance-approval", methods=["GET"])
def admin_attendance_approval():
    deny = admin_required()
    if deny:
        return deny

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                id,
                name_input,
                user_id,
                work_date,
                arrival_type,
                note,
                device_id,
                latitude,
                longitude,
                accuracy,
                photo_path,
                created_at,
                created_at AS created_at_wib
            FROM attendance_pending
            WHERE status='PENDING'
            ORDER BY created_at DESC
            LIMIT 200;
        """)
        pendings = cur.fetchall()

        cur.execute("""
            SELECT id, name, email
            FROM users
            WHERE role='employee'
            ORDER BY name ASC;
        """)
        employees = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template(
        "admin_attendance_approval.html",
        pendings=pendings,
        employees=employees
    )


@app.route("/admin/attendance-approval/approve", methods=["POST"])
def admin_attendance_approve():
    deny = admin_required()
    if deny:
        return deny

    pending_id = request.form.get("pending_id")
    user_id_form = request.form.get("user_id")
    if not pending_id:
        return redirect("/admin/attendance-approval")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                id,
                name_input,
                user_id,
                work_date,
                arrival_type,
                note,
                device_id,
                latitude,
                longitude,
                accuracy,
                photo_path,
                created_at,
                created_at AS created_at_wib
            FROM attendance_pending
            WHERE id=%s AND status='PENDING'
            LIMIT 1;
        """, (int(pending_id),))
        p = cur.fetchone()

        if not p:
            return redirect("/admin/attendance-approval")

        target_user_id = p.get("user_id") or (int(user_id_form) if user_id_form else None)
        if not target_user_id:
            return redirect("/admin/attendance-approval")

        created_at_wib = p.get("created_at_wib") or p.get("created_at")
        work_date = p.get("work_date") or (created_at_wib.date() if created_at_wib else date.today())
        checkin_at = created_at_wib

        arrival_type = (p.get("arrival_type") or "ONTIME").strip().upper()

        if arrival_type in ("ONTIME", "LATE"):
            status = "PRESENT"
        elif arrival_type == "SICK":
            status = "SICK"
        elif arrival_type == "LEAVE":
            status = "LEAVE"
        elif arrival_type == "ABSENT":
            status = "ABSENT"
        else:
            status = "PRESENT"
            arrival_type = "ONTIME"

        latv = p.get("latitude")
        lngv = p.get("longitude")
        map_url = f"https://www.google.com/maps?q={latv},{lngv}" if (latv is not None and lngv is not None) else None
        clean_note = (p.get("note") or "").strip()

        cur.execute("""
            UPDATE attendance_pending
            SET status='APPROVED',
                approved_user_id=%s,
                approved_by=%s,
                approved_at=NOW()
            WHERE id=%s;
        """, (int(target_user_id), session.get("user_id"), int(pending_id)))

        cur.execute("""
            INSERT INTO attendance
                (user_id, work_date, status, arrival_type, note, checkin_at,
                 device_id, latitude, longitude, accuracy, photo_path, map_url)
            VALUES
                (%s, %s, %s, %s, %s, %s,
                 %s, %s, %s, %s, %s, %s)
            ON CONFLICT (user_id, work_date)
            DO UPDATE SET
                status=EXCLUDED.status,
                arrival_type=EXCLUDED.arrival_type,
                note=EXCLUDED.note,
                checkin_at=EXCLUDED.checkin_at,
                device_id=EXCLUDED.device_id,
                latitude=EXCLUDED.latitude,
                longitude=EXCLUDED.longitude,
                accuracy=EXCLUDED.accuracy,
                photo_path=EXCLUDED.photo_path,
                map_url=EXCLUDED.map_url;
        """, (
            int(target_user_id),
            work_date,
            status,
            arrival_type,
            clean_note,
            checkin_at,
            p.get("device_id"),
            latv,
            lngv,
            p.get("accuracy"),
            p.get("photo_path"),
            map_url
        ))

        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/attendance-approval")


@app.route("/admin/attendance-approval/reject", methods=["POST"])
def admin_attendance_reject():
    deny = admin_required()
    if deny:
        return deny

    pending_id = request.form.get("pending_id")
    reason = (request.form.get("reason") or "").strip()
    if not pending_id:
        return redirect("/admin/attendance-approval")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            UPDATE attendance_pending
            SET status='REJECTED',
                rejected_by=%s,
                rejected_at=NOW(),
                reject_reason=%s
            WHERE id=%s AND status='PENDING';
        """, (session.get("user_id"), reason, int(pending_id)))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/attendance-approval")

# ---------- ATTENDANCE ----------
@app.route("/attendance")
def attendance_page():
    if not is_logged_in():
        return redirect("/login")
    if is_admin():
        return redirect("/admin")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT work_date, arrival_type, status, note, checkin_at FROM attendance WHERE user_id=%s ORDER BY work_date DESC, checkin_at DESC NULLS LAST;", (session["user_id"],))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("attendance.html", rows=rows)

@app.route("/attendance/add", methods=["POST"])
def attendance_add():
    if not is_logged_in():
        return redirect("/login")

    arrival_type = (request.form.get("arrival_type") or "ONTIME").strip().upper()
    note = (request.form.get("note") or "").strip()
    now = _now_wib_naive_from_form()
    work_date = now.date()

    device_id = (request.form.get("device_id") or "").strip()
    lat = request.form.get("latitude")
    lng = request.form.get("longitude")
    acc = request.form.get("accuracy")

    def _to_float(x):
        try:
            return float(x) if x not in (None, "", "null") else None
        except:
            return None

    lat_f = _to_float(lat)
    lng_f = _to_float(lng)
    acc_f = _to_float(acc)

    # selfie upload
    photo = request.files.get("selfie")
    photo_path = None
    if photo and photo.filename:
        os.makedirs(os.path.join("static", "uploads", "attendance_user"), exist_ok=True)
        today_tag = date.today().strftime("%Y_%m_%d")
        filename = f"att_{today_tag}_{uuid.uuid4().hex}.jpg"
        save_path = os.path.join("static", "uploads", "attendance_user", filename)
        photo.save(save_path)
        photo_path = f"uploads/attendance_user/{filename}"

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        # masukkan ke pending (menunggu admin)
        submit_at = _now_wib_naive_from_form()

        cur.execute("""
            INSERT INTO attendance_pending
                (user_id, work_date, arrival_type, note,
                name_input, device_id, latitude, longitude, accuracy, photo_path,
                ip_address, status, created_at)
            VALUES
                (%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,
                %s,'PENDING', %s)
        """, (
            session.get("user_id"),
            work_date,
            arrival_type,
            note,
            session.get("user_name"),
            device_id,
            lat_f,
            lng_f,
            acc_f,
            photo_path,
            _public_ip(),
            submit_at
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/attendance")

@app.route("/admin/attendance")
def admin_attendance():
    r = admin_guard()
    if r:
        return r

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, name, email FROM users WHERE role='employee' ORDER BY name ASC;")
    employees = cur.fetchall()
    cur.execute("""
        SELECT a.work_date, a.arrival_type, a.status, a.note, a.checkin_at, u.name AS employee_name
        FROM attendance a
        JOIN users u ON u.id=a.user_id
        ORDER BY a.work_date DESC, a.checkin_at DESC NULLS LAST
        LIMIT 80;
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admin_attendance.html", employees=employees, rows=rows)

@app.route("/admin/attendance/add", methods=["POST"])
def admin_attendance_add():
    deny = admin_required()
    if deny:
        return deny

    user_id = int(request.form["user_id"])
    arrival_type = (request.form.get("arrival_type") or "ONTIME").strip().upper()
    note = (request.form.get("note") or "").strip()
    manual_checkin = (request.form.get("manual_checkin") or "").strip()

    if arrival_type in ("SICK", "LEAVE", "ABSENT"):
        status = arrival_type
    else:
        status = "PRESENT"

    if user_id == session.get("user_id"):
        now = _now_wib_naive_from_form()
    else:
        now = _parse_manual_wib_naive(manual_checkin) or _now_wib_naive_from_form()

    work_date = now.date()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id FROM attendance WHERE user_id=%s AND work_date=%s LIMIT 1;", (user_id, work_date))
    existing = cur.fetchone()

    if existing:
        cur.execute("UPDATE attendance SET status=%s, arrival_type=%s, note=%s, checkin_at=%s WHERE id=%s;",
            (status, arrival_type, note, now, existing["id"]))
    else:
        cur.execute("INSERT INTO attendance (user_id, work_date, status, arrival_type, note, created_at, checkin_at) VALUES (%s, %s, %s, %s, %s, %s, %s);",
            (user_id, work_date, status, arrival_type, note, now, now))

    conn.commit()
    cur.close()
    conn.close()
    return redirect("/admin/attendance")

# ---------- QUICK ATTENDANCE (PUBLIC, NO LOGIN) ----------
@app.route("/quick-attendance/<token>", methods=["GET"])
def quick_attendance_form(token):
    if not is_token_valid(token):
        return "Link absensi tidak valid / sudah nonaktif.", 404
    return render_template("quick_attendance.html", token=token)

@app.route("/quick-attendance/<token>/submit", methods=["POST"])
def quick_attendance_submit(token):
    if not is_token_valid(token):
        return "Link absensi tidak valid / sudah nonaktif.", 404

    name_input = (request.form.get("name_input") or "").strip()
    device_id = (request.form.get("device_id") or "").strip()
    lat = request.form.get("latitude")
    lng = request.form.get("longitude")
    acc = request.form.get("accuracy")

    if not name_input:
        return render_template("quick_attendance.html", token=token, error="Nama wajib diisi.")
    if not device_id:
        return render_template("quick_attendance.html", token=token, error="Device tidak terdeteksi. Coba refresh halaman.")

    # selfie file
    photo = request.files.get("selfie")
    if not photo or photo.filename == "":
        return render_template("quick_attendance.html", token=token, error="Selfie wajib diambil.")

    # bersihin foto lama (harian)
    cleanup_old_quick_attendance_photos()

    # simpan file
    _ensure_upload_dir()
    today_tag = date.today().strftime("%Y_%m_%d")
    filename = f"qa_{today_tag}_{uuid.uuid4().hex}.jpg"
    save_path = os.path.join(UPLOAD_QA_DIR, filename)
    photo.save(save_path)

    # path untuk ditampilkan via web
    photo_path = f"uploads/quick_attendance/{filename}"

    # parse angka (boleh kosong)
    def _to_float(x):
        try:
            return float(x) if x not in (None, "", "null") else None
        except:
            return None

    lat_f = _to_float(lat)
    lng_f = _to_float(lng)
    acc_f = _to_float(acc)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        submit_at = _now_wib_naive_from_form()

        cur.execute("""
            INSERT INTO attendance_pending
            (name_input, device_id, latitude, longitude, accuracy, photo_path, ip_address, status, created_at)
            VALUES
            (%s, %s, %s, %s, %s, %s, %s, 'PENDING', %s)
            RETURNING id;
        """, (
            name_input,
            device_id,
            lat_f,
            lng_f,
            acc_f,
            photo_path,
            _public_ip(),
            submit_at
        ))
        row = cur.fetchone() or {}
        conn.commit()
        pending_id = row.get("id")
    except psycopg2.IntegrityError:
        conn.rollback()
        # biasanya kena UNIQUE uq_pending_device_per_day
        return render_template(
            "quick_attendance.html",
            token=token,
            error="Perangkat ini sudah melakukan absensi hari ini."
        )
    finally:
        cur.close()
        conn.close()

    return render_template(
        "quick_attendance.html",
        token=token,
        success=f"Absensi terkirim (ID #{pending_id}). Menunggu approval admin."
    )

# ---------- PAYROLL ----------
def count_workdays_only_sunday_off(start_date, end_date):
    d = start_date
    n = 0
    while d < end_date:
        if d.weekday() != 6:
            n += 1
        d = d.fromordinal(d.toordinal() + 1)
    return n

@app.route("/admin/payroll")
def admin_payroll():
    deny = admin_required()
    if deny:
        return deny

    month = request.args.get("month")
    if not month:
        today = date.today()
        month = f"{today.year:04d}-{today.month:02d}"

    year = int(month.split("-")[0])
    mon = int(month.split("-")[1])
    start_date = date(year, mon, 1)
    end_date = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)
    WORKDAYS = count_workdays_only_sunday_off(start_date, end_date)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT u.id, u.name, COALESCE(p.daily_salary, 0) AS daily_salary, COALESCE(p.monthly_salary, 0) AS monthly_salary,
            COALESCE(SUM(CASE WHEN a.status='PRESENT' THEN 1 ELSE 0 END), 0) AS days_present,
            COALESCE(SUM(CASE WHEN a.status='SICK' THEN 1 ELSE 0 END), 0) AS days_sick,
            COALESCE(SUM(CASE WHEN a.status='LEAVE' THEN 1 ELSE 0 END), 0) AS days_leave,
            COALESCE(SUM(CASE WHEN a.status='ABSENT' THEN 1 ELSE 0 END), 0) AS days_absent
        FROM users u
        LEFT JOIN payroll_settings p ON p.user_id = u.id
        LEFT JOIN attendance a ON a.user_id = u.id AND a.work_date >= %s AND a.work_date < %s
        WHERE u.role = 'employee'
        GROUP BY u.id, u.name, p.daily_salary, p.monthly_salary
        ORDER BY u.name ASC;
    """, (start_date, end_date))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    result = []
    for r in rows:
        daily_salary = int(r.get("daily_salary") or 0)
        monthly_salary = int(r.get("monthly_salary") or 0)
        if daily_salary == 0 and monthly_salary > 0 and WORKDAYS > 0:
            daily_salary = int(round(monthly_salary / WORKDAYS))
        days_present = int(r.get("days_present") or 0)
        result.append({
            "id": r["id"], "name": r["name"], "daily_salary": daily_salary, "workdays": int(WORKDAYS),
            "days_present": days_present, "days_sick": int(r.get("days_sick") or 0),
            "days_leave": int(r.get("days_leave") or 0), "days_absent": int(r.get("days_absent") or 0),
            "salary_paid": int(daily_salary * days_present),
        })

    return render_template("admin_payroll.html", month=month, rows=result, workdays=int(WORKDAYS))

# ---------- SALES ----------
@app.route("/sales", methods=["GET", "POST"])
def sales_user():
    if not is_logged_in():
        return redirect("/login")
    if session.get("role") == "admin":
        return redirect("/admin/sales")

    if request.method == "POST":
        product_id = request.form.get("product_id")
        qty = request.form.get("qty") or "0"
        note = (request.form.get("note") or "").strip()

        try:
            product_id = int(product_id)
            qty_int = int(qty)
        except Exception:
            return redirect("/sales")

        if qty_int <= 0:
            return redirect("/sales")

        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            # validasi produk global
            cur.execute("""
                SELECT id, name
                FROM products
                WHERE id=%s AND is_global=TRUE
                LIMIT 1;
            """, (product_id,))
            ok = cur.fetchone()

            if not ok:
                return redirect("/sales")

            cur.execute("""
                INSERT INTO sales_submissions (user_id, product_id, qty, note, status)
                VALUES (%s, %s, %s, %s, 'PENDING');
            """, (session["user_id"], product_id, qty_int, note))
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return redirect("/sales")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        # PRODUK untuk dropdown
        cur.execute("""
            SELECT id, name, price
            FROM products
            WHERE is_global=TRUE
            ORDER BY id DESC;
        """)
        products = cur.fetchall()

        # RIWAYAT submit user
        cur.execute("""
            SELECT
                s.id,
                s.qty,
                s.note,
                s.status,
                s.admin_note,
                s.created_at,
                (s.created_at + interval '7 hour') AS created_at_wib,
                COALESCE(p.name, '-') AS product_name
            FROM sales_submissions s
            LEFT JOIN products p ON p.id = s.product_id
            WHERE s.user_id=%s
            ORDER BY s.id DESC
            LIMIT 50;
        """, (session["user_id"],))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("sales.html", products=products, rows=rows)


@app.route("/admin/sales")
def admin_sales():
    admin_guard()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                s.id,
                s.qty,
                s.note,
                s.status,
                s.admin_note,
                s.created_at,
                (s.created_at + interval '7 hour') AS created_at_wib,
                u.name AS employee_name,
                COALESCE(p.name, '-') AS product_name
            FROM sales_submissions s
            JOIN users u ON u.id = s.user_id
            LEFT JOIN products p ON p.id = s.product_id
            ORDER BY
                (CASE WHEN s.status='PENDING' THEN 0 ELSE 1 END),
                s.created_at DESC
            LIMIT 300;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_sales.html", rows=rows)


@app.route("/admin/sales/approve/<int:sid>", methods=["POST"])
def admin_sales_approve(sid):
    admin_guard()
    admin_note = (request.form.get("admin_note") or "").strip()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE sales_submissions
            SET status='APPROVED',
                admin_note=%s,
                decided_at=CURRENT_TIMESTAMP,
                decided_by=%s
            WHERE id=%s;
        """, (admin_note, session["user_id"], sid))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/sales")


@app.route("/admin/sales/reject/<int:sid>", methods=["POST"])
def admin_sales_reject(sid):
    admin_guard()
    admin_note = (request.form.get("admin_note") or "").strip()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE sales_submissions
            SET status='REJECTED',
                admin_note=%s,
                decided_at=CURRENT_TIMESTAMP,
                decided_by=%s
            WHERE id=%s;
        """, (admin_note, session["user_id"], sid))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/admin/sales")


@app.route("/admin/sales/monitor")
def admin_sales_monitor():
    admin_guard()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                u.id,
                u.name AS employee_name,
                COALESCE(SUM(s.qty), 0) AS total_qty
            FROM users u
            LEFT JOIN sales_submissions s
                ON s.user_id = u.id
               AND s.status='APPROVED'
            WHERE u.role='employee'
            GROUP BY u.id, u.name
            ORDER BY total_qty DESC, u.name ASC;
        """)
        summary = cur.fetchall()

        cur.execute("""
            SELECT
                s.created_at,
                (s.created_at + interval '7 hour') AS created_at_wib,
                u.name AS employee_name,
                COALESCE(p.name, '-') AS product_name,
                s.qty,
                s.status,
                s.note,
                s.admin_note
            FROM sales_submissions s
            JOIN users u ON u.id = s.user_id
            LEFT JOIN products p ON p.id = s.product_id
            ORDER BY s.created_at DESC
            LIMIT 200;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_sales_monitor.html", summary=summary, rows=rows)

# ---------- STATS ----------
@app.route("/admin/stats")
def admin_stats():
    admin_guard()

    month = request.args.get("month")
    if not month:
        today = date.today()
        month = f"{today.year:04d}-{today.month:02d}"

    year = int(month.split("-")[0])
    mon = int(month.split("-")[1])
    start_date = date(year, mon, 1)
    end_date = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)  # <<< FIX DI SINI
    try:
        cur.execute("""
            SELECT u.id, u.name AS employee_name,
                COALESCE(SUM(CASE WHEN a.status='PRESENT' THEN 1 ELSE 0 END), 0) AS present_days,
                COALESCE(SUM(CASE WHEN a.arrival_type='LATE' THEN 1 ELSE 0 END), 0) AS late_days,
                COALESCE(SUM(CASE WHEN a.status='SICK' THEN 1 ELSE 0 END), 0) AS sick_days,
                COALESCE(SUM(CASE WHEN a.status='LEAVE' THEN 1 ELSE 0 END), 0) AS leave_days,
                COALESCE(SUM(CASE WHEN a.status='ABSENT' THEN 1 ELSE 0 END), 0) AS absent_days
            FROM users u
            LEFT JOIN attendance a
                ON a.user_id=u.id
                AND a.work_date >= %s AND a.work_date < %s
            WHERE u.role='employee'
            GROUP BY u.id, u.name
            ORDER BY u.name ASC;
        """, (start_date, end_date))
        att = cur.fetchall()

        cur.execute("""
            SELECT u.id, COALESCE(SUM(s.qty), 0) AS sales_qty
            FROM users u
            LEFT JOIN sales_submissions s
                ON s.user_id=u.id
                AND s.created_at >= %s AND s.created_at < %s
            WHERE u.role='employee'
            GROUP BY u.id;
        """, (start_date, end_date))
        sales = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    sales_map = {r["id"]: int(r["sales_qty"] or 0) for r in sales}

    rows = []
    totals = {"present": 0, "late": 0, "sick": 0, "leave": 0, "absent": 0, "sales": 0}

    for r in att:
        row = {
            "employee_name": r["employee_name"],
            "present_days": int(r["present_days"] or 0),
            "late_days": int(r["late_days"] or 0),
            "sick_days": int(r["sick_days"] or 0),
            "leave_days": int(r["leave_days"] or 0),
            "absent_days": int(r["absent_days"] or 0),
            "sales_qty": sales_map.get(r["id"], 0),
        }
        totals["present"] += row["present_days"]
        totals["late"] += row["late_days"]
        totals["sick"] += row["sick_days"]
        totals["leave"] += row["leave_days"]
        totals["absent"] += row["absent_days"]
        totals["sales"] += row["sales_qty"]
        rows.append(row)

    return render_template("admin_stats.html", month=month, rows=rows, totals=totals)

# --------- INVOICE ---------
@app.route("/invoice/new", methods=["GET", "POST"])
def invoice_new_user():
    if not is_logged_in():
        return redirect("/login")
    if session.get("role") == "admin":
        return redirect("/admin/invoice/new")

    ensure_invoice_schema()

    if request.method == "POST":
        return _save_invoice(is_admin_mode=False)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, price
            FROM products
            WHERE is_global=TRUE
            ORDER BY name ASC;
        """)
        products = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template(
        "invoice_form.html",
        products=products,
        is_admin_mode=False
    )


@app.route("/admin/invoice/new", methods=["GET", "POST"])
def invoice_new_admin():
    deny = admin_required()
    if deny:
        return deny

    ensure_invoice_schema()

    if request.method == "POST":
        return _save_invoice(is_admin_mode=True)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, price
            FROM products
            WHERE is_global=TRUE
            ORDER BY name ASC;
        """)
        products = cur.fetchall()

        cur.execute("""
            SELECT id, name, email
            FROM users
            WHERE role='employee'
            ORDER BY name ASC;
        """)
        employees = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template(
        "invoice_form.html",
        products=products,
        employees=employees,
        is_admin_mode=True
    )

@app.route("/invoice/<int:invoice_id>")
def invoice_view(invoice_id):
    if not is_logged_in():
        return redirect("/login")

    ensure_invoice_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT i.*, u.name AS created_by_name
            FROM invoices i
            JOIN users u ON u.id = i.created_by
            WHERE i.id=%s
            LIMIT 1;
        """, (invoice_id,))
        invoice = cur.fetchone()
        if not invoice:
            abort(404)

        cur.execute("""
            SELECT id, product_id, product_name, qty, price, subtotal
            FROM invoice_items
            WHERE invoice_id=%s
            ORDER BY id ASC;
        """, (invoice_id,))
        items = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    invoice["created_at_wib"] = _utc_naive_to_wib_string(invoice.get("created_at"))
    invoice["paid_at_wib"] = _utc_naive_to_wib_string(invoice.get("paid_at")) if invoice.get("paid_at") else None
    invoice.setdefault("is_paid", True)
    invoice.setdefault("discount", 0)
    invoice.setdefault("customer_phone", "")
    invoice.setdefault("company_name", "")
    invoice.setdefault("company_logo_path", None)

    return render_template("invoice_print.html", invoice=invoice, items=items)


@app.route("/invoice/<int:invoice_id>/json")
def invoice_json(invoice_id):
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    ensure_invoice_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT *
            FROM invoices
            WHERE id=%s
            LIMIT 1;
        """, (invoice_id,))
        invoice = cur.fetchone()
        if not invoice:
            return jsonify({"ok": False, "error": "Invoice tidak ditemukan"}), 404

        cur.execute("""
            SELECT product_name, qty, price, subtotal
            FROM invoice_items
            WHERE invoice_id=%s
            ORDER BY id ASC;
        """, (invoice_id,))
        items = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return jsonify({
        "ok": True,
        "invoice": invoice,
        "items": items
    })


@app.route("/invoice/<int:invoice_id>/mark-paid", methods=["POST"])
def invoice_mark_paid(invoice_id):
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    ensure_invoice_schema()

    data = request.get_json(silent=True) or {}
    is_paid = bool(data.get("is_paid"))
    paid_at = datetime.utcnow() if is_paid else None

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            UPDATE invoices
            SET is_paid=%s,
                paid_at=%s
            WHERE id=%s
            RETURNING id, is_paid, paid_at;
        """, (is_paid, paid_at, invoice_id))
        row = cur.fetchone()
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not row:
        return jsonify({"ok": False, "error": "Invoice tidak ditemukan"}), 404

    return jsonify({
        "ok": True,
        "invoice_id": row.get("id"),
        "is_paid": bool(row.get("is_paid")),
        "paid_at_wib": _utc_naive_to_wib_string(row.get("paid_at")) if row.get("paid_at") else None
    })

@app.route("/invoice/<int:invoice_id>/pdf")
def invoice_pdf(invoice_id):
    if not is_logged_in():
        return redirect("/login")

    ensure_invoice_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT i.*, u.name AS created_by_name
            FROM invoices i
            JOIN users u ON u.id = i.created_by
            WHERE i.id=%s
            LIMIT 1;
        """, (invoice_id,))
        invoice = cur.fetchone()
        if not invoice:
            abort(404)

        cur.execute("""
            SELECT id, product_id, product_name, qty, price, subtotal
            FROM invoice_items
            WHERE invoice_id=%s
            ORDER BY id ASC;
        """, (invoice_id,))
        items = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    invoice["created_at_wib"] = _utc_naive_to_wib_string(invoice.get("created_at"))
    invoice["paid_at_wib"]    = _utc_naive_to_wib_string(invoice.get("paid_at")) if invoice.get("paid_at") else None
    invoice.setdefault("is_paid", True)
    invoice.setdefault("discount", 0)
    invoice.setdefault("customer_phone", "")
    invoice.setdefault("company_name", "")
    invoice.setdefault("company_logo_path", None)

    html = render_template("invoice_pdf.html", invoice=invoice, items=items)

    filename = f"{invoice['invoice_no']}.pdf"

    # 1) coba WeasyPrint dulu
    try:
        from weasyprint import HTML

        pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e_weasy:
        # 2) fallback ke xhtml2pdf
        try:
            from xhtml2pdf import pisa

            pdf_io = io.BytesIO()
            pisa_status = pisa.CreatePDF(src=html, dest=pdf_io)
            if pisa_status.err:
                return (
                    f"Gagal membuat PDF. WeasyPrint error: {e_weasy} | xhtml2pdf juga gagal.",
                    500,
                )

            pdf_io.seek(0)
            return send_file(
                pdf_io,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )
        except Exception as e_xhtml:
            return (
                "Gagal membuat PDF.<br>"
                f"WeasyPrint error: {e_weasy}<br>"
                f"xhtml2pdf error: {e_xhtml}",
                500,
            )

# ---------- PRODUCTS ----------
@app.route("/products")
def products():
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, price, user_id, is_global
            FROM products
            ORDER BY id DESC;
        """)
        rows = cur.fetchall()
        return render_template("products.html", products=rows, error=None)
    finally:
        cur.close()
        conn.close()


@app.route("/products/add", methods=["POST"])
def products_add():
    if not is_logged_in():
        return redirect("/login")

    name = (request.form.get("name") or "").strip()
    price = (request.form.get("price") or "0").strip()
    if not name:
        return redirect("/products")

    try:
        price_int = int(price)
        if price_int < 0:
            price_int = 0
    except ValueError:
        price_int = 0

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO products (user_id, name, price, is_global)
            VALUES (%s, %s, %s, TRUE);
        """, (session["user_id"], name, price_int))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/products")


@app.route("/products/delete/<int:pid>")
def products_delete(pid):
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM products WHERE id=%s;", (pid,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/products")


@app.route("/products/edit/<int:pid>", methods=["GET", "POST"])
def products_edit(pid):
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, price, user_id, is_global
            FROM products
            WHERE id=%s
            LIMIT 1;
        """, (pid,))
        product = cur.fetchone()
        if not product:
            abort(404)

        if request.method == "GET":
            return render_template("product_edit.html", product=product, error=None)

        name = (request.form.get("name") or "").strip()
        price = (request.form.get("price") or "0").strip()
        if not name:
            return render_template("product_edit.html", product=product, error="Nama produk wajib diisi.")

        try:
            price_int = int(price)
            if price_int < 0:
                price_int = 0
        except ValueError:
            price_int = 0

        # update
        cur2 = conn.cursor()
        try:
            cur2.execute("UPDATE products SET name=%s, price=%s WHERE id=%s;", (name, price_int, pid))
            conn.commit()
        finally:
            cur2.close()

        return redirect("/products")
    finally:
        cur.close()
        conn.close()

# ---------- CONTENT ----------
@app.route("/init-content")
def init_content():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS content_plans (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            plan_date DATE NOT NULL,
            platform VARCHAR(30) NOT NULL,
            content_type VARCHAR(30) NOT NULL,
            notes TEXT,
            is_done BOOLEAN NOT NULL DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()
    conn.close()
    return "OK: tabel content_plans siap."

@app.route("/content", methods=["GET", "POST"])
def content():
    if not is_logged_in():
        return redirect("/login")

    user_id = session.get("user_id")

    if request.method == "POST":
        plan_date = (request.form.get("plan_date") or "").strip()
        platform = (request.form.get("platform") or "").strip()
        content_type = (request.form.get("content_type") or "").strip()
        notes = (request.form.get("notes") or "").strip()

        if plan_date and platform and content_type:
            conn = get_conn()
            cur = conn.cursor()
            try:
                cur.execute("""
                    INSERT INTO content_plans (user_id, plan_date, platform, content_type, notes, is_done)
                    VALUES (%s, %s, %s, %s, %s, FALSE);
                """, (user_id, plan_date, platform, content_type, notes))
                conn.commit()
            finally:
                cur.close()
                conn.close()

        return redirect("/content")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                id,
                user_id,
                plan_date,
                platform,
                content_type,
                notes,
                is_done,
                created_at
            FROM content_plans
            WHERE user_id = %s
            ORDER BY
                is_done ASC,
                plan_date ASC,
                id DESC;
        """, (user_id,))
        plans = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("content.html", plans=plans)

@app.route("/content/add", methods=["POST"])
def content_add():
    if not is_logged_in():
        return redirect("/login")
    plan_date = request.form.get("plan_date")
    platform = (request.form.get("platform") or "").strip()
    content_type = (request.form.get("content_type") or "").strip()
    notes = (request.form.get("notes") or "").strip()
    if not plan_date or not platform or not content_type:
        return redirect("/content")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO content_plans (user_id, plan_date, platform, content_type, notes) VALUES (%s, %s, %s, %s, %s);",
        (session["user_id"], plan_date, platform, content_type, notes))
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/content")

@app.route("/content/done/<int:cid>")
def content_done(cid):
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE content_plans
            SET is_done = TRUE
            WHERE id = %s AND user_id = %s;
        """, (cid, session.get("user_id")))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/content")


@app.route("/content/undo/<int:cid>")
def content_undo(cid):
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE content_plans
            SET is_done = FALSE
            WHERE id = %s AND user_id = %s;
        """, (cid, session.get("user_id")))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/content")


@app.route("/content/delete/<int:cid>")
def content_delete(cid):
    if not is_logged_in():
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            DELETE FROM content_plans
            WHERE id = %s AND user_id = %s;
        """, (cid, session.get("user_id")))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return redirect("/content")

# ---------- CAPTION ----------
@app.route("/caption", methods=["GET", "POST"])
def caption():
    if not is_logged_in():
        return redirect("/login")

    form = {"template": "promo", "biz_type": "produk", "tone": "santai", "product": "", "price": "", "wa": "", "location": "", "extra": ""}
    caption_text = None

    if request.method == "POST":
        form["template"] = request.form.get("template", "promo")
        form["biz_type"] = request.form.get("biz_type", "produk")
        form["tone"] = request.form.get("tone", "santai")
        form["product"] = request.form.get("product", "").strip()
        form["price"] = request.form.get("price", "").strip()
        form["wa"] = request.form.get("wa", "").strip()
        form["location"] = request.form.get("location", "").strip()
        form["extra"] = request.form.get("extra", "").strip()

        nama = form["product"]
        harga = rupiah(form["price"])
        loc = f"\n📍 Lokasi: {form['location']}" if form["location"] else ""
        extra = f"\nℹ️ Catatan: {form['extra']}" if form["extra"] else ""

        benefit = pick(["Kualitas terjaga", "Cocok untuk kebutuhan harian", "Praktis dan mudah digunakan", "Bisa untuk hadiah"])
        hook = pick(["Lagi cari yang pas buat kamu?", "Biar makin gampang, cek ini dulu 👇", "Yang ini lagi banyak dicari loh!"])
        cta = pick(["Chat aja ya 👉", "Langsung DM/WA ya 👉", "Pesan sekarang 👉"])

        if form["template"] == "promo":
            caption_text = f"{hook}\n🎯 {nama}\n💰 Mulai {harga}\n✅ {benefit}\n📌 Harga spesial{loc}{extra}\n\n{cta} {form['wa']}"
        elif form["template"] == "new":
            caption_text = f"{hook}\n✨ Rilis: {nama}\n💰 Harga: {harga}\n✅ {benefit}{loc}{extra}\n\n{cta} {form['wa']}"
        elif form["template"] == "testi":
            testi = pick(["\"Pelayanannya cepat dan responsif.\"", "\"Hasilnya sesuai ekspektasi, recommended!\"", "\"Worth it!\""])
            caption_text = f"{hook}\n⭐ Testimoni: {testi}\n💰 Mulai {harga}\n✅ {benefit}{loc}{extra}\n\n{cta} {form['wa']}"
        else:
            caption_text = f"{hook}\n⏰ Reminder: {nama}\n💰 Mulai {harga}\n✅ {benefit}\n📌 Slot terbatas{loc}{extra}\n\n{cta} {form['wa']}"

    return render_template("caption.html", caption=caption_text, form=form)

# ==================== AI PENGHITUNG HPP (v2 — matang) ====================

import json as _json

def _parse_materials(materials_raw):
    """Parse & validasi list bahan. Return (list[dict], total_material_cost)."""
    if isinstance(materials_raw, str):
        try:
            materials_raw = _json.loads(materials_raw)
        except Exception:
            materials_raw = []
    if not isinstance(materials_raw, list):
        materials_raw = []

    rows = []
    total = 0
    for m in materials_raw:
        name = str(m.get("name") or "").strip()
        unit = str(m.get("unit") or "").strip()
        try:
            cost = int(m.get("cost") or 0)
        except Exception:
            cost = 0
        if cost < 0:
            cost = 0
        total += cost
        rows.append({"name": name, "unit": unit, "cost": cost})
    return rows, total


def _hpp_calculate(product_name, materials_raw, labor_cost_raw, overhead_cost_raw, output_qty_raw):
    """Hitung HPP. Kembalikan dict result lengkap."""
    try:
        labor_cost = max(0, int(labor_cost_raw or 0))
    except Exception:
        labor_cost = 0
    try:
        overhead_cost = max(0, int(overhead_cost_raw or 0))
    except Exception:
        overhead_cost = 0
    try:
        output_qty = max(1, int(output_qty_raw or 1))
    except Exception:
        output_qty = 1

    material_rows, total_material_cost = _parse_materials(materials_raw)
    total_cost = total_material_cost + labor_cost + overhead_cost
    hpp_per_unit = round(total_cost / output_qty) if output_qty > 0 else 0

    # Saran harga jual (margin 20%, 40%, 60%)
    price_suggestions = {
        "safe":   round(hpp_per_unit * 1.20),
        "normal": round(hpp_per_unit * 1.40),
        "ideal":  round(hpp_per_unit * 1.60),
    }

    return {
        "product_name":        product_name,
        "materials":           material_rows,
        "total_material_cost": total_material_cost,
        "labor_cost":          labor_cost,
        "overhead_cost":       overhead_cost,
        "output_qty":          output_qty,
        "total_cost":          total_cost,
        "hpp_per_unit":        hpp_per_unit,
        "price_suggestions":   price_suggestions,
    }


def _hpp_ai_prompt(data):
    """Bangun prompt AI review HPP yang kaya konteks."""
    mat_lines = "\n".join(
        ("  - " + m["name"] + " (" + (m["unit"] or "-") + "): Rp " + str(m["cost"]))
        for m in data["materials"]
    ) or "  (tidak ada bahan yang diinput)"

    return f"""Kamu adalah konsultan keuangan UMKM Indonesia yang ahli dalam menghitung HPP (Harga Pokok Produksi).

Data HPP yang disubmit user:
Produk: "{data['product_name']}"
Bahan baku:
{mat_lines}
Biaya tenaga kerja (per batch): Rp {data['labor_cost']:,}
Biaya overhead (per batch): Rp {data['overhead_cost']:,}
Jumlah produk jadi: {data['output_qty']} unit
Total biaya produksi: Rp {data['total_cost']:,}
HPP per unit: Rp {data['hpp_per_unit']:,}

Tugasmu:
1. Analisis kelengkapan dan keakuratan input.
2. Sebutkan biaya yang mungkin terlewat (contoh: kemasan, ongkir bahan, penyusutan alat, biaya air/listrik, dsb.).
3. Evaluasi apakah HPP per unit terlihat realistis untuk jenis produk tersebut.
4. Berikan saran harga jual yang optimal beserta alasannya.
5. Jika ada potensi efisiensi biaya, sebutkan.

Format jawaban WAJIB:
- Analisis: [2-3 kalimat evaluasi keseluruhan]
- Yang kurang spesifik: [list poin dengan bullet]
- Biaya yang mungkin terlewat: [list poin]
- Saran: [rekomendasi konkret, termasuk range harga jual dan alasannya]

Jawab dalam Bahasa Indonesia yang ramah dan mudah dipahami pelaku UMKM. Maksimal 350 kata.""".replace(",", ",")


@app.route("/hpp-ai", methods=["GET", "POST"])
def hpp_ai_page():
    if not is_logged_in():
        return redirect("/login")

    result = None
    ai_notes = None
    form_data = {
        "product_name": "",
        "labor_cost":   "",
        "overhead_cost":"",
        "output_qty":   "",
        "materials_json": "[]",
    }

    if request.method == "POST":
        product_name   = (request.form.get("product_name")   or "").strip()
        labor_cost_raw = (request.form.get("labor_cost")      or "0").strip()
        overhead_raw   = (request.form.get("overhead_cost")   or "0").strip()
        qty_raw        = (request.form.get("output_qty")      or "1").strip()
        mats_json      = (request.form.get("materials_json")  or "[]").strip()

        form_data = {
            "product_name":  product_name,
            "labor_cost":    labor_cost_raw,
            "overhead_cost": overhead_raw,
            "output_qty":    qty_raw,
            "materials_json": mats_json,
        }

        if not product_name:
            return render_template("hpp_ai.html", result=None, ai_notes=None,
                                   form_data=form_data, error="Nama produk wajib diisi.")

        result = _hpp_calculate(product_name, mats_json, labor_cost_raw, overhead_raw, qty_raw)

        if oa_client:
            try:
                prompt = _hpp_ai_prompt(result)
                resp = oa_client.chat.completions.create(
                    model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.35,
                    max_tokens=500,
                )
                ai_notes = (resp.choices[0].message.content or "").strip()
            except Exception as ex:
                ai_notes = f"AI tidak tersedia saat ini ({type(ex).__name__}). Hasil HPP tetap valid."

    return render_template(
        "hpp_ai.html",
        result=result,
        ai_notes=ai_notes,
        form_data=form_data,
    )


@app.route("/api/hpp-calculate", methods=["POST"])
def api_hpp_calculate():
    """Pure server-side HPP calculation — no AI, fast, untuk validasi JS."""
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Silakan login dulu."}), 401

    data = request.get_json(silent=True) or {}
    try:
        result = _hpp_calculate(
            product_name   = (data.get("product_name") or "").strip(),
            materials_raw  = data.get("materials") or [],
            labor_cost_raw = data.get("labor_cost") or 0,
            overhead_cost_raw = data.get("overhead_cost") or 0,
            output_qty_raw = data.get("output_qty") or 1,
        )
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hpp-ai-review", methods=["POST"])
def api_hpp_ai_review():
    """AI review endpoint — dipanggil async dari frontend setelah kalkulasi."""
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Silakan login dulu."}), 401

    data = request.get_json(silent=True) or {}
    product_name  = (data.get("product_name") or "").strip()
    materials     = data.get("materials") or []
    labor_cost    = int(data.get("labor_cost")    or 0)
    overhead_cost = int(data.get("overhead_cost") or 0)
    output_qty    = max(1, int(data.get("output_qty") or 1))
    total_cost    = int(data.get("total_cost")    or 0)
    hpp_per_unit  = int(data.get("hpp_per_unit")  or 0)

    if not product_name:
        return jsonify({"ok": False, "error": "Nama produk wajib diisi."}), 400
    if not oa_client:
        return jsonify({"ok": False, "error": "AI belum dikonfigurasi (OPENAI_API_KEY missing)."}), 500

    calc_data = {
        "product_name": product_name,
        "materials":    [{"name": str(m.get("name","")).strip(),
                          "unit": str(m.get("unit","")).strip(),
                          "cost": int(m.get("cost") or 0)} for m in materials],
        "labor_cost":   labor_cost,
        "overhead_cost":overhead_cost,
        "output_qty":   output_qty,
        "total_cost":   total_cost or (sum(int(m.get("cost",0)) for m in materials) + labor_cost + overhead_cost),
        "hpp_per_unit": hpp_per_unit,
    }

    try:
        prompt = _hpp_ai_prompt(calc_data)
        resp = oa_client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.35,
            max_tokens=500,
        )
        review = (resp.choices[0].message.content or "").strip()
        return jsonify({"ok": True, "review": review})
    except Exception as e:
        return jsonify({"ok": False, "error": f"AI gagal: {type(e).__name__}: {str(e)}"}), 500

#-----NOTIFICATION-------
@app.route("/notifications")
def notifications():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, title, message, created_at
        FROM announcements
        WHERE is_active = TRUE
        ORDER BY created_at DESC
        LIMIT 50;
    """)
    announcements = cur.fetchall()
    cur.close()
    conn.close()

    return render_template("notifications.html", announcements=announcements)

@app.route("/notifications/read/<int:ann_id>")
def mark_notification_read(ann_id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO announcement_reads (announcement_id, user_id)
        VALUES (%s, %s)
        ON CONFLICT DO NOTHING
    """, (ann_id, session["user_id"]))

    conn.commit()
    conn.close()

    return redirect("/notifications")

@app.route("/admin/announcements")
def admin_announcements():
    if session.get("role") != "admin":
        return abort(403)

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT * FROM announcements ORDER BY created_at DESC")
    data = cur.fetchall()

    conn.close()

    return render_template("admin_announcements.html", data=data)

@app.route("/admin/announcements/add", methods=["POST"])
def add_announcement():
    if session.get("role") != "admin":
        return abort(403)

    title = request.form["title"]
    message = request.form["message"]

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO announcements (title, message, created_by)
        VALUES (%s, %s, %s)
    """, (title, message, session["user_id"]))

    conn.commit()
    conn.close()

    return redirect("/admin/announcements")

@app.route("/admin/announcements/delete/<int:id>")
def delete_announcement(id):
    if session.get("role") != "admin":
        return abort(403)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM announcements WHERE id=%s", (id,))
    conn.commit()
    conn.close()

    return redirect("/admin/announcements")

# ---------- AI ----------
@app.route("/ai-test")
def ai_test():
    try:
        r = client.responses.create(model="gpt-4.1-mini", input="Buatkan caption jualan kopi susu yang santai")
        return r.output_text
    except Exception as e:
        return f"ERROR: {e}"

@app.route("/caption/ai", methods=["POST"])
def caption_ai():
    if not is_logged_in():
        return redirect("/login")
    product = request.form.get("product")
    price = request.form.get("price", "")
    style = request.form.get("style", "Santai")
    try:
        caption = generate_caption_ai(product, price, style)
    except Exception:
        caption = "⚠️ AI sedang sibuk, coba lagi sebentar."
    return render_template("caption.html", ai_result=caption, product=product, price=price, style=style)

@app.route("/api/caption-ai", methods=["POST"])
def api_caption_ai():
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Silakan login dulu."}), 401
    product = (request.form.get("product") or "").strip()
    price = (request.form.get("price") or "").strip()
    style = (request.form.get("style") or "Santai").strip()
    brand = (request.form.get("brand") or "").strip()
    platform = (request.form.get("platform") or "Instagram").strip()
    notes = (request.form.get("notes") or "").strip()

    if not product:
        return jsonify({"ok": False, "error": "Nama produk wajib diisi."}), 400
    try:
        caption = generate_caption_ai(product, price, style, brand=brand, platform=platform, notes=notes)
        return jsonify({"ok": True, "caption": caption})
    except Exception:
        return jsonify({"ok": False, "error": "AI sedang sibuk."}), 500

@app.route("/api/chat", methods=["POST"])
def api_chat():
    data = request.get_json(silent=True) or {}
    msg = (data.get("message") or "").strip()
    if not msg:
        return jsonify({"ok": False, "error": "Pesan kosong."}), 400
    if not oa_client:
        return jsonify({"ok": False, "error": "OPENAI_API_KEY belum dikonfigurasi."}), 500

    hist = session.get("chat_history") or []
    hist = [h for h in hist if isinstance(h, dict) and h.get("role") and h.get("content")]
    hist = hist[-12:]
    base_url = request.host_url.rstrip("/")
    app_url = base_url

    system_prompt = f"""
Kamu "Asisten UMGAP" untuk UMKM. Tujuan: bantu user + arahkan ke fitur UMGAP (soft-selling).
Fitur: Absensi, Monitor Penjualan, AI Caption, Kelola Karyawan.
Jawab pakai bahasa Indonesia ramah. Akhiri dengan CTA ke {app_url}/login atau {app_url}/register.
""".strip()

    messages = [{"role": "system", "content": system_prompt}] + hist + [{"role": "user", "content": msg}]
    try:
        resp = oa_client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            messages=messages, temperature=0.7, max_tokens=450,
        )
        reply = (resp.choices[0].message.content or "").strip()
        hist.append({"role": "user", "content": msg})
        hist.append({"role": "assistant", "content": reply})
        session["chat_history"] = hist[-12:]
        return jsonify({"ok": True, "reply": reply, "app_url": app_url})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Gagal memproses AI: {str(e)}"}), 500

@app.route("/api/caption", methods=["POST"])
def api_caption():
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Silakan login dulu."}), 401
    data = request.get_json(silent=True) or {}
    for k in ["template", "biz_type", "tone", "product", "price", "wa"]:
        if not (data.get(k) or "").strip():
            return jsonify({"ok": False, "error": f"Field '{k}' wajib diisi."}), 400
    caption, vid = build_caption(data)
    return jsonify({"ok": True, "caption": caption, "variant_id": vid})

# ---------- POINTS ----------
@app.route("/admin/points")
def admin_points():
    deny = admin_required()
    if deny:
        return deny
    if os.getenv('RUN_SCHEMA_ON_REQUEST','').lower()=='true':
        ensure_points_schema()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, name, email, COALESCE(points, 0) AS points, COALESCE(points_admin, 0) AS points_admin
        FROM users WHERE role = 'employee' ORDER BY name ASC;
    """)
    employees = cur.fetchall()
    cur.execute("""
        SELECT l.created_at, u.name AS user_name, l.delta, l.note, a.name AS admin_name
        FROM points_logs l
        JOIN users u ON u.id = l.user_id
        JOIN users a ON a.id = l.admin_id
        ORDER BY l.created_at DESC LIMIT 50;
    """)
    logs = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("input_poin.html", user_name=session.get("user_name"), notif_count=0, employees=employees, logs=logs)

@app.route("/admin/points/add", methods=["POST"])
def admin_points_add():
    deny = admin_required()
    if deny:
        return deny
    if os.getenv('RUN_SCHEMA_ON_REQUEST','').lower()=='true':
        ensure_points_schema()
    user_id = int(request.form["user_id"])
    delta_raw = (request.form.get("delta") or "").strip()
    note = (request.form.get("note") or "").strip()
    try:
        delta = int(delta_raw)
    except:
        return "Delta poin harus angka.", 400
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("UPDATE users SET points_admin = COALESCE(points_admin,0) + %s WHERE id=%s AND role='employee';", (delta, user_id))
    cur.execute("INSERT INTO points_logs (user_id, admin_id, delta, note) VALUES (%s, %s, %s, %s);", (user_id, session.get("user_id"), delta, note if note else None))
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/admin/points")

# ---------- EXPORT ----------
def _validate_range(start, end):
    if end < start:
        return False, "Tanggal akhir harus >= tanggal awal."
    days = (end - start).days + 1
    if days > 31:
        return False, "Maksimal range 31 hari."
    today = datetime.now(ZoneInfo("Asia/Jakarta")).date()
    if start < (today - timedelta(days=183)):
        return False, "Range hanya boleh dari 6 bulan terakhir."
    if start.year != end.year or start.month != end.month:
        return False, "Range harus dalam bulan yang sama."
    return True, ""

def _autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

def _build_attendance_xlsx(rows_detail, recap_rows, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws.append([title])
    ws.append([])
    ws.append(["Nama", "Tanggal", "Jam Kehadiran", "Status", "Gaji Harian", "Catatan"])
    for r in rows_detail:
        ws.append([r.get("name", ""), r.get("work_date", ""), r.get("checkin_time", ""), r.get("status", ""), r.get("daily_salary", 0), r.get("note", "")])
    _autosize_columns(ws)

    ws2 = wb.create_sheet("Rekap")
    ws2.append([title])
    ws2.append([])
    ws2.append(["Nama", "Hadir", "Sakit", "Izin", "Absen", "Total Gaji"])
    for rr in recap_rows:
        ws2.append([rr["name"], rr["present"], rr["sick"], rr["leave"], rr["absent"], rr["total_salary"]])
    _autosize_columns(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@app.route("/admin/data/range.xlsx")
def admin_download_range_xlsx():
    deny = admin_required()
    if deny:
        return deny
    ensure_hr_v2_schema()

    start = _parse_date(request.args.get("start"))
    end = _parse_date(request.args.get("end"))
    if not start or not end:
        return "start & end wajib (YYYY-MM-DD)", 400

    ok, msg = _validate_range(start, end)
    if not ok:
        return msg, 400

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT u.name, COALESCE(p.daily_salary, 0) AS daily_salary, a.work_date, a.status, a.note, a.checkin_at
        FROM attendance a
        JOIN users u ON u.id = a.user_id
        LEFT JOIN payroll_settings p ON p.user_id = u.id
        WHERE u.role='employee' AND a.work_date >= %s AND a.work_date <= %s
        ORDER BY u.name ASC, a.work_date ASC, a.checkin_at ASC NULLS LAST;
    """, (start, end))
    raw = cur.fetchall()
    cur.close()
    conn.close()

    detail = []
    recap = {}
    for r in raw:
        name = r["name"]
        ds = int(r["daily_salary"] or 0)
        status = (r["status"] or "").upper()
        checkin_time = r["checkin_at"].strftime("%H:%M:%S") if r["checkin_at"] else ""
        detail.append({"name": name, "work_date": r["work_date"].isoformat() if r["work_date"] else "", "checkin_time": checkin_time, "status": status, "daily_salary": ds, "note": (r.get("note") or "").strip()})
        if name not in recap:
            recap[name] = {"name": name, "present": 0, "sick": 0, "leave": 0, "absent": 0, "total_salary": 0}
        if status == "PRESENT":
            recap[name]["present"] += 1
            recap[name]["total_salary"] += ds
        elif status == "SICK":
            recap[name]["sick"] += 1
        elif status == "LEAVE":
            recap[name]["leave"] += 1
        elif status == "ABSENT":
            recap[name]["absent"] += 1

    recap_rows = list(recap.values())
    title = f"Rekap Absensi & Gaji ({start.isoformat()} s/d {end.isoformat()})"
    xlsx_bytes = _build_attendance_xlsx(detail, recap_rows, title)
    filename = f"rekap_{start.isoformat()}_{end.isoformat()}.xlsx"
    return Response(xlsx_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@app.route("/admin/data/range_user.xlsx")
def admin_download_range_user_xlsx():
    deny = admin_required()
    if deny:
        return deny
    ensure_hr_v2_schema()

    user_id = request.args.get("user_id")
    if not user_id:
        return "user_id wajib", 400
    start = _parse_date(request.args.get("start"))
    end = _parse_date(request.args.get("end"))
    if not start or not end:
        return "start & end wajib (YYYY-MM-DD)", 400

    ok, msg = _validate_range(start, end)
    if not ok:
        return msg, 400

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, name FROM users WHERE id=%s LIMIT 1;", (int(user_id),))
    urow = cur.fetchone()
    if not urow:
        cur.close()
        conn.close()
        return "User tidak ditemukan", 404
    user_name = urow["name"]

    cur.execute("""
        SELECT COALESCE(p.daily_salary, 0) AS daily_salary, a.work_date, a.status, a.note, a.checkin_at
        FROM attendance a
        LEFT JOIN payroll_settings p ON p.user_id = a.user_id
        WHERE a.user_id=%s AND a.work_date >= %s AND a.work_date <= %s
        ORDER BY a.work_date ASC, a.checkin_at ASC NULLS LAST;
    """, (int(user_id), start, end))
    raw = cur.fetchall()
    cur.close()
    conn.close()

    detail = []
    recap = {"name": user_name, "present": 0, "sick": 0, "leave": 0, "absent": 0, "total_salary": 0}
    for r in raw:
        ds = int(r["daily_salary"] or 0)
        status = (r["status"] or "").upper()
        checkin_time = r["checkin_at"].strftime("%H:%M:%S") if r["checkin_at"] else ""
        detail.append({"name": user_name, "work_date": r["work_date"].isoformat() if r["work_date"] else "", "checkin_time": checkin_time, "status": status, "daily_salary": ds, "note": (r.get("note") or "").strip()})
        if status == "PRESENT":
            recap["present"] += 1
            recap["total_salary"] += ds
        elif status == "SICK":
            recap["sick"] += 1
        elif status == "LEAVE":
            recap["leave"] += 1
        elif status == "ABSENT":
            recap["absent"] += 1

    title = f"Rekap {user_name} ({start.isoformat()} s/d {end.isoformat()})"
    xlsx_bytes = _build_attendance_xlsx(detail, [recap], title)
    filename = f"rekap_{user_name}_{start.isoformat()}_{end.isoformat()}.xlsx".replace(" ", "_")
    return Response(xlsx_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# ---------- PREVIEW ----------
@app.route("/preview/<name>")
def preview_template(name):
    allowed = {"login": "login.html", "register": "register.html", "dashboard": "dashboard.html", "products": "products.html"}
    if name not in allowed:
        abort(404)
    dummy = {
        "user_name": "UMKM Demo", "total_products": 3, "total_contents": 5, "total_done": 2,
        "products": [{"id": 1, "name": "Kopi Susu", "price": 12000}, {"id": 2, "name": "Roti Bakar", "price": 15000}, {"id": 3, "name": "Teh Manis", "price": 6000}],
        "error": None
    }
    return render_template(allowed[name], **dummy)

# ==================== THERMAL PRINT API ====================

def _build_escpos(invoice, items, paper_width=80):
    """Build ESC/POS byte array for ZJ-5809 II and similar BT thermal printers."""
    import struct

    ESC, GS, LF = 0x1B, 0x1D, 0x0A
    W = 32 if paper_width == 58 else 48  # char width

    def enc(s):
        return [c if c < 128 else 0x3F for c in s.encode('ascii', errors='replace')]

    def pad_row(left, right, w):
        left, right = str(left or ''), str(right or '')
        gap = max(1, w - len(left) - len(right))
        line = left + ' ' * gap + right
        return enc(line[:w]) + [LF]

    buf = []
    b = buf.extend

    # Init
    b([ESC, 0x40])              # init
    b([ESC, 0x74, 0x00])        # codepage PC437

    # Header - center + big
    b([ESC, 0x61, 0x01])        # align center
    b([ESC, 0x21, 0x30])        # double width+height
    b([ESC, 0x45, 0x01])        # bold on
    b(enc('UMGAP') + [LF])
    b([ESC, 0x21, 0x00])        # normal size
    b([ESC, 0x45, 0x00])        # bold off
    b(enc('Nota Penjualan') + [LF])

    # Dashes
    dash = enc('-' * W) + [LF]

    b(dash)
    b([ESC, 0x61, 0x00])        # align left

    # Invoice info
    inv_no = str(invoice.get('invoice_no', '') or '')
    created = ''
    if invoice.get('created_at'):
        try:
            created = invoice['created_at'].strftime('%d/%m/%Y %H:%M')
        except Exception:
            created = str(invoice['created_at'])[:16]

    b(pad_row('No', inv_no[:W-3], W))
    b(pad_row('Tanggal', created, W))
    b(pad_row('Customer', str(invoice.get('customer_name') or '-')[:W-10], W))
    b(pad_row('Kasir', str(invoice.get('created_by_name') or '-')[:W-6], W))
    b(pad_row('Bayar', str(invoice.get('payment_method') or 'CASH'), W))
    b(dash)

    # Items
    for item in items:
        name = str(item.get('product_name', '') or '')
        qty  = int(item.get('qty', 1) or 1)
        price = int(item.get('price', 0) or 0)
        sub   = int(item.get('subtotal', 0) or 0)
        price_fmt = '{:,}'.format(price).replace(',', '.')
        sub_fmt   = '{:,}'.format(sub).replace(',', '.')
        b([ESC, 0x45, 0x01])    # bold
        b(enc(name[:W]) + [LF])
        b([ESC, 0x45, 0x00])    # bold off
        detail_left  = '  {} x Rp {}'.format(qty, price_fmt)
        detail_right = 'Rp {}'.format(sub_fmt)
        b(pad_row(detail_left, detail_right, W))

    b(dash)

    # Total - double height
    grand = int(invoice.get('grand_total', 0) or 0)
    grand_fmt = '{:,}'.format(grand).replace(',', '.')
    b([ESC, 0x61, 0x02])        # align right
    b([ESC, 0x21, 0x10])        # double height
    b([ESC, 0x45, 0x01])        # bold
    b(enc('TOTAL: Rp {}'.format(grand_fmt)) + [LF])
    b([ESC, 0x21, 0x00])        # normal
    b([ESC, 0x45, 0x00])        # bold off
    b([ESC, 0x61, 0x00])        # align left
    b(dash)

    # Notes
    notes = str(invoice.get('notes') or '').strip()
    if notes:
        b(enc('Catatan: ' + notes[:W-9]) + [LF])
        b(dash)

    # Footer
    b([ESC, 0x61, 0x01])        # center
    b(enc('Terima kasih sudah berbelanja!') + [LF])
    b([LF, LF, LF])

    # Cut
    b([GS, 0x56, 0x42, 0x20])  # partial cut

    return bytes(buf)


def _get_invoice_with_items(invoice_id):
    """Fetch invoice + items dict from DB."""
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT i.*, u.name AS created_by_name
            FROM invoices i JOIN users u ON u.id = i.created_by
            WHERE i.id=%s LIMIT 1;
        """, (invoice_id,))
        inv = cur.fetchone()
        if not inv:
            return None, []
        cur.execute("""
            SELECT product_name, qty, price, subtotal
            FROM invoice_items WHERE invoice_id=%s ORDER BY id;
        """, (invoice_id,))
        items = cur.fetchall()
        return dict(inv), [dict(r) for r in items]
    finally:
        cur.close()
        conn.close()


@app.route("/invoice/<int:invoice_id>/escpos")
def invoice_escpos(invoice_id):
    """Download raw ESC/POS binary - for RawBT or manual use."""
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    ensure_invoice_schema()
    inv, items = _get_invoice_with_items(invoice_id)
    if not inv:
        abort(404)

    pw_str = (inv.get('print_size') or '80mm').replace('mm', '')
    pw = 58 if pw_str == '58' else 80

    data = _build_escpos(inv, items, paper_width=pw)
    fname = 'nota_{}.bin'.format(inv.get('invoice_no', str(invoice_id)))

    return Response(
        data,
        mimetype='application/octet-stream',
        headers={'Content-Disposition': 'attachment; filename="{}"'.format(fname)}
    )


@app.route("/invoice/<int:invoice_id>/print-server", methods=["POST"])
def invoice_print_server(invoice_id):
    """Print via pyserial to Bluetooth COM port (Windows only)."""
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if not SERIAL_OK:
        return jsonify({"ok": False, "error": "pyserial tidak terinstall di server. Jalankan: pip install pyserial"}), 500

    ensure_invoice_schema()
    inv, items = _get_invoice_with_items(invoice_id)
    if not inv:
        return jsonify({"ok": False, "error": "Invoice tidak ditemukan"}), 404

    com_port = (request.json or {}).get('port', '').strip()
    pw_str   = (inv.get('print_size') or '80mm').replace('mm', '')
    pw = 58 if pw_str == '58' else 80

    # Auto-detect if no port given
    if not com_port:
        ports = serial.tools.list_ports.comports()
        for p in ports:
            desc = (p.description or '').lower()
            if 'bluetooth' in desc or 'spp' in desc or 'serial' in desc or 'com' in desc.lower():
                com_port = p.device
                break
        if not com_port and ports:
            com_port = ports[0].device

    if not com_port:
        return jsonify({"ok": False, "error": "Tidak ada COM port ditemukan. Pastikan printer sudah di-pair di Windows."}), 400

    try:
        data = _build_escpos(inv, items, paper_width=pw)
        with serial.Serial(com_port, baudrate=9600, timeout=3) as ser:
            ser.write(data)
        return jsonify({"ok": True, "port": com_port, "bytes": len(data)})
    except serial.SerialException as e:
        return jsonify({"ok": False, "error": "Serial error: {}".format(str(e))}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/invoice/<int:invoice_id>/list-ports")
def invoice_list_ports(invoice_id):
    """List available COM ports on server (Windows)."""
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    if not SERIAL_OK:
        return jsonify({"ok": False, "ports": [], "error": "pyserial tidak terinstall"}), 200
    ports = [{"device": p.device, "desc": p.description} for p in serial.tools.list_ports.comports()]
    return jsonify({"ok": True, "ports": ports})


# ==================== HARGA BELI SCRAP ====================

def ensure_buy_prices_schema():
    """Buat tabel harga beli scrap material untuk landing page."""
    conn = get_conn()
    cur  = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS buy_prices (
                id         SERIAL PRIMARY KEY,
                material   VARCHAR(100) NOT NULL,
                grade      VARCHAR(150) NOT NULL DEFAULT '',
                unit       VARCHAR(20)  NOT NULL DEFAULT 'kg',
                price      NUMERIC(10,2) NOT NULL DEFAULT 0,
                note       TEXT,
                is_active  BOOLEAN      NOT NULL DEFAULT TRUE,
                sort_order INTEGER      NOT NULL DEFAULT 0,
                updated_at TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_at TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)
        # safe migration: add columns if not exist (idempotent)
        for col_sql in [
            "ALTER TABLE buy_prices ADD COLUMN IF NOT EXISTS note TEXT;",
            "ALTER TABLE buy_prices ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE;",
            "ALTER TABLE buy_prices ADD COLUMN IF NOT EXISTS sort_order INTEGER NOT NULL DEFAULT 0;",
            "ALTER TABLE buy_prices ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP;",
            "ALTER TABLE buy_prices ADD COLUMN IF NOT EXISTS created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP;",
        ]:
            try:
                cur.execute(col_sql)
            except Exception:
                conn.rollback()
        cur.execute("SELECT COUNT(*) AS n FROM buy_prices;")
        if cur.fetchone()[0] == 0:
            defaults = [
                ("Tembaga", "TM", "kg", 205, 10),
                ("Tembaga", "TS / TS Halus", "kg", 202, 11),
                ("Tembaga", "BC", "kg", 198.5, 12),
                ("Tembaga", "Telkom / TB Pipa", "kg", 197.5, 13),
                ("Tembaga", "TB / TB Bakar", "kg", 189, 14),
                ("Tembaga", "TB Putih Lidi", "kg", 189, 15),
                ("Tembaga", "TB Putih", "kg", 184, 16),
                ("Tembaga", "DD", "kg", 182, 17),
                ("Tembaga", "Gram TB", "kg", 150, 18),
                ("Tembaga", "Gram TB Lembut", "kg", 140, 19),
                ("Tembaga", "Jarum TB Putih", "kg", 147.5, 20),
                ("Tembaga", "Jarum TB Hitam", "kg", 150.5, 21),
                ("Tembaga", "RD TB", "kg", 165, 22),
                ("Tembaga", "TB Kotak", "kg", 122, 23),
                ("Tembaga", "TB Bakau Super", "kg", 90, 24),
                ("Kuningan", "Bron", "kg", 169, 30),
                ("Kuningan", "Bron Putih", "kg", 164, 31),
                ("Kuningan", "Plat KN", "kg", 125, 32),
                ("Kuningan", "Patrum Bersih", "kg", 128, 33),
                ("Kuningan", "KN Kasar", "kg", 122, 34),
                ("Kuningan", "KN Rosok / KN Puler", "kg", 119, 35),
                ("Kuningan", "KN Kipas", "kg", 121, 36),
                ("Kuningan", "KN Rambut", "kg", 120, 37),
                ("Kuningan", "KN Gelang", "kg", 107, 38),
                ("Kuningan", "AISI", "kg", 123, 39),
                ("Kuningan", "AISI Kawul", "kg", 120, 40),
                ("Kuningan", "RD KN", "kg", 115, 41),
                ("Kuningan", "RD KN Lepas", "kg", 110, 42),
                ("Kuningan", "KN Totok", "kg", 115, 43),
                ("Kuningan", "KN Paten", "kg", 103, 44),
                ("Kuningan", "KN Tanjek", "kg", 103, 45),
                ("Kuningan", "Gram ME", "kg", 118, 46),
                ("Kuningan", "Gram Merah Lembut", "kg", 113, 47),
                ("Kuningan", "Gram KN As", "kg", 106, 48),
                ("Kuningan", "Gram Kemprotok", "kg", 108, 49),
                ("Kuningan", "Gram KN Kawul", "kg", 103, 50),
                ("Kuningan", "Gram Juwana", "kg", 100, 51),
                ("Kuningan", "Awon KN", "kg", 77, 52),
                ("Aluminium", "AK", "kg", 50, 60),
                ("Aluminium", "AK Bakar", "kg", 49, 61),
                ("Aluminium", "Kusen", "kg", 47.5, 62),
                ("Aluminium", "Siku", "kg", 43, 63),
                ("Aluminium", "Siku Cat", "kg", 42, 64),
                ("Aluminium", "Plat Koran", "kg", 48, 65),
                ("Aluminium", "Plat KPU", "kg", 43.5, 66),
                ("Aluminium", "Plat A", "kg", 42.5, 67),
                ("Aluminium", "Pelek Mobil", "kg", 44, 68),
                ("Aluminium", "Pelek Mobil Krom", "kg", 43.5, 69),
                ("Aluminium", "Seker", "kg", 40, 70),
                ("Aluminium", "Blok", "kg", 39.5, 71),
                ("Aluminium", "Blok 2", "kg", 36, 72),
                ("Aluminium", "Blok Parabola", "kg", 32.5, 73),
                ("Aluminium", "Kampas B", "kg", 35.5, 74),
                ("Aluminium", "Kampas K", "kg", 27.5, 75),
                ("Aluminium", "Plat B", "kg", 38, 76),
                ("Aluminium", "Plat Nomor", "kg", 39.5, 77),
                ("Aluminium", "Plat Lembutan", "kg", 31, 78),
                ("Aluminium", "Plat Jeruk", "kg", 34.5, 79),
                ("Aluminium", "Parfum B", "kg", 39.5, 80),
                ("Aluminium", "Parfum Kotor", "kg", 18.5, 81),
                ("Aluminium", "Nium Dinamo", "kg", 31.5, 82),
                ("Aluminium", "RD N Utuh", "kg", 34, 83),
                ("Aluminium", "RD N Lepas", "kg", 31, 84),
                ("Aluminium", "Panci LPK", "kg", 37.5, 85),
                ("Aluminium", "PC", "kg", 37, 86),
                ("Aluminium", "PC Silitan Bersih", "kg", 33, 87),
                ("Aluminium", "Kaleng", "kg", 36, 88),
                ("Aluminium", "Wajan", "kg", 31, 89),
                ("Aluminium", "Elemen", "kg", 25.5, 90),
                ("Aluminium", "Kerey", "kg", 24, 91),
                ("Aluminium", "Gram Nium", "kg", 19.5, 92),
                ("Aluminium", "Gram Nium Kemprotok", "kg", 21, 93),
                ("Aluminium", "Lelehan Nium", "kg", 16, 94),
                ("Aluminium", "Ring", "kg", 21.5, 95),
                ("Aluminium", "Tutup", "kg", 15, 96),
                ("Aluminium", "Nium Api", "kg", 15, 97),
                ("Aluminium", "Foil", "kg", 11, 98),
                ("Stainless", "Monel", "kg", 15.5, 110),
                ("Stainless", "Monel Cat", "kg", 15, 111),
                ("Stainless", "India", "kg", 6, 112),
                ("Timah & Aki", "Timah KPL", "kg", 29, 120),
                ("Timah & Aki", "Nium KPL", "kg", 25, 121),
                ("Timah & Aki", "Budeng", "kg", 30, 122),
                ("Timah & Aki", "Lakson", "kg", 32, 123),
                ("Timah & Aki", "Lakson RBS", "kg", 15.5, 124),
                ("Timah & Aki", "Aki Bersih Bebas Air", "kg", 16.4, 125),
            ]
            for mat, grade, unit, price, sort in defaults:
                cur.execute("""
                    INSERT INTO buy_prices (material, grade, unit, price, sort_order)
                    VALUES (%s,%s,%s,%s,%s);
                """, (mat, grade, unit, price, sort))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


@app.route("/api/buy-prices", methods=["GET","OPTIONS"])
def api_buy_prices():
    """Public API — harga beli untuk landing page (CORS enabled)."""
    # Preflight
    if request.method == "OPTIONS":
        resp = Response("", status=200)
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    ensure_buy_prices_schema()
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, material, grade, unit, price, note, updated_at
            FROM buy_prices
            WHERE is_active = TRUE
            ORDER BY sort_order ASC, id ASC;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    items = []
    for r in rows:
        items.append({
            "id":       r["id"],
            "material": r["material"],
            "grade":    r["grade"],
            "unit":     r["unit"],
            "price":    float(r["price"]) if r["price"] else 0,
            "note":     r["note"] or "",
            "updated_at": r["updated_at"].strftime("%d/%m/%Y") if r["updated_at"] else "-",
        })
    groups = {}
    for it in items:
        m = it["material"]
        if m not in groups:
            groups[m] = []
        groups[m].append(it)

    resp = jsonify({"ok": True, "groups": [{"material": k, "items": v} for k, v in groups.items()]})
    resp.headers["Access-Control-Allow-Origin"]  = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    resp.headers["Cache-Control"] = "public, max-age=300"
    return resp


@app.route("/admin/buy-prices")
def admin_buy_prices():
    """Halaman kelola harga beli scrap."""
    deny = admin_required()
    if deny:
        return deny
    ensure_buy_prices_schema()
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, material, grade, unit, price, note, is_active, sort_order, updated_at
            FROM buy_prices ORDER BY sort_order ASC, id ASC;
        """)
        raw_rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    # Konversi ke plain dict — Decimal & datetime tidak bisa tojson Jinja2
    rows = []
    for r in raw_rows:
        rows.append({
            "id":         int(r["id"]),
            "material":   str(r["material"] or ""),
            "grade":      str(r["grade"] or ""),
            "unit":       str(r["unit"] or "kg"),
            "price":      float(r["price"]) if r["price"] is not None else 0.0,
            "note":       str(r["note"] or ""),
            "is_active":  bool(r["is_active"]) if r["is_active"] is not None else True,
            "sort_order": int(r["sort_order"]) if r["sort_order"] is not None else 0,
            "updated_at_str": r["updated_at"].strftime("%d/%m/%Y %H:%M") if r["updated_at"] else "-",
        })
    return render_template("admin_buy_prices.html", rows=rows)


@app.route("/admin/buy-prices/save", methods=["POST"])
def admin_buy_prices_save():
    """Simpan perubahan harga beli (upsert via JSON)."""
    deny = admin_required()
    if deny:
        return deny
    ensure_buy_prices_schema()
    data = request.get_json(silent=True) or {}
    action = data.get("action", "")
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if action == "update":
            pid    = int(data.get("id", 0))
            price  = max(0.0, float(data.get("price", 0) or 0))
            note   = (data.get("note") or "").strip()
            is_active = bool(data.get("is_active", True))
            cur.execute("""
                UPDATE buy_prices
                SET price=%s, note=%s, is_active=%s,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=%s
                RETURNING id, price, note, is_active, updated_at;
            """, (price, note, is_active, pid))
            row = cur.fetchone()
            conn.commit()
            if not row:
                return jsonify({"ok": False, "error": "Not found"}), 404
            return jsonify({"ok": True, "row": {
                "id":            int(row["id"]),
                "price":         float(row["price"]) if row["price"] is not None else 0.0,
                "note":          str(row["note"] or ""),
                "is_active":     bool(row["is_active"]),
                "updated_at_str": row["updated_at"].strftime("%d/%m/%Y %H:%M"),
            }})
        elif action == "add":
            material = (data.get("material") or "").strip()
            grade    = (data.get("grade") or "").strip()
            unit     = (data.get("unit") or "kg").strip()
            price    = max(0.0, float(data.get("price", 0) or 0))
            note     = (data.get("note") or "").strip()
            if not material:
                return jsonify({"ok": False, "error": "Material wajib diisi"}), 400
            cur.execute("""
                INSERT INTO buy_prices (material, grade, unit, price, note, sort_order)
                VALUES (%s,%s,%s,%s,%s,
                    (SELECT COALESCE(MAX(sort_order),0)+1 FROM buy_prices))
                RETURNING id, material, grade, unit, price, note, is_active, sort_order, updated_at;
            """, (material, grade, unit, price, note))
            row = cur.fetchone()
            conn.commit()
            return jsonify({"ok": True, "row": {
                "id":            int(row["id"]),
                "material":      str(row["material"] or ""),
                "grade":         str(row["grade"] or ""),
                "unit":          str(row["unit"] or "kg"),
                "price":         float(row["price"]) if row["price"] is not None else 0.0,
                "note":          str(row["note"] or ""),
                "is_active":     bool(row["is_active"]) if row["is_active"] is not None else True,
                "sort_order":    int(row["sort_order"]) if row["sort_order"] is not None else 0,
                "updated_at_str": row["updated_at"].strftime("%d/%m/%Y %H:%M"),
            }})
        elif action == "delete":
            pid = int(data.get("id", 0))
            cur.execute("DELETE FROM buy_prices WHERE id=%s;", (pid,))
            conn.commit()
            return jsonify({"ok": True})
        else:
            return jsonify({"ok": False, "error": "Unknown action"}), 400
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

# ==================== CCTV ROUTES ====================
@app.route("/cctv_stream/<camera_code>")
def cctv_stream_multi(camera_code):
    deny = admin_required()
    if deny:
        return deny

    cam = get_camera_config(camera_code)
    if not cam:
        return "Camera not found", 404

    source_url = get_camera_url(camera_code)
    if not source_url:
        return "Sumber kamera tidak ditemukan", 404

    return Response(
        generate_frames_multi(
            rtsp_url=source_url,
            camera_name=cam.get("name") or cam.get("code") or "CCTV",
            camera_code=(cam.get("code") or camera_code).upper(),
        ),
        mimetype="multipart/x-mixed-replace; boundary=frame",
    )

@app.route("/admin/cctv-multi")
def admin_cctv_multi():
    deny = admin_required()
    if deny:
        return deny

    cameras = [cam for cam in CCTV_CAMERAS if (cam.get("rtsp_url") or "").strip()]

    return render_template(
        "admin_cctv_multi.html",
        cameras=cameras,
        user_name=session.get("user_name", "Admin"),
    )


@app.route("/admin/cctv-multi/json")
def admin_cctv_multi_json():
    deny = admin_required()
    if deny:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    cameras = []
    for cam in CCTV_CAMERAS:
        if not (cam.get("rtsp_url") or "").strip():
            continue

        cameras.append({
            "code": cam.get("code"),
            "name": cam.get("name"),
            "location": cam.get("location"),
            "stream_url": url_for("cctv_stream_multi", camera_code=cam.get("code")),
        })

    return jsonify({"ok": True, "cameras": cameras})

@app.route("/admin/cctv-analytics")
def admin_cctv_analytics():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_activity_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                camera_code,
                work_date,
                moving_seconds,
                idle_seconds,
                last_status,
                last_motion_score,
                updated_at
            FROM cctv_activity_daily
            WHERE work_date = CURRENT_DATE
            ORDER BY camera_code ASC;
        """)
        raw_rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    rows = []
    for r in raw_rows:
        row = dict(r)
        row["moving_hms"] = _seconds_to_hms(r.get("moving_seconds") or 0)
        row["idle_hms"] = _seconds_to_hms(r.get("idle_seconds") or 0)
        rows.append(row)

    return render_template(
        "admin_cctv_analytics.html",
        rows=rows,
        user_name=session.get("user_name", "Admin"),
    )

@app.route("/admin/cctv-analytics/json")
def admin_cctv_analytics_json():
    deny = admin_required()
    if deny:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    ensure_cctv_activity_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                camera_code,
                work_date,
                moving_seconds,
                idle_seconds,
                last_status,
                last_motion_score,
                updated_at
            FROM cctv_activity_daily
            WHERE work_date = CURRENT_DATE
            ORDER BY camera_code ASC;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    data = []
    for r in rows:
        data.append({
            "camera_code": r["camera_code"],
            "work_date": str(r["work_date"]),
            "moving_seconds": int(r.get("moving_seconds") or 0),
            "idle_seconds": int(r.get("idle_seconds") or 0),
            "moving_hms": _seconds_to_hms(r.get("moving_seconds") or 0),
            "idle_hms": _seconds_to_hms(r.get("idle_seconds") or 0),
            "last_status": r.get("last_status"),
            "last_motion_score": int(r.get("last_motion_score") or 0),
            "updated_at": str(r.get("updated_at") or ""),
        })

    return jsonify({"ok": True, "rows": data})

@app.route("/admin/cctv-zones", methods=["GET", "POST"])
def admin_cctv_zones():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_activity_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()

            if action == "add":
                camera_code = (request.form.get("camera_code") or "").strip().upper()
                user_id = int(request.form.get("user_id") or 0)
                zone_name = (request.form.get("zone_name") or "").strip()
                x1 = int(request.form.get("x1") or 0)
                y1 = int(request.form.get("y1") or 0)
                x2 = int(request.form.get("x2") or 0)
                y2 = int(request.form.get("y2") or 0)
                idle_threshold_seconds = int(request.form.get("idle_threshold_seconds") or 600)

                if camera_code and zone_name and user_id > 0:
                    cur.execute("""
                        INSERT INTO cctv_employee_zones
                            (camera_code, user_id, zone_name, x1, y1, x2, y2, idle_threshold_seconds, is_active, updated_at)
                        VALUES
                            (%s, %s, %s, %s, %s, %s, %s, %s, TRUE, CURRENT_TIMESTAMP);
                    """, (
                        camera_code, user_id, zone_name, x1, y1, x2, y2, idle_threshold_seconds
                    ))
                    conn.commit()

            elif action == "toggle":
                zid = int(request.form.get("id") or 0)
                cur.execute("""
                    UPDATE cctv_employee_zones
                    SET is_active = NOT is_active,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id=%s;
                """, (zid,))
                conn.commit()

            elif action == "delete":
                zid = int(request.form.get("id") or 0)
                cur.execute("DELETE FROM cctv_employee_zones WHERE id=%s;", (zid,))
                conn.commit()

        cur.execute("""
            SELECT id, name
            FROM users
            WHERE role='employee'
            ORDER BY name ASC;
        """)
        employees = cur.fetchall()

        cur.execute("""
            SELECT
                z.*,
                u.name AS user_name
            FROM cctv_employee_zones z
            LEFT JOIN users u ON u.id = z.user_id
            ORDER BY z.camera_code ASC, z.id ASC;
        """)
        zones = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    cameras = [cam for cam in CCTV_CAMERAS if (cam.get("rtsp_url") or "").strip()]

    return render_template(
        "admin_cctv_zones.html",
        cameras=cameras,
        employees=employees,
        zones=zones,
    )

@app.route("/admin/cctv-employee-analytics")
def admin_cctv_employee_analytics():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_activity_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                a.camera_code,
                a.work_date,
                a.zone_id,
                a.user_id,
                u.name AS user_name,
                z.zone_name,
                a.moving_seconds,
                a.idle_seconds,
                a.absent_seconds,
                a.last_status,
                a.last_motion_score,
                a.updated_at
            FROM cctv_employee_activity_daily a
            LEFT JOIN users u ON u.id = a.user_id
            LEFT JOIN cctv_employee_zones z ON z.id = a.zone_id
            WHERE a.work_date = CURRENT_DATE
            ORDER BY a.camera_code ASC, z.zone_name ASC;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_cctv_employee_analytics.html", rows=rows)

@app.route("/admin/cctv-idle-snapshots")
def admin_cctv_idle_snapshots():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_activity_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                s.*,
                u.name AS user_name,
                z.zone_name
            FROM cctv_idle_snapshots s
            LEFT JOIN users u ON u.id = s.user_id
            LEFT JOIN cctv_employee_zones z ON z.id = s.zone_id
            ORDER BY s.created_at DESC
            LIMIT 200;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_cctv_idle_snapshots.html", rows=rows)

@app.route("/admin/cctv-events", methods=["GET", "POST"])
def admin_cctv_events():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_report_schema()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if request.method == "POST":
            snapshot_id = int(request.form.get("snapshot_id") or 0)
            subject_label = (request.form.get("subject_label") or "").strip()
            note = (request.form.get("note") or "").strip()

            if snapshot_id > 0 and subject_label:
                cur.execute("""
                    INSERT INTO cctv_snapshot_labels
                        (snapshot_id, subject_label, note, created_by)
                    VALUES
                        (%s, %s, %s, %s);
                """, (
                    snapshot_id,
                    subject_label,
                    note,
                    session.get("user_id"),
                ))
                conn.commit()

        cur.execute("""
            SELECT
                s.id,
                s.camera_code,
                s.event_type,
                s.snapshot_path,
                s.motion_score,
                s.note,
                s.created_at,
                l.subject_label,
                l.note AS label_note
            FROM cctv_event_snapshots s
            LEFT JOIN LATERAL (
                SELECT subject_label, note
                FROM cctv_snapshot_labels
                WHERE snapshot_id = s.id
                ORDER BY id DESC
                LIMIT 1
            ) l ON TRUE
            ORDER BY s.created_at DESC
            LIMIT 200;
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_cctv_events.html", rows=rows)

@app.route("/admin/cctv-daily-report")
def admin_cctv_daily_report():
    deny = admin_required()
    if deny:
        return deny

    ensure_cctv_report_schema()

    report_date = request.args.get("date") or date.today().isoformat()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                s.camera_code,
                COALESCE(l.subject_label, 'BELUM DILABEL') AS subject_label,
                COUNT(*) AS total_snapshots,
                MIN(s.created_at) AS first_seen,
                MAX(s.created_at) AS last_seen,
                MAX(s.snapshot_path) AS sample_snapshot
            FROM cctv_event_snapshots s
            LEFT JOIN LATERAL (
                SELECT subject_label
                FROM cctv_snapshot_labels
                WHERE snapshot_id = s.id
                ORDER BY id DESC
                LIMIT 1
            ) l ON TRUE
            WHERE DATE(s.created_at) = %s
            GROUP BY s.camera_code, COALESCE(l.subject_label, 'BELUM DILABEL')
            ORDER BY s.camera_code ASC, subject_label ASC;
        """, (report_date,))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return render_template("admin_cctv_daily_report.html", rows=rows, report_date=report_date)


#--------- MOBILE APP ROUTE ----------
@app.route("/api/mobile/login", methods=["POST"])
def api_mobile_login():
    ensure_mobile_api_schema()

    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()
    device_name = (data.get("device_name") or "").strip()

    if not email or not password:
        return mobile_api_response(
            ok=False,
            message="Email dan password wajib diisi.",
            status_code=400
        )

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, email, password_hash, role
            FROM users
            WHERE email=%s
            LIMIT 1;
        """, (email,))
        user = cur.fetchone()

        if not user or not check_password_hash(user["password_hash"], password):
            return mobile_api_response(
                ok=False,
                message="Email atau password salah.",
                status_code=401
            )

        token = uuid.uuid4().hex + uuid.uuid4().hex

        cur.execute("""
            INSERT INTO mobile_api_tokens (user_id, token, device_name, is_active)
            VALUES (%s, %s, %s, TRUE)
            RETURNING id;
        """, (
            user["id"],
            token,
            device_name or "Android Device"
        ))
        cur.fetchone()
        conn.commit()

        return mobile_api_response(
            ok=True,
            message="Login berhasil.",
            data={
                "token": token,
                "user": {
                    "id": user["id"],
                    "name": user["name"],
                    "email": user["email"],
                    "role": user["role"],
                }
            }
        )
    finally:
        cur.close()
        conn.close()


@app.route("/api/mobile/logout", methods=["POST"])
@mobile_api_login_required
def api_mobile_logout():
    token = get_bearer_token()
    if not token:
        return mobile_api_response(
            ok=False,
            message="Token tidak ditemukan.",
            status_code=400
        )

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE mobile_api_tokens
            SET is_active=FALSE
            WHERE token=%s;
        """, (token,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return mobile_api_response(
        ok=True,
        message="Logout berhasil."
    )

@app.route("/api/mobile/me", methods=["GET"])
@mobile_api_login_required
def api_mobile_me():
    user = request.mobile_user

    return mobile_api_response(
        ok=True,
        message="Data user berhasil diambil.",
        data={
            "user": {
                "id": user["id"],
                "name": user["name"],
                "email": user["email"],
                "role": user["role"],
                "points": int(user.get("points") or 0),
                "points_admin": int(user.get("points_admin") or 0),
            }
        }
    )

@app.route("/api/mobile/dashboard", methods=["GET"])
@mobile_api_login_required
def api_mobile_dashboard():
    user = request.mobile_user
    user_id = user["id"]
    role = user["role"]

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if role == "admin":
            cur.execute("SELECT COUNT(*) AS total FROM users WHERE role='employee';")
            total_employees = int((cur.fetchone() or {}).get("total", 0))

            cur.execute("SELECT COUNT(*) AS total FROM products;")
            total_products = int((cur.fetchone() or {}).get("total", 0))

            cur.execute("SELECT COUNT(*) AS total FROM attendance WHERE work_date=%s AND status='PRESENT';", (date.today(),))
            total_attendance_today = int((cur.fetchone() or {}).get("total", 0))

            data = {
                "role": role,
                "summary": {
                    "total_employees": total_employees,
                    "total_products": total_products,
                    "total_attendance_today": total_attendance_today,
                }
            }
        else:
            cur.execute("""
                SELECT COUNT(*) AS total
                FROM attendance
                WHERE user_id=%s AND status='PRESENT';
            """, (user_id,))
            total_present = int((cur.fetchone() or {}).get("total", 0))

            cur.execute("""
                SELECT COUNT(*) AS total
                FROM sales_submissions
                WHERE user_id=%s;
            """, (user_id,))
            total_sales_submissions = int((cur.fetchone() or {}).get("total", 0))

            cur.execute("""
                SELECT COUNT(*) AS total
                FROM content_plans
                WHERE user_id=%s AND is_done=FALSE;
            """, (user_id,))
            pending_content = int((cur.fetchone() or {}).get("total", 0))

            data = {
                "role": role,
                "summary": {
                    "total_present": total_present,
                    "total_sales_submissions": total_sales_submissions,
                    "pending_content": pending_content,
                }
            }

        return mobile_api_response(
            ok=True,
            message="Dashboard berhasil diambil.",
            data=data
        )
    finally:
        cur.close()
        conn.close()       

@app.route("/api/mobile/products", methods=["GET"])
@mobile_api_login_required
def api_mobile_products():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, price, is_global, created_at
            FROM products
            ORDER BY id DESC;
        """)
        rows = cur.fetchall()

        products = []
        for r in rows:
            products.append({
                "id": r["id"],
                "name": r["name"],
                "price": int(r.get("price") or 0),
                "is_global": bool(r.get("is_global")),
                "created_at": str(r.get("created_at") or ""),
            })

        return mobile_api_response(
            ok=True,
            message="Daftar produk berhasil diambil.",
            data={"products": products}
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/attendance", methods=["GET"])
@mobile_api_login_required
def api_mobile_attendance():
    user = request.mobile_user
    user_id = user["id"]
    role = user["role"]

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if role == "admin":
            cur.execute("""
                SELECT
                    a.id,
                    a.user_id,
                    u.name AS user_name,
                    a.work_date,
                    a.status,
                    a.arrival_type,
                    a.note,
                    a.checkin_at
                FROM attendance a
                JOIN users u ON u.id = a.user_id
                ORDER BY a.work_date DESC, a.checkin_at DESC NULLS LAST
                LIMIT 100;
            """)
        else:
            cur.execute("""
                SELECT
                    a.id,
                    a.user_id,
                    %s AS user_name,
                    a.work_date,
                    a.status,
                    a.arrival_type,
                    a.note,
                    a.checkin_at
                FROM attendance a
                WHERE a.user_id=%s
                ORDER BY a.work_date DESC, a.checkin_at DESC NULLS LAST
                LIMIT 100;
            """, (user["name"], user_id))

        rows = cur.fetchall()

        attendance = []
        for r in rows:
            attendance.append({
                "id": r["id"],
                "user_id": r["user_id"],
                "user_name": r.get("user_name"),
                "work_date": str(r.get("work_date") or ""),
                "status": r.get("status") or "",
                "arrival_type": r.get("arrival_type") or "",
                "note": r.get("note") or "",
                "checkin_at": r["checkin_at"].astimezone(wib).strftime("%Y-%m-%d %H:%M:%S") if r.get("checkin_at") else "",
            })

        return mobile_api_response(
            ok=True,
            message="Data absensi berhasil diambil.",
            data={"attendance": attendance}
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/attendance/checkin", methods=["POST"])
@mobile_api_login_required
def api_mobile_attendance_checkin():
    user = request.mobile_user
    user_id = user["id"]

    data = request.get_json(silent=True) or {}
    arrival_type = (data.get("arrival_type") or "ONTIME").strip().upper()
    note = (data.get("note") or "").strip()

    allowed_arrival = {"ONTIME", "LATE", "SICK", "LEAVE", "ABSENT"}
    if arrival_type not in allowed_arrival:
        return mobile_api_response(
            ok=False,
            message="arrival_type tidak valid.",
            status_code=400
        )

    wib = pytz.timezone("Asia/Jakarta")
    now = datetime.now(pytz.timezone("Asia/Jakarta"))
    work_date = now.date()

    if arrival_type in ("SICK", "LEAVE", "ABSENT"):
        status = arrival_type
    else:
        status = "PRESENT"

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id
            FROM attendance
            WHERE user_id=%s AND work_date=%s
            LIMIT 1;
        """, (user_id, work_date))
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE attendance
                SET status=%s,
                    arrival_type=%s,
                    note=%s,
                    checkin_at=%s
                WHERE id=%s
                RETURNING id;
            """, (
                status,
                arrival_type,
                note,
                now,
                existing["id"]
            ))
            row = cur.fetchone()
            conn.commit()

            return mobile_api_response(
                ok=True,
                message="Absensi berhasil diperbarui.",
                data={
                    "attendance_id": row["id"],
                    "work_date": str(work_date),
                    "status": status,
                    "arrival_type": arrival_type,
                    "note": note,
                }
            )
        else:
            cur.execute("""
                INSERT INTO attendance
                    (user_id, work_date, status, arrival_type, note, created_at, checkin_at)
                VALUES
                    (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id;
            """, (
                user_id,
                work_date,
                status,
                arrival_type,
                note,
                now,
                now
            ))
            row = cur.fetchone()
            conn.commit()

            return mobile_api_response(
                ok=True,
                message="Absensi berhasil dikirim.",
                data={
                    "attendance_id": row["id"],
                    "work_date": str(work_date),
                    "status": status,
                    "arrival_type": arrival_type,
                    "note": note,
                }
            )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/notifications", methods=["GET"])
@mobile_api_login_required
def api_mobile_notifications():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT
                id,
                title,
                message,
                is_active,
                created_at
            FROM announcements
            WHERE is_active=TRUE
            ORDER BY created_at DESC
            LIMIT 50;
        """)
        rows = cur.fetchall()

        notifications = []
        for r in rows:
            notifications.append({
                "id": r["id"],
                "title": r.get("title") or "",
                "message": r.get("message") or "",
                "is_active": bool(r.get("is_active")),
                "created_at": str(r.get("created_at") or ""),
            })

        return mobile_api_response(
            ok=True,
            message="Notifikasi berhasil diambil.",
            data={"notifications": notifications}
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/sales", methods=["GET"])
@mobile_api_login_required
def api_mobile_sales():
    user = request.mobile_user
    user_id = user["id"]
    role = user["role"]

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        if role == "admin":
            cur.execute("""
                SELECT
                    s.id,
                    s.user_id,
                    u.name AS user_name,
                    s.product_id,
                    COALESCE(p.name, '-') AS product_name,
                    s.qty,
                    s.note,
                    s.status,
                    s.admin_note,
                    s.created_at
                FROM sales_submissions s
                JOIN users u ON u.id = s.user_id
                LEFT JOIN products p ON p.id = s.product_id
                ORDER BY s.id DESC
                LIMIT 100;
            """)
        else:
            cur.execute("""
                SELECT
                    s.id,
                    s.user_id,
                    %s AS user_name,
                    s.product_id,
                    COALESCE(p.name, '-') AS product_name,
                    s.qty,
                    s.note,
                    s.status,
                    s.admin_note,
                    s.created_at
                FROM sales_submissions s
                LEFT JOIN products p ON p.id = s.product_id
                WHERE s.user_id=%s
                ORDER BY s.id DESC
                LIMIT 100;
            """, (user["name"], user_id))

        rows = cur.fetchall()

        sales = []
        for r in rows:
            sales.append({
                "id": r["id"],
                "user_id": r["user_id"],
                "user_name": r.get("user_name"),
                "product_id": r.get("product_id"),
                "product_name": r.get("product_name") or "-",
                "qty": int(r.get("qty") or 0),
                "note": r.get("note") or "",
                "status": r.get("status") or "",
                "admin_note": r.get("admin_note") or "",
                "created_at": str(r.get("created_at") or ""),
            })

        return mobile_api_response(
            ok=True,
            message="Data sales berhasil diambil.",
            data={"sales": sales}
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/sales", methods=["POST"])
@mobile_api_login_required
def api_mobile_sales_submit():
    user = request.mobile_user
    user_id = user["id"]

    data = request.get_json(silent=True) or {}
    product_id = data.get("product_id")
    qty = data.get("qty")
    note = (data.get("note") or "").strip()

    try:
        product_id = int(product_id)
        qty = int(qty)
    except Exception:
        return mobile_api_response(
            ok=False,
            message="product_id dan qty harus angka.",
            status_code=400
        )

    if qty <= 0:
        return mobile_api_response(
            ok=False,
            message="Qty harus lebih dari 0.",
            status_code=400
        )

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name
            FROM products
            WHERE id=%s
            LIMIT 1;
        """, (product_id,))
        product = cur.fetchone()

        if not product:
            return mobile_api_response(
                ok=False,
                message="Produk tidak ditemukan.",
                status_code=404
            )

        cur.execute("""
            INSERT INTO sales_submissions
                (user_id, product_id, qty, note, status)
            VALUES
                (%s, %s, %s, %s, 'PENDING')
            RETURNING id;
        """, (
            user_id,
            product_id,
            qty,
            note
        ))
        row = cur.fetchone()
        conn.commit()

        return mobile_api_response(
            ok=True,
            message="Sales berhasil dikirim.",
            data={
                "sales_id": row["id"],
                "product_id": product_id,
                "product_name": product["name"],
                "qty": qty,
                "note": note,
                "status": "PENDING"
            }
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/profile", methods=["GET"])
@mobile_api_login_required
def api_mobile_profile():
    user = request.mobile_user

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, name, email, role,
                   COALESCE(points, 0) AS points,
                   COALESCE(points_admin, 0) AS points_admin,
                   created_at
            FROM users
            WHERE id=%s
            LIMIT 1;
        """, (user["id"],))
        row = cur.fetchone()

        if not row:
            return mobile_api_response(
                ok=False,
                message="User tidak ditemukan.",
                status_code=404
            )

        return mobile_api_response(
            ok=True,
            message="Profil berhasil diambil.",
            data={
                "profile": {
                    "id": row["id"],
                    "name": row["name"],
                    "email": row["email"],
                    "role": row["role"],
                    "points": int(row.get("points") or 0),
                    "points_admin": int(row.get("points_admin") or 0),
                    "created_at": str(row.get("created_at") or ""),
                }
            }
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/profile", methods=["PUT"])
@mobile_api_login_required
def api_mobile_profile_update():
    user = request.mobile_user
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()

    if not name or not email:
        return mobile_api_response(
            ok=False,
            message="Nama dan email wajib diisi.",
            status_code=400
        )

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id
            FROM users
            WHERE email=%s AND id<>%s
            LIMIT 1;
        """, (email, user["id"]))
        existing = cur.fetchone()

        if existing:
            return mobile_api_response(
                ok=False,
                message="Email sudah dipakai user lain.",
                status_code=400
            )

        cur.execute("""
            UPDATE users
            SET name=%s, email=%s
            WHERE id=%s
            RETURNING id, name, email, role;
        """, (name, email, user["id"]))
        row = cur.fetchone()
        conn.commit()

        return mobile_api_response(
            ok=True,
            message="Profil berhasil diperbarui.",
            data={
                "profile": {
                    "id": row["id"],
                    "name": row["name"],
                    "email": row["email"],
                    "role": row["role"],
                }
            }
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/change-password", methods=["POST"])
@mobile_api_login_required
def api_mobile_change_password():
    user = request.mobile_user
    data = request.get_json(silent=True) or {}

    old_password = (data.get("old_password") or "").strip()
    new_password = (data.get("new_password") or "").strip()

    if not old_password or not new_password:
        return mobile_api_response(
            ok=False,
            message="Password lama dan password baru wajib diisi.",
            status_code=400
        )

    if len(new_password) < 6:
        return mobile_api_response(
            ok=False,
            message="Password baru minimal 6 karakter.",
            status_code=400
        )

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT id, password_hash
            FROM users
            WHERE id=%s
            LIMIT 1;
        """, (user["id"],))
        row = cur.fetchone()

        if not row:
            return mobile_api_response(
                ok=False,
                message="User tidak ditemukan.",
                status_code=404
            )

        if not check_password_hash(row["password_hash"], old_password):
            return mobile_api_response(
                ok=False,
                message="Password lama salah.",
                status_code=400
            )

        new_hash = generate_password_hash(new_password)

        cur.execute("""
            UPDATE users
            SET password_hash=%s
            WHERE id=%s;
        """, (new_hash, user["id"]))
        conn.commit()

        return mobile_api_response(
            ok=True,
            message="Password berhasil diganti."
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/logout-all", methods=["POST"])
@mobile_api_login_required
def api_mobile_logout_all():
    user = request.mobile_user

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE mobile_api_tokens
            SET is_active=FALSE
            WHERE user_id=%s;
        """, (user["id"],))
        conn.commit()

        return mobile_api_response(
            ok=True,
            message="Semua device berhasil logout."
        )
    finally:
        cur.close()
        conn.close()

@app.route("/api/mobile/health", methods=["GET"])
def api_mobile_health():
    return mobile_api_response(
        ok=True,
        message="Mobile API aktif.",
        data={
            "app_name": "UMGAP Mobile API",
            "server_time": str(datetime.now()),
            "version": "1.0.0"
        }
    )


@app.route("/api/mobile/info", methods=["GET"])
def api_mobile_info():
    return mobile_api_response(
        ok=True,
        message="Daftar endpoint mobile.",
        data={
            "base_prefix": "/api/mobile",
            "endpoints": [
                {"method": "GET", "path": "/api/mobile/health"},
                {"method": "POST", "path": "/api/mobile/login"},
                {"method": "POST", "path": "/api/mobile/logout"},
                {"method": "GET", "path": "/api/mobile/me"},
                {"method": "GET", "path": "/api/mobile/dashboard"},
                {"method": "GET", "path": "/api/mobile/products"},
                {"method": "GET", "path": "/api/mobile/attendance"},
                {"method": "POST", "path": "/api/mobile/attendance/checkin"},
                {"method": "GET", "path": "/api/mobile/notifications"},
                {"method": "GET", "path": "/api/mobile/sales"},
                {"method": "POST", "path": "/api/mobile/sales"},
                {"method": "GET", "path": "/api/mobile/profile"},
                {"method": "PUT", "path": "/api/mobile/profile"},
                {"method": "POST", "path": "/api/mobile/change-password"},
                {"method": "POST", "path": "/api/mobile/logout-all"}
            ]
        }
    )

@app.route('/api/mobile/register', methods=['POST'])
def api_mobile_register():
    try:
        data = request.get_json(silent=True) or {}

        first_name = (data.get('first_name') or '').strip()
        last_name = (data.get('last_name') or '').strip()
        email = (data.get('email') or '').strip().lower()
        password = data.get('password') or ''
        year = (data.get('year') or '').strip()
        month = (data.get('month') or '').strip()
        day = (data.get('day') or '').strip()

        if not first_name:
            return jsonify({'ok': False, 'message': 'First name wajib diisi.'}), 400
        if not last_name:
            return jsonify({'ok': False, 'message': 'Last name wajib diisi.'}), 400
        if not email:
            return jsonify({'ok': False, 'message': 'Email wajib diisi.'}), 400
        if not password:
            return jsonify({'ok': False, 'message': 'Password wajib diisi.'}), 400

        full_name = f"{first_name} {last_name}".strip()

        conn = get_conn()
        cur = conn.cursor()

        cur.execute("SELECT id FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1", (email,))
        existing = cur.fetchone()
        if existing:
            cur.close()
            conn.close()
            return jsonify({'ok': False, 'message': 'Email sudah terdaftar.'}), 409

        birthday = None
        if year and month and day:
            try:
                birthday = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
            except Exception:
                birthday = None

        password_hash = generate_password_hash(password)

        # Aman kalau kolom birthday belum ada: insert tanpa birthday
        try:
            cur.execute("""
                INSERT INTO users (name, email, password_hash, role, points, points_admin, birthday)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id, name, email, role
            """, (full_name, email, password_hash, 'employee', 0, 0, birthday))
        except Exception:
            conn.rollback()
            cur.execute("""
                INSERT INTO users (name, email, password_hash, role, points, points_admin)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id, name, email, role
            """, (full_name, email, password_hash, 'employee', 0, 0))

        user = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Register berhasil.',
            'data': {
                'user': {
                    'id': user[0],
                    'name': user[1],
                    'email': user[2],
                    'role': user[3]
                }
            }
        }), 201

    except Exception as e:
        return jsonify({'ok': False, 'message': f'Register gagal: {str(e)}'}), 500


@app.route('/api/mobile/forgot-password', methods=['POST'])
def api_mobile_forgot_password():
    try:
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip().lower()

        if not email:
            return jsonify({'ok': False, 'message': 'Email wajib diisi.'}), 400

        conn = get_conn()
        cur = conn.cursor()

        cur.execute("SELECT id, email FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1", (email,))
        user = cur.fetchone()

        if not user:
            cur.close()
            conn.close()
            return jsonify({'ok': False, 'message': 'Email tidak ditemukan.'}), 404

        otp_code = str(random.randint(100000, 999999))
        expired_at = datetime.now() + timedelta(minutes=10)

        cur.execute("""
            INSERT INTO password_reset_otps (email, otp_code, expired_at, used)
            VALUES (%s, %s, %s, FALSE)
        """, (email, otp_code, expired_at))

        conn.commit()
        cur.close()
        conn.close()

        # Untuk sementara OTP dikembalikan ke response
        # Nanti bisa diganti kirim email/WhatsApp
        return jsonify({
            'ok': True,
            'message': 'OTP reset password berhasil dibuat.',
            'data': {
                'email': email,
                'otp_code': otp_code,
                'expired_in_minutes': 10
            }
        }), 200

    except Exception as e:
        return jsonify({'ok': False, 'message': f'Forgot password gagal: {str(e)}'}), 500


@app.route('/api/mobile/reset-password', methods=['POST'])
def api_mobile_reset_password():
    try:
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip().lower()
        otp_code = (data.get('otp_code') or '').strip()
        new_password = data.get('new_password') or ''

        if not email:
            return jsonify({'ok': False, 'message': 'Email wajib diisi.'}), 400
        if not otp_code:
            return jsonify({'ok': False, 'message': 'OTP wajib diisi.'}), 400
        if not new_password:
            return jsonify({'ok': False, 'message': 'Password baru wajib diisi.'}), 400

        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, expired_at, used
            FROM password_reset_otps
            WHERE LOWER(email)=LOWER(%s) AND otp_code=%s
            ORDER BY id DESC
            LIMIT 1
        """, (email, otp_code))
        otp_row = cur.fetchone()

        if not otp_row:
            cur.close()
            conn.close()
            return jsonify({'ok': False, 'message': 'OTP tidak valid.'}), 400

        otp_id, expired_at, used = otp_row

        if used:
            cur.close()
            conn.close()
            return jsonify({'ok': False, 'message': 'OTP sudah digunakan.'}), 400

        if expired_at < datetime.now():
            cur.close()
            conn.close()
            return jsonify({'ok': False, 'message': 'OTP sudah kadaluarsa.'}), 400

        password_hash = generate_password_hash(new_password)

        cur.execute("""
            UPDATE users
            SET password_hash=%s
            WHERE LOWER(email)=LOWER(%s)
        """, (password_hash, email))

        cur.execute("""
            UPDATE password_reset_otps
            SET used=TRUE
            WHERE id=%s
        """, (otp_id,))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Password berhasil direset.'
        }), 200

    except Exception as e:
        return jsonify({'ok': False, 'message': f'Reset password gagal: {str(e)}'}), 500

@app.route("/api/mobile/<path:anything>", methods=["OPTIONS"])
def api_mobile_options(anything):
    return ("", 204)

# ── Force init endpoint (admin only, satu kali pakai) ──
@app.route("/init-buy-prices")
def init_buy_prices_route():
    """Force buat tabel buy_prices dan seed data. Akses sekali saat deploy."""
    deny = admin_required()
    if deny:
        return deny
    try:
        ensure_buy_prices_schema()
        conn2 = get_conn()
        cur2  = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM buy_prices;")
        n = cur2.fetchone()[0]
        cur2.close(); conn2.close()
        return f"OK: buy_prices siap. Total rows: {n}"
    except Exception as e:
        return f"ERROR: {e}", 500

@app.route("/api-tester")
def api_tester():
    return """
    <h2>API Tester - Mobile Login</h2>

    <input id="email" placeholder="Email"><br><br>
    <input id="password" placeholder="Password" type="password"><br><br>

    <button onclick="login()">Login</button>

    <pre id="result"></pre>

    <script>
    async function login() {
        const email = document.getElementById('email').value;
        const password = document.getElementById('password').value;

        const res = await fetch('/api/mobile/login', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json'
            },
            body: JSON.stringify({
                email: email,
                password: password,
                device_name: "Web Tester"
            })
        });

        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }
    </script>
    """

@app.route("/api-tester-2")
def api_tester_2():
    return """
    <h2>API Tester Tahap 2</h2>

    <label>Token Bearer</label><br>
    <input id="token" style="width:500px" placeholder="Paste token dari login"><br><br>

    <button onclick="getProducts()">GET Products</button>
    <button onclick="getAttendance()">GET Attendance</button>
    <button onclick="checkin()">POST Checkin</button>
    <button onclick="getNotifications()">GET Notifications</button>

    <pre id="result" style="margin-top:20px; background:#111; color:#0f0; padding:15px;"></pre>

    <script>
    function authHeaders() {
        return {
            'Content-Type': 'application/json',
            'Authorization': 'Bearer ' + document.getElementById('token').value
        };
    }

    async function getProducts() {
        const res = await fetch('/api/mobile/products', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function getAttendance() {
        const res = await fetch('/api/mobile/attendance', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function checkin() {
        const res = await fetch('/api/mobile/attendance/checkin', {
            method: 'POST',
            headers: authHeaders(),
            body: JSON.stringify({
                arrival_type: "ONTIME",
                note: "Checkin dari tester API"
            })
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function getNotifications() {
        const res = await fetch('/api/mobile/notifications', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }
    </script>
    """

@app.route("/api-tester-3")
def api_tester_3():
    return """
    <h2>API Tester Tahap 3</h2>

    <label>Token Bearer</label><br>
    <input id="token" style="width:600px" placeholder="Paste token dari login"><br><br>

    <button onclick="getSales()">GET Sales</button>
    <button onclick="submitSales()">POST Sales</button>
    <button onclick="getProfile()">GET Profile</button>
    <button onclick="updateProfile()">PUT Profile</button>
    <button onclick="changePassword()">POST Change Password</button>
    <button onclick="logoutAll()">POST Logout All</button>

    <pre id="result" style="margin-top:20px; background:#111; color:#0f0; padding:15px;"></pre>

    <script>
    function authHeaders() {
        return {
            'Content-Type': 'application/json',
            'Authorization': 'Bearer ' + document.getElementById('token').value
        };
    }

    async function getSales() {
        const res = await fetch('/api/mobile/sales', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function submitSales() {
        const productId = prompt("Masukkan product_id:", "1");
        const qty = prompt("Masukkan qty:", "1");
        const note = prompt("Masukkan note:", "Input dari API tester");

        const res = await fetch('/api/mobile/sales', {
            method: 'POST',
            headers: authHeaders(),
            body: JSON.stringify({
                product_id: productId,
                qty: qty,
                note: note
            })
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function getProfile() {
        const res = await fetch('/api/mobile/profile', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function updateProfile() {
        const name = prompt("Masukkan nama baru:");
        const email = prompt("Masukkan email baru:");

        const res = await fetch('/api/mobile/profile', {
            method: 'PUT',
            headers: authHeaders(),
            body: JSON.stringify({
                name: name,
                email: email
            })
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function changePassword() {
        const oldPassword = prompt("Masukkan password lama:");
        const newPassword = prompt("Masukkan password baru:");

        const res = await fetch('/api/mobile/change-password', {
            method: 'POST',
            headers: authHeaders(),
            body: JSON.stringify({
                old_password: oldPassword,
                new_password: newPassword
            })
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function logoutAll() {
        const res = await fetch('/api/mobile/logout-all', {
            method: 'POST',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }
    </script>
    """

@app.route("/api-tester-4")
def api_tester_4():
    return """
    <h2>API Tester Tahap 4</h2>

    <label>Base URL</label><br>
    <input id="base_url" style="width:600px" value="http://127.0.0.1:5000"><br><br>

    <label>Token Bearer</label><br>
    <input id="token" style="width:600px" placeholder="Paste token jika perlu"><br><br>

    <button onclick="health()">GET Health</button>
    <button onclick="info()">GET Info</button>
    <button onclick="me()">GET Me</button>

    <pre id="result" style="margin-top:20px; background:#111; color:#0f0; padding:15px;"></pre>

    <script>
    function getBaseUrl() {
        return document.getElementById('base_url').value.replace(/\\/$/, '');
    }

    function authHeaders() {
        const token = document.getElementById('token').value.trim();
        const headers = { 'Content-Type': 'application/json' };
        if (token) {
            headers['Authorization'] = 'Bearer ' + token;
        }
        return headers;
    }

    async function health() {
        const res = await fetch(getBaseUrl() + '/api/mobile/health', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function info() {
        const res = await fetch(getBaseUrl() + '/api/mobile/info', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }

    async function me() {
        const res = await fetch(getBaseUrl() + '/api/mobile/me', {
            method: 'GET',
            headers: authHeaders()
        });
        const data = await res.json();
        document.getElementById('result').innerText = JSON.stringify(data, null, 2);
    }
    </script>
    """

@app.route("/test-cctv")
def test_cctv():
    return "OK CCTV ROUTE"


# ==================== RUN ====================
if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)

# Init DB on startup
def safe_init_db():
    try:
        if os.getenv('RUN_SCHEMA_ON_REQUEST','').lower()=='true':
            ensure_points_schema()
        ensure_hr_v2_schema()
        init_points_v1()
        ensure_announcements_schema()
        ensure_buy_prices_schema()   # ← TAMBAH INI
        print("DB init OK — buy_prices schema ready")
    except Exception as e:
        print("Init error:", e)

# Jangan auto-init saat import di gunicorn.
# Jalankan hanya kalau kamu set env INIT_DB_ON_STARTUP=true
if os.getenv("INIT_DB_ON_STARTUP", "").lower() == "true":
    safe_init_db()
